"""사정률 예측 로직 Tab2/Tab3 인터랙션 E2E 테스트

검증 대상:
1. Tab1 — /analysis/rate-buckets/{id} 3가지 모드 응답 정합성
2. Tab2 — /analysis/company-rates/{id} 검색 5종 파라미터 적용
3. Tab3 — /analysis/correlation/{id} 3가지 방법 + 종합 1순위
4. 학습 저장 — GET/PUT /users/me/prediction-settings 영속화
5. 엑셀 export — /analysis/export/{type} (buckets/correlation/bid_list) 4종

실행:
    cd backend && python3 -m pytest tests/test_prediction_e2e.py -v
"""
from __future__ import annotations

import io
import os
import sys
import uuid
from datetime import datetime, timedelta

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from server import (  # noqa: E402
    app, SessionLocal, User, get_password_hash,
    BidAnnouncement, BidResult,
)


# ── Fixtures ──────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    return TestClient(app)


@pytest.fixture(scope="module")
def auth_token(client):
    """테스트용 admin 토큰"""
    email = "prediction_e2e@example.com"
    password = "testpass123"
    db = SessionLocal()
    u = db.query(User).filter(User.email == email).first()
    if not u:
        u = User(
            username=email, email=email,
            hashed_password=get_password_hash(password),
            name="Prediction E2E", role="admin", is_active=True,
            joined_at=datetime.now(),
        )
        db.add(u); db.commit()
    db.close()
    resp = client.post(f"/api/v1/auth/login?username={email}&password={password}")
    assert resp.status_code == 200, resp.text
    return resp.json()["access_token"]


def _h(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="module")
def fixture_announcement(client, auth_token):
    """테스트용 공고 + BidResult 5건 직접 적재 — API 의존 없이 분석 가능 데이터 확보"""
    db = SessionLocal()
    bid_no = f"E2E-PRED-{uuid.uuid4().hex[:8]}"
    ann = BidAnnouncement(
        id=str(uuid.uuid4()),
        source="TEST", bid_number=bid_no, bid_ord="000",
        title="E2E 예측 로직 테스트 공고",
        ordering_org_name="테스트청",
        category="용역",
        base_amount=1_000_000_000,
        announced_at=datetime.now() - timedelta(days=10),
        status="진행중",
    )
    db.add(ann); db.commit()
    # 사정률 분포 5건 (98.5 ~ 101.5 범위)
    for i, rate in enumerate([99.5, 99.8, 100.0, 100.0, 100.3]):
        r = BidResult(
            id=str(uuid.uuid4()),
            announcement_id=ann.id,
            assessment_rate=rate,
            first_place_rate=rate - 0.05,
            first_place_amount=int(1_000_000_000 * (rate - 0.05) / 100),
            winning_amount=int(1_000_000_000 * rate / 100),
            opened_at=datetime.now() - timedelta(days=5 - i),
        )
        db.add(r)
    db.commit()
    ann_id = ann.id
    db.close()
    yield ann_id
    # cleanup
    db = SessionLocal()
    db.query(BidResult).filter(BidResult.announcement_id == ann_id).delete()
    db.query(BidAnnouncement).filter(BidAnnouncement.id == ann_id).delete()
    db.commit(); db.close()


# ── Tab1: rate-buckets ────────────────────────────────────────────────

def test_tab1_rate_buckets_returns_3_modes(client, auth_token, fixture_announcement):
    resp = client.get(
        f"/api/v1/analysis/rate-buckets/{fixture_announcement}"
        "?period_months=12&category_filter=service&detail_rule=max_gap",
        headers=_h(auth_token),
    )
    assert resp.status_code == 200, resp.text
    d = resp.json()
    # 3가지 모드 모두 키 존재
    assert set(d["buckets"].keys()) == {"A", "B", "C"}
    # 데이터가 있을 때 모드 A 는 빈도 최대 1순위 보장
    if d["data_count"] > 0 and d["buckets"]["A"]:
        assert d["buckets"]["A"][0]["rank"] == 1
        assert d["buckets"]["A"][0]["mode"] == "A"


def test_tab1_detail_rules(client, auth_token, fixture_announcement):
    """3가지 detail_rule 모두 정상 응답"""
    for rule in ["first_after", "last_after", "max_gap"]:
        resp = client.get(
            f"/api/v1/analysis/rate-buckets/{fixture_announcement}?detail_rule={rule}",
            headers=_h(auth_token),
        )
        assert resp.status_code == 200, f"{rule}: {resp.text}"
        d = resp.json()
        assert d["params"]["detail_rule"] == rule


# ── Tab2: company-rates 검색 5종 ──────────────────────────────────────

def test_tab2_company_rates_basic(client, auth_token, fixture_announcement):
    resp = client.get(
        f"/api/v1/analysis/company-rates/{fixture_announcement}"
        "?rate_range_start=99&rate_range_end=101&period_months=12",
        headers=_h(auth_token),
    )
    assert resp.status_code == 200, resp.text


def test_tab2_search_params_accepted(client, auth_token, fixture_announcement):
    """5종 검색 파라미터 모두 적용 시 정상 응답 + 0건 가능"""
    resp = client.get(
        f"/api/v1/analysis/company-rates/{fixture_announcement}"
        "?rate_range_start=99&rate_range_end=101"
        "&org_search=테스트"
        "&price_volatility=2.0"
        "&base_amount_min=100000"
        "&base_amount_max=10000000000"
        "&industry_filter=421",
        headers=_h(auth_token),
    )
    assert resp.status_code == 200, resp.text


# ── Tab3: correlation ─────────────────────────────────────────────────

def test_tab3_correlation_returns_3_methods(client, auth_token, fixture_announcement):
    resp = client.get(
        f"/api/v1/analysis/correlation/{fixture_announcement}"
        "?period_months=12&category_filter=service&bucket_mode=A"
        "&rate_range_start=99&rate_range_end=101",
        headers=_h(auth_token),
    )
    assert resp.status_code == 200, resp.text
    d = resp.json()
    assert len(d["methods"]) == 3
    for i, expected in enumerate(["발생빈도", "갭 분석", "결합"]):
        assert expected in d["methods"][i]["name"]
    assert "correlation" in d
    assert "agreement" in d["correlation"]


def test_tab3_correlation_modes(client, auth_token, fixture_announcement):
    """bucket_mode A/B/C 모두 응답"""
    for mode in ["A", "B", "C"]:
        resp = client.get(
            f"/api/v1/analysis/correlation/{fixture_announcement}?bucket_mode={mode}",
            headers=_h(auth_token),
        )
        assert resp.status_code == 200
        assert resp.json()["params"]["bucket_mode"] == mode


# ── 학습 저장 (PredictionSettings) ────────────────────────────────────

def test_settings_initial_get_returns_defaults(client, auth_token):
    resp = client.get("/api/v1/users/me/prediction-settings", headers=_h(auth_token))
    assert resp.status_code == 200
    d = resp.json()
    # 처음 호출 시 exists=False 또는 존재
    assert "period_months" in d
    assert "bucket_mode" in d


def test_settings_put_then_get_roundtrip(client, auth_token):
    """PUT 으로 설정 저장 → GET 으로 동일 값 회수"""
    payload = {
        "period_months": 9,
        "category_filter": "construction",
        "bucket_mode": "B",
        "detail_rule": "first_after",
        "rate_range_start": 98.5,
        "rate_range_end": 101.5,
        "confirmed_rate": 99.7,
    }
    r1 = client.put("/api/v1/users/me/prediction-settings",
                    json=payload, headers=_h(auth_token))
    assert r1.status_code == 200, r1.text
    assert r1.json()["status"] == "ok"

    r2 = client.get("/api/v1/users/me/prediction-settings", headers=_h(auth_token))
    assert r2.status_code == 200
    d = r2.json()
    assert d["period_months"] == 9
    assert d["category_filter"] == "construction"
    assert d["bucket_mode"] == "B"
    assert d["detail_rule"] == "first_after"
    assert d["rate_range_start"] == 98.5
    assert d["rate_range_end"] == 101.5
    assert d["confirmed_rate"] == 99.7
    assert d["exists"] is True


def test_settings_put_partial_update(client, auth_token):
    """일부 필드만 보내도 다른 필드 보존"""
    # 먼저 알려진 상태 만들기
    full = {"period_months": 6, "bucket_mode": "A", "detail_rule": "max_gap"}
    client.put("/api/v1/users/me/prediction-settings", json=full, headers=_h(auth_token))

    # period_months 만 변경
    client.put("/api/v1/users/me/prediction-settings",
               json={"period_months": 24}, headers=_h(auth_token))
    d = client.get("/api/v1/users/me/prediction-settings",
                   headers=_h(auth_token)).json()
    assert d["period_months"] == 24
    assert d["bucket_mode"] == "A"   # 보존 확인


# ── Excel Export (4종) ────────────────────────────────────────────────

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _common_export_qs(ann_id: str) -> str:
    return (
        f"announcement_id={ann_id}"
        "&period_months=12&category_filter=service"
        "&bucket_mode=A&detail_rule=max_gap"
        "&rate_range_start=99&rate_range_end=101"
    )


@pytest.mark.parametrize("export_type", ["buckets", "company", "correlation", "bid_list"])
def test_export_xlsx_returns_file(client, auth_token, fixture_announcement, export_type):
    resp = client.get(
        f"/api/v1/analysis/export/{export_type}?{_common_export_qs(fixture_announcement)}",
        headers=_h(auth_token),
    )
    assert resp.status_code == 200, f"{export_type}: {resp.text[:200]}"
    assert resp.headers["content-type"] == XLSX_MIME
    # 다운로드 헤더 + 파일명 패턴
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd and ".xlsx" in cd
    # 빈 파일이 아닌지
    assert len(resp.content) > 1000  # 최소 1KB


def test_export_invalid_type_returns_400(client, auth_token, fixture_announcement):
    resp = client.get(
        f"/api/v1/analysis/export/INVALID?{_common_export_qs(fixture_announcement)}",
        headers=_h(auth_token),
    )
    assert resp.status_code == 400


def test_export_xlsx_contains_announcement_meta(client, auth_token, fixture_announcement):
    """xlsx 파일 안에 공고명/기초금액이 들어있는지 (openpyxl 로 재로드)"""
    from openpyxl import load_workbook
    resp = client.get(
        f"/api/v1/analysis/export/buckets?{_common_export_qs(fixture_announcement)}",
        headers=_h(auth_token),
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # 공고 메타 6번째 줄 위쪽
    cells = [ws.cell(row=r, column=2).value for r in range(1, 5)]
    assert any("E2E 예측 로직 테스트 공고" in str(c) for c in cells)
    assert 1_000_000_000 in cells  # 기초금액
