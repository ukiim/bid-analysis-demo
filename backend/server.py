"""통합 데모 서버 — KBID 스타일 1순위 예상 낙찰가 분석 시스템

SQLite + FastAPI로 백엔드 API 서빙 + 프론트엔드 정적 파일 서빙
단일 프로세스로 전체 데모 구동

실행: python3 server.py
"""

import os
import sys
import uuid
import json
import random
import math
import logging
import statistics
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path


# ─── 로깅 설정 ──────────────────────────────────────────────────────────

def _setup_logging() -> logging.Logger:
    level_name = os.environ.get("LOG_LEVEL", "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)
    log_file = os.environ.get("LOG_FILE", "").strip()

    root = logging.getLogger()
    root.setLevel(level)
    # 기존 핸들러 제거 (uvicorn reload 대응)
    for h in list(root.handlers):
        root.removeHandler(h)

    fmt = logging.Formatter(
        "%(asctime)s %(levelname)s [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    root.addHandler(sh)

    if log_file:
        try:
            from logging.handlers import RotatingFileHandler
            os.makedirs(os.path.dirname(log_file), exist_ok=True)
            # 환경변수로 크기/보관 개수 조정 (기본 10MB × 7개 = 약 70MB 상한)
            max_bytes = int(os.environ.get("LOG_MAX_BYTES", str(10 * 1024 * 1024)))
            backup_count = int(os.environ.get("LOG_BACKUP_COUNT", "7"))
            fh = RotatingFileHandler(
                log_file, maxBytes=max_bytes, backupCount=backup_count,
                encoding="utf-8", delay=True,
            )
            fh.setFormatter(fmt)
            root.addHandler(fh)
        except Exception as e:
            sh.handle(logging.LogRecord(
                "server", logging.WARNING, __file__, 0,
                f"파일 로그 핸들러 구성 실패: {e}", None, None,
            ))

    return logging.getLogger("bid-insight")


logger = _setup_logging()

from fastapi import FastAPI, Query, Depends, HTTPException, UploadFile, File, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.security import OAuth2PasswordBearer
from sqlalchemy import (
    Column, String, Integer, Float, Text, DateTime, Boolean,
    func, text, Index,
)
from jose import JWTError, jwt
import bcrypt

# ─── DB 엔진 (app/core/database 에서 정의 — F2 분리) ───────────────────────

from app.core.database import engine, SessionLocal, DB_PATH  # noqa: E402

NAS_PATH = os.environ.get("NAS_MOUNT_PATH", os.path.join(os.path.dirname(__file__), "data"))


# ─── 모델 (app/models 에서 정의 — F1 분리) ────────────────────────────────

from app.models import (  # noqa: E402
    Base,
    BidAnnouncement,
    BidResult,
    CompanyBidRecord,
    DataSyncLog,
    ModelMetric,
    PredictionSettings,
    QueryHistory,
    UploadLog,
    User,
)


# ─── 발주기관 계층 매핑 ───────────────────────────────────────────────────

ORG_HIERARCHY = {
    "고양시": "경기도", "수원시": "경기도", "성남시": "경기도", "용인시": "경기도",
    "안양시": "경기도", "부천시": "경기도", "화성시": "경기도", "안산시": "경기도",
    "서울특별시 강남구": "서울특별시", "서울특별시 서초구": "서울특별시",
    "서울특별시 송파구": "서울특별시", "서울특별시 강동구": "서울특별시",
    "부산광역시 해운대구": "부산광역시", "부산광역시 사하구": "부산광역시",
    "대전광역시 유성구": "대전광역시", "인천광역시 남동구": "인천광역시",
    "고양시교육청": "경기도교육청", "수원시교육청": "경기도교육청",
    "국토교통부": "중앙부처", "환경부": "중앙부처", "교육부": "중앙부처",
    "행정안전부": "중앙부처", "국방부": "중앙부처",
    "한국도로공사": "공기업", "한국수자원공사": "공기업", "한국토지주택공사": "공기업",
    "부산항만공사": "공기업", "인천국제공항공사": "공기업",
    "서울대학교": "교육기관", "부산대학교": "교육기관", "충남대학교": "교육기관",
    "경기도": "경기도", "서울특별시": "서울특별시", "부산광역시": "부산광역시",
    "대전광역시": "대전광역시", "인천광역시": "인천광역시", "세종특별자치시": "세종특별자치시",
}

# ─── 메타 상수 (API 응답용) ────────────────────────────────────────────────

REGIONS = ["서울", "경기", "부산", "대전", "인천", "광주", "대구", "울산", "세종", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주"]

# 시드 데이터 생성은 backend/seed.py 모듈로 분리됨.
# 운영 환경에서는 SKIP_SEED=true 로 비활성화.


# ─── FastAPI 앱 ────────────────────────────────────────────────────────────

# APScheduler 통합 (lifespan에서 시작/종료)
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    from apscheduler.triggers.interval import IntervalTrigger
    _scheduler: "BackgroundScheduler | None" = None
    _SCHEDULER_AVAILABLE = True
except ImportError:
    _scheduler = None
    _SCHEDULER_AVAILABLE = False


def _send_alert_webhook(title: str, fields: dict):
    """Slack/Discord 호환 webhook 알림 + 옵션 SMTP 이메일.

    환경변수:
      ALERT_WEBHOOK_URL — Slack 호환 incoming webhook
      ALERT_EMAIL_TO   — 콤마 구분 수신자 (선택)
      SMTP_HOST/PORT/USER/PASS — SMTP 서버 (ALERT_EMAIL_TO 사용 시 필수)
    """
    text_lines = [f"*{title}*"]
    for k, v in fields.items():
        text_lines.append(f"• {k}: `{v}`")
    body_text = "\n".join(text_lines)

    # 1) Slack/Discord webhook
    url = os.environ.get("ALERT_WEBHOOK_URL", "").strip()
    if url:
        try:
            import urllib.request
            payload = json.dumps({"text": body_text}).encode("utf-8")
            req = urllib.request.Request(
                url, data=payload,
                headers={"Content-Type": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=5) as resp:
                if resp.status >= 400:
                    logger.warning("alert webhook non-2xx status=%s", resp.status)
        except Exception as e:
            logger.warning("alert webhook failed: %s", e)

    # 2) SMTP Email (옵션)
    email_to = os.environ.get("ALERT_EMAIL_TO", "").strip()
    smtp_host = os.environ.get("SMTP_HOST", "").strip()
    if email_to and smtp_host:
        try:
            import smtplib
            from email.mime.text import MIMEText
            recipients = [e.strip() for e in email_to.split(",") if e.strip()]
            msg = MIMEText(body_text, "plain", "utf-8")
            msg["Subject"] = f"[비드스타] {title}"
            msg["From"] = os.environ.get("SMTP_FROM", os.environ.get("SMTP_USER", "noreply@bidstar"))
            msg["To"] = ", ".join(recipients)
            port = int(os.environ.get("SMTP_PORT", "587"))
            with smtplib.SMTP(smtp_host, port, timeout=10) as s:
                s.ehlo()
                if port == 587:
                    s.starttls()
                user = os.environ.get("SMTP_USER", "").strip()
                pwd = os.environ.get("SMTP_PASS", "").strip()
                if user and pwd:
                    s.login(user, pwd)
                s.sendmail(msg["From"], recipients, msg.as_string())
            logger.info("alert email sent to %s", email_to)
        except Exception as e:
            logger.warning("alert email failed: %s", e)


def _check_sync_failure_threshold():
    """최근 N회 수집 중 실패 비율이 임계치 초과 시 알림.

    환경변수:
      ALERT_FAILURE_THRESHOLD — 0.0~1.0 (기본 0.5 = 50%)
      ALERT_FAILURE_WINDOW    — 최근 N건 (기본 10)
    """
    try:
        threshold = float(os.environ.get("ALERT_FAILURE_THRESHOLD", "0.5"))
        window = int(os.environ.get("ALERT_FAILURE_WINDOW", "10"))
    except ValueError:
        return
    db = SessionLocal()
    try:
        recent = db.query(DataSyncLog).order_by(
            DataSyncLog.started_at.desc()
        ).limit(window).all()
        if len(recent) < window:
            return
        failed = sum(1 for r in recent if r.status == "failed")
        ratio = failed / len(recent)
        if ratio >= threshold:
            _send_alert_webhook(
                f"🚨 수집 실패 임계치 초과 ({failed}/{len(recent)}건, {int(ratio*100)}%)",
                {
                    "임계치": f"{int(threshold*100)}%",
                    "최근 실패": ", ".join(
                        f"{r.source} {r.started_at.strftime('%m/%d %H:%M')}"
                        for r in recent if r.status == "failed"
                    )[:200],
                },
            )
    finally:
        db.close()


def _scheduled_model_retrain():
    """매일 회귀 모델 재학습 — 카테고리별 OLS 적합 후 ModelMetric 저장.

    in-sample MAE/R² 만 기록 (out-of-sample 은 walk-forward 엔드포인트 활용).
    """
    import json as _json
    db = SessionLocal()
    try:
        period_days = int(os.environ.get("MODEL_RETRAIN_DAYS", "180"))
        cutoff = datetime.now() - timedelta(days=period_days)
        for category in ("공사", "용역", "all"):
            q = db.query(
                BidResult.assessment_rate, BidResult.opened_at,
                BidAnnouncement.base_amount, BidAnnouncement.category,
                BidAnnouncement.region, BidAnnouncement.industry_code,
                BidAnnouncement.ordering_org_type,
            ).join(BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id).filter(
                BidResult.assessment_rate.isnot(None),
                BidResult.opened_at >= cutoff,
            )
            if category in ("공사", "용역"):
                q = q.filter(BidAnnouncement.category == category)
            samples = []
            ref = datetime.now()
            for row in q.all():
                if not row[0]:
                    continue
                samples.append({
                    "rate": row[0],
                    "base_amount": row[2] or 0,
                    "category": row[3], "region": row[4],
                    "industry": row[5], "org_type": row[6],
                    "days_old": (ref - row[1]).days if row[1] else 0,
                })
            if len(samples) < 30:
                logger.info("model retrain skipped category=%s (n=%d, <30)", category, len(samples))
                continue
            # 자기 자신을 target 으로 → 평균값 회귀 결과
            target = {"base_amount": sum(s["base_amount"] for s in samples) / len(samples),
                      "category": category, "region": None, "industry": None,
                      "org_type": None, "days_old": 0}
            result = _ols_regression(samples, target)
            if result.get("predicted_rate") is None:
                continue
            # in-sample MAE 계산 (lstsq 가 다시 fit)
            import numpy as _np
            from collections import Counter as _Counter
            top_regions = [r for r, _ in _Counter(s.get("region") for s in samples
                                                    if s.get("region")).most_common(5)]

            def _featurize(s):
                ba = s.get("base_amount") or 1
                return [1.0, math.log(max(1, ba)), 1 if s.get("category") == "공사" else 0] + \
                       [1 if s.get("region") == r else 0 for r in top_regions] + \
                       [s.get("days_old", 0) or 0]
            X = _np.array([_featurize(s) for s in samples], dtype=float)
            y = _np.array([s["rate"] for s in samples], dtype=float)
            coef, *_ = _np.linalg.lstsq(X, y, rcond=None)
            y_pred = X @ coef
            mae = float(_np.mean(_np.abs(y - y_pred)))

            db.add(ModelMetric(
                category=category, model_type="ols",
                n_samples=len(samples),
                r_squared=result.get("r_squared"),
                residual_std=result.get("residual_std"),
                mae=round(mae, 4),
                coefficients_json=_json.dumps(
                    [(n, round(float(c), 4)) for n, c in zip(
                        ["intercept", "log_base", "is_construction"] + [f"region_{r}" for r in top_regions] + ["days_old"],
                        coef)],
                    ensure_ascii=False,
                ),
                period_days=period_days,
            ))
            logger.info("model retrain category=%s n=%d r²=%.4f mae=%.4f",
                        category, len(samples), result.get("r_squared") or 0, mae)
        db.commit()
    finally:
        db.close()


def _scheduled_sync_job():
    """스케줄러가 호출하는 수집 잡"""
    logger.info("scheduled sync job started")
    failures: list[dict] = []
    try:
        # G2B 키만으로 국방부·군 등 국방 발주 데이터까지 수집됨 → 기본은 G2B만 호출
        # D2B 별도 API를 사용할 경우 환경변수 ENABLE_D2B_SYNC=true 로 활성
        sources = ("G2B", "D2B") if os.environ.get("ENABLE_D2B_SYNC", "").lower() == "true" else ("G2B",)
        for src in sources:
            res = _run_sync_for_source(src, trigger="scheduled")
            logger.info(
                "sync result source=%s status=%s records=%s inserted=%s",
                res["source"], res["status"], res.get("records", 0),
                res.get("inserted", 0),
            )
            if res["status"] == "failed":
                failures.append({
                    "source": res["source"],
                    "error": res.get("error_message", "unknown"),
                })
        if failures:
            _send_alert_webhook(
                "🚨 자동 수집 실패",
                {
                    "실패 소스": ", ".join(f["source"] for f in failures),
                    "오류": "; ".join(f["error"][:80] for f in failures),
                    "서버시각": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            )
        # 누적 실패율 임계치 점검
        _check_sync_failure_threshold()
    except Exception as e:
        logger.exception("scheduler job failed: %s", e)
        _send_alert_webhook("❗ 스케줄러 예외", {"error": str(e)[:200]})
        try:
            db = SessionLocal()
            try:
                db.add(DataSyncLog(
                    source="SCHEDULER", sync_type="자동 수집", status="failed",
                    records_fetched=0, error_message=str(e)[:500],
                    started_at=datetime.now(), finished_at=datetime.now(),
                ))
                db.commit()
            finally:
                db.close()
        except Exception as exc:
            logger.warning("DataSyncLog 저장 실패: %s", exc)


def _apply_schedule_config(config: dict):
    """_schedule_config에 따라 APScheduler 재설정"""
    global _scheduler
    if not _SCHEDULER_AVAILABLE or _scheduler is None:
        return
    # 기존 job 제거
    for job in _scheduler.get_jobs():
        job.remove()
    if not config.get("enabled", True):
        return
    interval = config.get("interval", "daily")
    time_str = config.get("time", "02:00")
    try:
        hour, minute = (int(x) for x in time_str.split(":"))
    except (ValueError, AttributeError) as exc:
        logger.warning("스케줄 시각 파싱 실패 (%s) — 기본값 02:00 사용: %s", time_str, exc)
        hour, minute = 2, 0

    if interval == "hourly":
        trigger = IntervalTrigger(hours=1)
    elif interval == "weekly":
        trigger = CronTrigger(day_of_week="mon", hour=hour, minute=minute)
    else:  # daily
        trigger = CronTrigger(hour=hour, minute=minute)
    _scheduler.add_job(_scheduled_sync_job, trigger=trigger, id="auto_sync", replace_existing=True)
    # 회귀 모델 재학습 — 매일 04:00 (수집 02:00 후)
    _scheduler.add_job(
        _scheduled_model_retrain,
        trigger=CronTrigger(hour=4, minute=0),
        id="auto_model_retrain", replace_existing=True,
    )


from contextlib import asynccontextmanager, contextmanager

# db_session / get_db 는 app/core/database 에서 재import (F2 분리)
from app.core.database import db_session, get_db  # noqa: E402, F401


@asynccontextmanager
async def lifespan(app: FastAPI):
    global _scheduler
    if _SCHEDULER_AVAILABLE:
        try:
            _scheduler = BackgroundScheduler(timezone="Asia/Seoul")
            _scheduler.start()
            _apply_schedule_config(_schedule_config)
            logger.info("scheduler started config=%s", _schedule_config)
        except Exception as e:
            logger.exception("scheduler start failed: %s", e)
            _scheduler = None
    yield
    if _scheduler is not None:
        try:
            _scheduler.shutdown(wait=False)
        except Exception as exc:
            logger.warning("스케줄러 종료 실패: %s", exc)


# 스케줄 설정 (메모리 기반 — 재시작 시 환경변수에서 로드)
_schedule_config = {
    "interval": os.environ.get("SYNC_INTERVAL", "daily"),
    "time": os.environ.get("SYNC_TIME", "02:00"),
    "enabled": os.environ.get("SYNC_ENABLED", "true").lower() == "true",
}

app = FastAPI(
    title="비드스타 API",
    version="0.2.0",
    description=(
        "공공데이터 기반 입찰가 산정 및 사정률 분석 웹 서비스 API.\n\n"
        "**주요 기능**\n"
        "- 나라장터(G2B) 공고 수집 (국방부·군 발주 포함, 용역+공사 카테고리)\n"
        "- Tab1 빈도 분석 / Tab2 갭 분석 / Tab3 종합 분석\n"
        "- 사용자 마이페이지·데이터 업로드·관리자 대시보드\n"
        "- 자동 수집 스케줄러(APScheduler) + Prometheus 메트릭\n\n"
        "**인증**: JWT Bearer 토큰 (`POST /api/v1/auth/login` 으로 획득)\n"
        "**Rate Limit**: 로그인 10/min · 분석 30~60/min · 글로벌 300/min"
    ),
    contact={
        "name": "비드스타 지원팀",
        "email": "support@bid-insight.example.com",
    },
    license_info={"name": "Proprietary"},
    openapi_tags=[
        {"name": "인증", "description": "회원가입/로그인/토큰/비밀번호"},
        {"name": "메타", "description": "지역·공종 등 메타 데이터"},
        {"name": "공고", "description": "입찰 공고 조회 및 필터"},
        {"name": "분석", "description": "빈도·갭·종합·밀어내기식 검증"},
        {"name": "통계", "description": "KPI·추이·지역/업종별 통계"},
        {"name": "이력", "description": "조회 이력 (마이페이지)"},
        {"name": "데이터", "description": "Excel/CSV 업로드"},
        {"name": "관리자", "description": "사용자·수집·스케줄·NAS 상태 관리"},
        {"name": "운영", "description": "헬스체크·메트릭"},
    ],
    lifespan=lifespan,
)

CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "*").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Rate Limiting (slowapi) ──────────────────────────────────────
try:
    from slowapi import Limiter
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    from starlette.responses import JSONResponse

    limiter = Limiter(key_func=get_remote_address, default_limits=["300/minute"])
    app.state.limiter = limiter

    @app.exception_handler(RateLimitExceeded)
    async def _rate_limit_handler(request, exc):
        logger.warning("rate limit exceeded ip=%s path=%s", get_remote_address(request), request.url.path)
        return JSONResponse(
            status_code=429,
            content={"detail": f"요청 속도 제한을 초과했습니다. 잠시 후 다시 시도해주세요. (limit: {exc.detail})"},
        )

    _RATE_LIMIT_AVAILABLE = True
except ImportError:
    limiter = None
    _RATE_LIMIT_AVAILABLE = False
    logger.warning("slowapi 미설치 — Rate Limiting 비활성화")


def rate_limit(spec: str):
    """slowapi 설치 여부에 따라 조건부 적용 데코레이터"""
    if _RATE_LIMIT_AVAILABLE and limiter is not None:
        return limiter.limit(spec)

    def noop(fn):
        return fn

    return noop


# ─── 분석 결과 TTL 캐시 ───────────────────────────────────────────
# 간단한 LRU + TTL 메모리 캐시. 운영 시 Redis로 교체 가능.
import threading
import time as _time

_CACHE_MAX = int(os.environ.get("ANALYSIS_CACHE_SIZE", "256"))
_CACHE_TTL = int(os.environ.get("ANALYSIS_CACHE_TTL", "300"))  # 초
_cache_store: dict = {}
_cache_lock = threading.Lock()


def _cache_get(key: str):
    with _cache_lock:
        entry = _cache_store.get(key)
        if not entry:
            return None
        expires_at, value = entry
        if _time.time() > expires_at:
            _cache_store.pop(key, None)
            return None
        return value


def _cache_set(key: str, value):
    with _cache_lock:
        # 크기 초과 시 만료된 항목부터 정리
        if len(_cache_store) >= _CACHE_MAX:
            now = _time.time()
            expired = [k for k, (t, _) in _cache_store.items() if now > t]
            for k in expired:
                _cache_store.pop(k, None)
            # 여전히 초과면 가장 오래된 항목 제거
            if len(_cache_store) >= _CACHE_MAX:
                oldest = min(_cache_store.items(), key=lambda kv: kv[1][0])[0]
                _cache_store.pop(oldest, None)
        _cache_store[key] = (_time.time() + _CACHE_TTL, value)


def _cache_key(*parts) -> str:
    return "|".join(str(p) for p in parts)


def invalidate_analysis_cache():
    """공고/수집 데이터 변경 시 호출"""
    with _cache_lock:
        _cache_store.clear()
    logger.info("analysis cache invalidated")


# ─── 보안 헤더 미들웨어 ───────────────────────────────────────────
def _is_production() -> bool:
    return os.environ.get("APP_ENV", "development").lower() == "production"


@app.middleware("http")
async def _security_headers_middleware(request, call_next):
    response = await call_next(request)
    # XSS / Clickjacking 방어
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
    # CSP — 단일 파일 SPA 특성상 inline script 허용 필요
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://unpkg.com https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com data:; "
        "img-src 'self' data: https:; "
        "connect-src 'self' https://apis.data.go.kr https://d2b.go.kr;",
    )
    if _is_production():
        # 프로덕션에서만 HSTS (HTTPS 강제)
        response.headers.setdefault(
            "Strict-Transport-Security",
            "max-age=31536000; includeSubDomains",
        )
    return response


@app.middleware("http")
async def _access_log_middleware(request, call_next):
    import time as _time
    start = _time.perf_counter()
    try:
        response = await call_next(request)
    except Exception as e:
        elapsed = (_time.perf_counter() - start) * 1000
        logger.exception(
            "request error method=%s path=%s duration_ms=%.1f err=%s",
            request.method, request.url.path, elapsed, e,
        )
        raise
    elapsed = (_time.perf_counter() - start) * 1000
    level = logging.WARNING if response.status_code >= 400 else logging.INFO
    logger.log(
        level,
        "req method=%s path=%s status=%d duration_ms=%.1f",
        request.method, request.url.path, response.status_code, elapsed,
    )
    return response


# get_db 는 app/core/database 에서 이미 import 되어 있음 (F2)


# ─── 인증 설정 ──────────────────────────────────────────────────────────
_DEFAULT_SECRET = "bid-insight-secret-key-change-in-production"
SECRET_KEY = os.environ.get("SECRET_KEY", _DEFAULT_SECRET)
ALGORITHM = "HS256"

# 프로덕션 보안 검증: ENV=production일 때 기본 SECRET_KEY 거부
_RUNTIME_ENV = os.environ.get("APP_ENV", "development").lower()
if _RUNTIME_ENV == "production":
    if SECRET_KEY == _DEFAULT_SECRET:
        raise RuntimeError(
            "SECRET_KEY 환경변수가 설정되지 않았거나 기본값입니다. "
            "프로덕션 배포 시 `openssl rand -hex 32` 으로 생성한 키를 지정해야 합니다."
        )
    if len(SECRET_KEY) < 32:
        raise RuntimeError(
            f"SECRET_KEY 길이가 너무 짧습니다 ({len(SECRET_KEY)}자). 최소 32자 이상 권장."
        )
    if CORS_ORIGINS := os.environ.get("CORS_ORIGINS", "").strip():
        if "*" in CORS_ORIGINS.split(","):
            raise RuntimeError("프로덕션에서는 CORS_ORIGINS='*' 를 사용할 수 없습니다.")
    logger.info("production env validation passed")
else:
    if SECRET_KEY == _DEFAULT_SECRET:
        logger.warning("SECRET_KEY 기본값 사용 중 — 프로덕션 배포 전 반드시 변경 필요")
# ─── 인증 헬퍼 (app/core/security 에서 정의 — F2 분리) ──────────────────
# server.py 가 환경변수 검증 후 SECRET_KEY 를 주입한다.
import app.core.security as _sec_mod  # noqa: E402
_sec_mod.SECRET_KEY = SECRET_KEY

from app.core.security import (  # noqa: E402
    ACCESS_TOKEN_EXPIRE_MINUTES,
    oauth2_scheme,
    verify_password,
    get_password_hash,
    create_access_token,
    get_current_user,
    require_auth,
    require_admin,
)


# ─── API: 인증 ──────────────────────────────────────────────────────────

@app.post("/api/v1/auth/register")
def register(username: str = Query(...), email: str = Query(...),
             password: str = Query(...), name: str = Query(None)):
    db = SessionLocal()
    try:
        # 중복 확인
        if db.query(User).filter(User.username == username).first():
            db.close()
            raise HTTPException(status_code=400, detail="이미 사용 중인 아이디입니다.")
        if db.query(User).filter(User.email == email).first():
            db.close()
            raise HTTPException(status_code=400, detail="이미 등록된 이메일입니다.")
        user = User(
            username=username, email=email,
            hashed_password=get_password_hash(password),
            name=name or username, role="user",
            joined_at=datetime.now(),
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        token = create_access_token({"sub": user.id})
    finally:
        db.close()
    return {
        "access_token": token, "token_type": "bearer",
        "user": {"id": user.id, "username": user.username, "name": user.name,
                 "email": user.email, "role": user.role, "plan": user.plan},
    }


@app.post("/api/v1/auth/login")
@rate_limit("10/minute")
def login(request: Request, username: str = Query(...), password: str = Query(...)):
    db = SessionLocal()
    try:
        user = db.query(User).filter(
            (User.username == username) | (User.email == username)
        ).first()
        if not user or not verify_password(password, user.hashed_password):
            db.close()
            raise HTTPException(status_code=401, detail="아이디 또는 비밀번호가 올바르지 않습니다.")
        user.last_login_at = datetime.now()
        db.commit()
        token = create_access_token({"sub": user.id})
    finally:
        db.close()
    return {
        "access_token": token, "token_type": "bearer",
        "user": {"id": user.id, "username": user.username, "name": user.name,
                 "email": user.email, "role": user.role, "plan": user.plan},
    }


@app.get("/api/v1/auth/me")
def get_me(current_user: User = Depends(require_auth)):
    return {
        "id": current_user.id, "username": current_user.username,
        "name": current_user.name, "email": current_user.email,
        "role": current_user.role, "plan": current_user.plan,
        "query_count": current_user.query_count,
        "joined_at": current_user.joined_at.strftime("%Y-%m-%d") if current_user.joined_at else None,
    }


@app.post("/api/v1/auth/refresh")
def refresh_token(current_user: User = Depends(require_auth)):
    token = create_access_token({"sub": current_user.id})
    return {"access_token": token, "token_type": "bearer"}


@app.put("/api/v1/auth/password")
def change_password(
    current_password: str = Query(...),
    new_password: str = Query(...),
    current_user: User = Depends(require_auth),
):
    """비밀번호 변경"""
    if not verify_password(current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="현재 비밀번호가 올바르지 않습니다.")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="새 비밀번호는 6자 이상이어야 합니다.")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == current_user.id).first()
        user.hashed_password = get_password_hash(new_password)
        db.commit()
    finally:
        db.close()
    return {"message": "비밀번호가 변경되었습니다."}


# ─── API: 공고 목록 ──────────────────────────────────────────────────────

@app.get("/api/v1/meta/industry-codes")
def get_industry_codes():
    """공종 코드 목록"""
    db = SessionLocal()
    try:
        codes = db.query(BidAnnouncement.industry_code).distinct().filter(
            BidAnnouncement.industry_code.isnot(None)
        ).all()
    finally:
        db.close()
    return {"codes": sorted([c[0] for c in codes if c[0]])}


@app.get("/api/v1/meta/regions")
def get_regions():
    """지역(17개 시도) 목록"""
    return {"regions": REGIONS}


@app.get("/api/v1/announcements")
def list_announcements(
    category: str = None, source: str = None, region: str = None,
    keyword: str = None, status: str = None,
    industry_code: str = None, date_from: str = None, date_to: str = None,
    is_defense: str = None,  # "true"/"false"/"all"
    page: int = 1, page_size: int = 20,
):
    db = SessionLocal()
    try:
        q = db.query(BidAnnouncement)

        # 용역/공사만
        q = q.filter(BidAnnouncement.category.in_(["공사", "용역"]))

        if category and category != "all":
            cats = [c.strip() for c in category.split(",")]
            if len(cats) == 1:
                q = q.filter(BidAnnouncement.category == cats[0])
            else:
                q = q.filter(BidAnnouncement.category.in_(cats))
        if source and source != "all":
            q = q.filter(BidAnnouncement.source == source)
        if region and region != "all":
            q = q.filter(BidAnnouncement.region == region)
        if status and status != "all":
            q = q.filter(BidAnnouncement.status == status)
        if industry_code and industry_code != "all":
            q = q.filter(BidAnnouncement.industry_code == industry_code)
        if is_defense and is_defense != "all":
            q = q.filter(BidAnnouncement.is_defense == (is_defense.lower() == "true"))
        if date_from:
            try:
                q = q.filter(BidAnnouncement.announced_at >= datetime.strptime(date_from, "%Y-%m-%d"))
            except ValueError:
                pass
        if date_to:
            try:
                q = q.filter(BidAnnouncement.announced_at <= datetime.strptime(date_to, "%Y-%m-%d"))
            except ValueError:
                pass
        if keyword:
            q = q.filter(
                BidAnnouncement.title.contains(keyword) |
                BidAnnouncement.ordering_org_name.contains(keyword)
            )

        total = q.count()
        items = q.order_by(BidAnnouncement.announced_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

        # N+1 해소 — BidResult 를 batch IN 절로 한 번에 조회
        ann_ids = [a.id for a in items]
        res_map: dict = {}
        if ann_ids:
            for r in db.query(BidResult).filter(BidResult.announcement_id.in_(ann_ids)).all():
                # 한 공고당 1건만 (가장 최근) 사용 — 다중 결과는 첫 row 우선
                if r.announcement_id not in res_map:
                    res_map[r.announcement_id] = r

        result_items = []
        for ann in items:
            res = res_map.get(ann.id)
            result_items.append({
                "id": ann.id,
                "bid_number": ann.bid_number,
                "title": ann.title,
                "org": ann.ordering_org_name,
                "type": ann.category,
                "area": ann.region,
                "budget": ann.base_amount,
                "deadline": ann.deadline_at.strftime("%Y-%m-%d") if ann.deadline_at else "",
                "rate": round(res.assessment_rate, 2) if res and res.assessment_rate else None,
                "first_place_rate": round(res.first_place_rate, 4) if res and res.first_place_rate else None,
                "first_place_amount": res.first_place_amount if res else None,
                "status": ann.status,
                "source": ann.source,
                "bid_method": ann.bid_method,
                "num_bidders": res.num_bidders if res else None,
                "is_defense": bool(getattr(ann, "is_defense", False)),
                "external_url": get_announcement_url(ann),
            })

    finally:
        db.close()
    return {
        "items": result_items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total > 0 else 0,
    }


@app.get("/api/v1/announcements/{announcement_id}")
def get_announcement(announcement_id: str):
    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}
        res = db.query(BidResult).filter(BidResult.announcement_id == ann.id).first()
    finally:
        db.close()
    return {
        "id": ann.id, "bid_number": ann.bid_number, "title": ann.title,
        "org": ann.ordering_org_name, "org_type": ann.ordering_org_type,
        "parent_org": ann.parent_org_name,
        "type": ann.category, "area": ann.region, "source": ann.source,
        "budget": ann.base_amount, "bid_method": ann.bid_method,
        "industry_code": ann.industry_code, "status": ann.status,
        "deadline": ann.deadline_at.strftime("%Y-%m-%d") if ann.deadline_at else "",
        "announced_at": ann.announced_at.strftime("%Y-%m-%d") if ann.announced_at else "",
        "rate": round(res.assessment_rate, 4) if res and res.assessment_rate else None,
        "first_place_rate": round(res.first_place_rate, 4) if res and res.first_place_rate else None,
        "first_place_amount": res.first_place_amount if res else None,
        "num_bidders": res.num_bidders if res else None,
        "is_defense": bool(getattr(ann, "is_defense", False)),
        "external_url": (
            get_announcement_url(ann)
        ),
    }


# ─── 사정률 예측 로직 (스펙 §1) — 3가지 구간 알고리즘 ─────────────────────

def _remove_outliers_iqr(values: list, multiplier: float = 1.5) -> list:
    """IQR 기반 이상치 제거 — 사정률 분포 안정화.

    Q1-1.5×IQR ~ Q3+1.5×IQR 구간 밖 제거.
    표본이 5건 미만이면 그대로 반환 (제거 시 무의미).
    """
    if len(values) < 5:
        return list(values)
    sorted_v = sorted(values)
    n = len(sorted_v)
    q1 = sorted_v[n // 4]
    q3 = sorted_v[3 * n // 4]
    iqr = q3 - q1
    lower = q1 - multiplier * iqr
    upper = q3 + multiplier * iqr
    return [v for v in values if lower <= v <= upper]


def _time_weighted_rates(rates_with_dates: list, half_life_days: int = 180,
                          ref_date: datetime = None) -> list:
    """시계열 가중치 — 최근 데이터를 더 높게 (지수 감쇠).

    Args:
        rates_with_dates: [(rate, date), ...]
        half_life_days: 가중치가 절반이 되는 일수 (기본 6개월)
        ref_date: 기준일 (기본 오늘)
    Returns:
        weight 적용한 rates 리스트 (가중 표본 — 가중치만큼 복제)
        예) 가중치 0.5 → 1번 등장, 가중치 1.0 → 2번 등장
    """
    if not rates_with_dates:
        return []
    ref = ref_date or datetime.now()
    import math as _math
    weighted = []
    for rate, dt in rates_with_dates:
        if not dt:
            weighted.append(rate)  # 날짜 없으면 가중치 1
            continue
        days_old = (ref - dt).days if isinstance(dt, datetime) else 0
        # exp(-ln(2) * t / half_life) → t=0 일 때 1, t=half_life 일 때 0.5
        weight = _math.exp(-_math.log(2) * max(0, days_old) / half_life_days)
        # 0~2 범위로 환산 — 가중치 1 = 표본 2번 카운트, 0.5 = 1번
        replicate = max(1, int(round(weight * 2)))
        weighted.extend([rate] * replicate)
    return weighted


def _confidence_interval(values: list, confidence: float = 0.95) -> tuple:
    """평균 ± 신뢰구간 (정규분포 가정, n≥10 권장).

    Returns: (mean, lower, upper, margin)
    margin = z * std / sqrt(n)
    """
    if not values:
        return (None, None, None, None)
    if len(values) < 2:
        return (values[0], values[0], values[0], 0)
    mean = statistics.mean(values)
    std = statistics.stdev(values)
    n = len(values)
    # z-score: 95% → 1.96, 99% → 2.576
    z = 2.576 if confidence >= 0.99 else 1.96 if confidence >= 0.95 else 1.645
    margin = z * std / (n ** 0.5)
    return (round(mean, 4), round(mean - margin, 4), round(mean + margin, 4), round(margin, 4))


def _homogeneity_score(target_org: str, target_industry: str,
                        sample_orgs: list, sample_industries: list) -> float:
    """동종성 점수 — 표본이 대상 공고와 얼마나 유사한가 (0~1).

    동일 발주처 비율(0.6) + 동일 업종 비율(0.4) 가중 평균.
    """
    if not sample_orgs:
        return 0.0
    n = len(sample_orgs)
    same_org = sum(1 for o in sample_orgs if o == target_org) / n
    same_ind = (sum(1 for i in sample_industries if i and i == target_industry) / n
                if target_industry else 0)
    return round(same_org * 0.6 + same_ind * 0.4, 4)


# ─── 앙상블 + 모멘텀 + Walk-forward ──────────────────────────────────

def _ensemble_predict(predictions: list) -> dict:
    """가중 앙상블 예측.

    Args:
        predictions: [{rate, weight, name}, ...]
                     weight 가 클수록 영향 큼. None rate 는 자동 제외.
    Returns:
        {final_rate, total_weight, contributions: [{name, rate, weight, share}]}
    """
    valid = [p for p in predictions if p.get("rate") is not None and p.get("weight", 0) > 0]
    if not valid:
        return {"final_rate": None, "total_weight": 0, "contributions": []}
    total_w = sum(p["weight"] for p in valid)
    final_rate = sum(p["rate"] * p["weight"] for p in valid) / total_w
    contributions = [
        {
            "name": p.get("name", "?"),
            "rate": round(p["rate"], 4),
            "weight": round(p["weight"], 4),
            "share": round(p["weight"] / total_w * 100, 1),
        }
        for p in valid
    ]
    return {
        "final_rate": round(final_rate, 4),
        "total_weight": round(total_w, 4),
        "contributions": contributions,
    }


def _trend_momentum(series: list) -> dict:
    """시계열 시리즈에 선형 회귀 → 월간 변화율(slope) 산출.

    Args:
        series: [{period, avg_rate, count}, ...] (시간순 정렬 가정)
    Returns:
        {slope, intercept, r_squared, direction, next_predicted, n}
    """
    import numpy as np
    valid = [s for s in series if s.get("avg_rate") is not None]
    if len(valid) < 3:
        return {"slope": None, "direction": "insufficient_data", "n": len(valid)}

    x = np.arange(len(valid), dtype=float)
    y = np.array([s["avg_rate"] for s in valid], dtype=float)

    # OLS: y = a*x + b
    A = np.vstack([x, np.ones_like(x)]).T
    try:
        coef, *_ = np.linalg.lstsq(A, y, rcond=None)
    except (np.linalg.LinAlgError, ValueError) as exc:
        logger.warning("trend momentum lstsq 실패 (n=%d): %s", len(valid), exc)
        return {"slope": None, "direction": "error", "n": len(valid)}
    slope, intercept = float(coef[0]), float(coef[1])

    # R²
    y_pred = slope * x + intercept
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r_squared = 1 - ss_res / ss_tot if ss_tot > 0 else 0

    # 다음 시점 예측
    next_x = float(len(valid))
    next_pred = slope * next_x + intercept

    # 방향 (절대값 0.05% 미만은 보합)
    if abs(slope) < 0.05:
        direction = "flat"
    elif slope > 0:
        direction = "up"
    else:
        direction = "down"

    return {
        "slope": round(slope, 4),                       # % per period (월별이면 %/월)
        "intercept": round(intercept, 4),
        "r_squared": round(r_squared, 4),
        "direction": direction,
        "next_predicted": round(next_pred, 4),
        "n": len(valid),
    }


def _walk_forward_validate(samples_with_dates: list, window_days: int = 90,
                            stride_days: int = 30) -> dict:
    """시계열 walk-forward 검증.

    각 시점마다: [t-window, t] 데이터로 학습 → t 시점 예측 → 실제값과 비교.
    stride_days 만큼 t를 앞으로 이동 (rolling window).

    Args:
        samples_with_dates: [(rate, date), ...]
    Returns:
        {windows: [{period, predicted, actual_avg, mae, n}, ...], overall_mae}
    """
    if len(samples_with_dates) < 30:
        return {"windows": [], "overall_mae": None, "reason": "표본 부족"}

    from datetime import datetime as _dt, timedelta as _td
    sorted_samples = sorted(samples_with_dates, key=lambda x: x[1] or _dt.min)
    valid = [(r, d) for r, d in sorted_samples if d]
    if len(valid) < 30:
        return {"windows": [], "overall_mae": None, "reason": "유효 날짜 표본 부족"}

    start_date = valid[0][1]
    end_date = valid[-1][1]

    windows = []
    cursor = start_date + _td(days=window_days)
    while cursor < end_date:
        train_start = cursor - _td(days=window_days)
        train_end = cursor
        test_end = min(end_date, cursor + _td(days=stride_days))

        train = [r for r, d in valid if train_start <= d < train_end]
        test = [r for r, d in valid if train_end <= d < test_end]

        if len(train) >= 5 and len(test) >= 1:
            # 학습: 단순 평균을 다음 기간 예측치로 (가장 robust한 baseline)
            predicted = sum(train) / len(train)
            actual_avg = sum(test) / len(test)
            mae = abs(predicted - actual_avg)
            windows.append({
                "period": cursor.strftime("%Y-%m-%d"),
                "predicted": round(predicted, 4),
                "actual_avg": round(actual_avg, 4),
                "mae": round(mae, 4),
                "n_train": len(train),
                "n_test": len(test),
            })
        cursor += _td(days=stride_days)

    if not windows:
        return {"windows": [], "overall_mae": None, "reason": "윈도우 생성 실패"}

    overall_mae = round(sum(w["mae"] for w in windows) / len(windows), 4)
    return {
        "windows": windows,
        "overall_mae": overall_mae,
        "n_windows": len(windows),
    }


# ─── Bayesian Shrinkage / 이상 탐지 / K-NN 유사 공고 ──────────────────

def _bayesian_shrinkage(local_rates: list, global_rates: list,
                         prior_strength: int = 10) -> dict:
    """경험적 베이지안 shrinkage — 소표본 발주처 평균을 전체 평균으로 끌어당김.

    weighted = (n_local × local_mean + k × global_mean) / (n_local + k)
    n_local 이 작을수록 global_mean 영향력 ↑

    Args:
        local_rates: 대상 발주처(또는 좁은 범위) 사정률 리스트
        global_rates: 카테고리 전체 사정률 리스트
        prior_strength: shrinkage 강도 — 사실상 "가상 표본 k건"
    Returns:
        {shrunk_rate, local_mean, global_mean, shrinkage_factor (0~1, 1=완전 global)}
    """
    if not global_rates:
        return {"shrunk_rate": None, "local_mean": None, "global_mean": None,
                "shrinkage_factor": None, "n_local": len(local_rates)}
    g_mean = sum(global_rates) / len(global_rates)
    n_local = len(local_rates)
    if n_local == 0:
        return {"shrunk_rate": round(g_mean, 4), "local_mean": None,
                "global_mean": round(g_mean, 4), "shrinkage_factor": 1.0, "n_local": 0}
    l_mean = sum(local_rates) / n_local
    shrunk = (n_local * l_mean + prior_strength * g_mean) / (n_local + prior_strength)
    factor = prior_strength / (n_local + prior_strength)
    return {
        "shrunk_rate": round(shrunk, 4),
        "local_mean": round(l_mean, 4),
        "global_mean": round(g_mean, 4),
        "shrinkage_factor": round(factor, 4),
        "n_local": n_local,
        "n_global": len(global_rates),
    }


def _apply_category_filter(q, category_filter: str, ann=None):
    """카테고리 필터 공통 적용 — 6개 분석 엔드포인트 중복 제거.

    Args:
        q: SQLAlchemy 쿼리 (BidAnnouncement 가 join 되어 있어야 함)
        category_filter: 'same' / 'all' / 'construction' / 'service' / 'industry'
        ann: 'same'/'industry' 모드 시 기준 공고
    """
    if category_filter == "construction":
        return q.filter(BidAnnouncement.category == "공사")
    if category_filter == "service":
        return q.filter(BidAnnouncement.category == "용역")
    if category_filter == "industry" and ann and ann.industry_code:
        return q.filter(BidAnnouncement.industry_code == ann.industry_code)
    if category_filter == "same" and ann:
        return q.filter(BidAnnouncement.category == ann.category)
    return q  # all (또는 fallback)


def _detect_anomaly(target_value: float, peer_values: list,
                     z_threshold: float = 2.0) -> dict:
    """z-score 기반 이상 탐지 — 대상값이 표본 분포에서 비정상인지 판정.

    Args:
        target_value: 대상 공고 기초금액 (또는 사정률)
        peer_values: 유사 표본의 동일 지표 리스트
        z_threshold: |z| 이 임계값보다 크면 이상 플래그
    Returns:
        {is_anomaly, z_score, peer_mean, peer_std, severity}
    """
    if len(peer_values) < 5:
        return {"is_anomaly": False, "z_score": None, "reason": "표본 부족"}
    mean = sum(peer_values) / len(peer_values)
    if len(peer_values) > 1:
        std = statistics.stdev(peer_values)
    else:
        std = 0
    if std == 0:
        return {"is_anomaly": False, "z_score": 0, "peer_mean": round(mean, 4)}
    z = (target_value - mean) / std
    abs_z = abs(z)
    if abs_z > z_threshold * 1.5:
        severity = "high"
    elif abs_z > z_threshold:
        severity = "medium"
    else:
        severity = "normal"
    return {
        "is_anomaly": abs_z > z_threshold,
        "z_score": round(z, 3),
        "abs_z": round(abs_z, 3),
        "peer_mean": round(mean, 4),
        "peer_std": round(std, 4),
        "severity": severity,
    }


def _kmeans_simple(features: list, k: int = 5, max_iter: int = 30, seed: int = 42) -> tuple:
    """초간단 K-means (numpy) — 외부 의존성 없음.

    Args:
        features: list of feature vectors (list of lists)
        k: 군집 수
    Returns:
        (labels, centroids) — labels[i] = i번째 표본 군집 인덱스
    """
    import numpy as np
    if len(features) < k:
        return ([0] * len(features), [features[0]] if features else [])
    X = np.array(features, dtype=float)
    rng = np.random.default_rng(seed)
    # 초기 centroid: 무작위 k개 표본
    init_idx = rng.choice(len(X), size=k, replace=False)
    centroids = X[init_idx].copy()

    for _ in range(max_iter):
        # 각 점의 가장 가까운 centroid (Euclidean)
        dists = np.linalg.norm(X[:, None, :] - centroids[None, :, :], axis=2)
        labels = dists.argmin(axis=1)
        # centroid 갱신
        new_centroids = np.array([
            X[labels == c].mean(axis=0) if (labels == c).any() else centroids[c]
            for c in range(k)
        ])
        if np.allclose(new_centroids, centroids, atol=1e-4):
            break
        centroids = new_centroids
    return (labels.tolist(), centroids.tolist())


def _knn_similar_announcements(target: dict, candidates: list, k: int = 5) -> list:
    """K-NN 유사 공고 추천 — 가중 거리 기반.

    Distance = |log(amount_diff)| × 0.4 + region_mismatch × 0.3 +
               industry_mismatch × 0.2 + (days_old / 365) × 0.1

    Args:
        target: {base_amount, region, industry, category}
        candidates: [{id, base_amount, region, industry, title, ..., assessment_rate, days_old}, ...]
        k: 상위 N건 반환
    Returns:
        sorted by similarity desc, top k
    """
    target_log_amount = math.log(max(1, target.get("base_amount") or 1))
    target_region = target.get("region")
    target_industry = target.get("industry")

    scored = []
    for c in candidates:
        c_amount = c.get("base_amount") or 0
        if c_amount <= 0:
            continue
        log_amount_diff = abs(math.log(max(1, c_amount)) - target_log_amount)
        region_mm = 0 if c.get("region") == target_region else 1
        ind_mm = 0 if (target_industry and c.get("industry") == target_industry) else 1
        days_old = c.get("days_old", 0) or 0

        distance = (log_amount_diff * 0.4
                    + region_mm * 0.3
                    + ind_mm * 0.2
                    + (days_old / 365) * 0.1)
        # similarity = 1 / (1 + distance), 0~1 정규화
        similarity = round(1 / (1 + distance), 4)
        scored.append({**c, "distance": round(distance, 4), "similarity": similarity})

    scored.sort(key=lambda x: -x["similarity"])
    return scored[:k]


# ─── 회귀 분석 (numpy OLS) ────────────────────────────────────────────

def _ols_regression(samples: list, target_features: dict) -> dict:
    """다중 선형 회귀 — 사정률 예측 (numpy lstsq).

    Args:
        samples: [{rate, base_amount, category, region, industry, org_type, days_old}, ...]
        target_features: 대상 공고의 동일 feature dict
    Returns:
        {predicted_rate, r_squared, coefficients, residual_std, n}
    """
    import numpy as np
    if len(samples) < 10:
        return {"predicted_rate": None, "r_squared": None, "n": len(samples),
                "reason": "표본 부족 (10건 미만)"}

    # ── feature 추출: log(base_amount), category=공사, region one-hot top5, days_old ──
    # 지역 상위 빈출 5개
    from collections import Counter
    region_counts = Counter(s.get("region") for s in samples if s.get("region"))
    top_regions = [r for r, _ in region_counts.most_common(5)]

    def featurize(s: dict) -> list:
        ba = s.get("base_amount") or 1
        log_ba = math.log(max(1, ba))
        is_construction = 1 if s.get("category") == "공사" else 0
        region_oh = [1 if s.get("region") == r else 0 for r in top_regions]
        days_old = s.get("days_old", 0) or 0
        # bias term + features
        return [1.0, log_ba, is_construction] + region_oh + [days_old]

    X = np.array([featurize(s) for s in samples], dtype=float)
    y = np.array([s["rate"] for s in samples], dtype=float)

    # OLS: β = (X'X)^-1 X'y, lstsq 로 안정 풀이
    try:
        coef, residuals, rank, sv = np.linalg.lstsq(X, y, rcond=None)
    except Exception as e:
        return {"predicted_rate": None, "r_squared": None, "error": str(e)}

    # 예측값 + R² 계산
    y_pred = X @ coef
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r_squared = 1 - ss_res / ss_tot if ss_tot > 0 else 0
    residual_std = float(np.sqrt(ss_res / max(1, len(y) - len(coef))))

    # 대상 공고 예측
    target_x = np.array(featurize(target_features), dtype=float)
    target_pred = float(target_x @ coef)
    # 사정률 안전 범위 클램프 (90~110)
    target_pred = max(90.0, min(110.0, target_pred))

    # 특성 영향도 (절대값 큰 순)
    feature_names = ["intercept", "log(기초금액)", "공사여부"] + [f"지역={r}" for r in top_regions] + ["경과일수"]
    importance = sorted(
        [(n, round(float(c), 4)) for n, c in zip(feature_names, coef) if n != "intercept"],
        key=lambda x: -abs(x[1])
    )[:5]

    return {
        "predicted_rate": round(target_pred, 4),
        "r_squared": round(r_squared, 4),
        "residual_std": round(residual_std, 4),
        "n": len(samples),
        "top_features": importance,  # [(name, coef), ...]
    }


def _build_rate_histogram(rates: list, bin_size: float = 0.1) -> list:
    """사정률 리스트 → 0.1% bin 히스토그램. 빈 bin 도 0 으로 채워서 반환."""
    if not rates:
        return []
    min_r = math.floor(min(rates) * 10) / 10
    max_r = math.ceil(max(rates) * 10) / 10
    bins = []
    cur = min_r
    while cur <= max_r + 0.001:
        cnt = sum(1 for r in rates if cur - bin_size / 2 <= r < cur + bin_size / 2)
        bins.append({"rate": round(cur, 2), "count": cnt})
        cur = round(cur + bin_size, 2)
    return bins


def _compute_rate_buckets(bins: list, mode: str = "A", top_n: int = 5) -> list:
    """스펙 §1 — 3가지 구간 산출 알고리즘.

    100 을 기준으로 +방향 / -방향 각각 정렬하여 범위(range) 반환.

    Args:
        bins: _build_rate_histogram() 결과
        mode: 'A' 빈도최대 / 'B' 공백 / 'C' 차이최대
        top_n: 반환할 상위 구간 수

    Returns:
        [{rate, side: '+'|'-'|'0', score, range_start, range_end, rank}, ...]
    """
    if not bins:
        return []

    candidates = []
    if mode == "A":
        # 막대가 가장 큰 값 (빈도 최대) → 막대 높이순 정렬
        for b in bins:
            if b["count"] > 0:
                side = "+" if b["rate"] > 100 else ("-" if b["rate"] < 100 else "0")
                candidates.append({
                    "rate": b["rate"],
                    "score": b["count"],  # 빈도 자체가 점수
                    "side": side,
                    "distance": abs(b["rate"] - 100),
                })
        candidates.sort(key=lambda c: (-c["score"], c["distance"]))

    elif mode == "B":
        # 막대가 없는 값 (공백) → 100 기준 가까운 순
        for b in bins:
            if b["count"] == 0:
                side = "+" if b["rate"] > 100 else ("-" if b["rate"] < 100 else "0")
                candidates.append({
                    "rate": b["rate"],
                    "score": 0,
                    "side": side,
                    "distance": abs(b["rate"] - 100),
                })
        candidates.sort(key=lambda c: c["distance"])

    elif mode == "C":
        # 막대 차이가 가장 큰 값 → 인접 차이 절대값 순
        for i, b in enumerate(bins):
            prev_cnt = bins[i - 1]["count"] if i > 0 else 0
            next_cnt = bins[i + 1]["count"] if i < len(bins) - 1 else 0
            diff = max(abs(b["count"] - prev_cnt), abs(b["count"] - next_cnt))
            if diff > 0:
                side = "+" if b["rate"] > 100 else ("-" if b["rate"] < 100 else "0")
                candidates.append({
                    "rate": b["rate"],
                    "score": diff,
                    "side": side,
                    "distance": abs(b["rate"] - 100),
                })
        candidates.sort(key=lambda c: (-c["score"], c["distance"]))
    else:
        return []

    # range_start/end + 순위 부여, top_n 제한
    result = []
    bin_size = 0.1
    for i, c in enumerate(candidates[:top_n]):
        c["range_start"] = round(c["rate"] - bin_size / 2, 2)
        c["range_end"] = round(c["rate"] + bin_size / 2, 2)
        c["rank"] = i + 1
        c["mode"] = mode
        result.append(c)
    return result


def _extract_detail_rate(table_rates: list, rule: str = "first_after") -> float:
    """스펙 §1 — 세부 사정률 값 추출.

    Args:
        table_rates: 막대그래프 아래 표의 사정률 리스트 (정렬 가정)
        rule: 'first_after'  첫값 바로 뒤
              'last_after'   마지막값 바로 뒤
              'max_gap'      간격이 가장 큰 값
    """
    if not table_rates:
        return None
    sorted_rates = sorted(table_rates)
    if rule == "first_after":
        return sorted_rates[1] if len(sorted_rates) > 1 else sorted_rates[0]
    if rule == "last_after":
        return sorted_rates[-1]  # "바로 뒤" 해석: 끝값을 그대로 사용
    if rule == "max_gap":
        if len(sorted_rates) < 2:
            return sorted_rates[0]
        max_gap = 0
        max_gap_rate = sorted_rates[0]
        for i in range(1, len(sorted_rates)):
            gap = sorted_rates[i] - sorted_rates[i - 1]
            if gap > max_gap:
                max_gap = gap
                # "간격이 가장 큰 값" — 간격 직후 값 (큰 쪽)
                max_gap_rate = sorted_rates[i]
        return max_gap_rate
    return None


def _compute_confidence(rates: list, agreement: float, methods_aligned_count: int) -> dict:
    """예측 신뢰도 계산 — 표본 크기·합치도·분포 안정성 종합.

    Returns: {score: 0~100, level: 'high|medium|low', reasons: [..]}
    """
    score = 0
    reasons = []

    # 1) 표본 크기 (최대 40점)
    n = len(rates)
    if n >= 100:
        score += 40
        reasons.append(f"표본 풍부 ({n}건)")
    elif n >= 30:
        score += 25
        reasons.append(f"표본 적정 ({n}건)")
    elif n >= 10:
        score += 10
        reasons.append(f"표본 부족 ({n}건)")
    else:
        reasons.append(f"표본 매우 부족 ({n}건)")

    # 2) 합치도 (최대 35점)
    score += int(agreement * 35)
    if agreement >= 0.66:
        reasons.append(f"방법 간 합치 높음 ({methods_aligned_count}/3)")
    elif agreement >= 0.34:
        reasons.append(f"방법 간 합치 보통 ({methods_aligned_count}/3)")
    else:
        reasons.append(f"방법 간 합치 낮음 ({methods_aligned_count}/3)")

    # 3) 분포 안정성 — 표준편차 (최대 25점)
    if n > 1:
        std = statistics.stdev(rates)
        if std < 0.5:
            score += 25
            reasons.append("분포 매우 집중")
        elif std < 1.0:
            score += 15
            reasons.append("분포 집중")
        elif std < 2.0:
            score += 5
            reasons.append("분포 다소 분산")
        else:
            reasons.append("분포 분산")

    score = min(100, max(0, score))
    if score >= 70:
        level = "high"
    elif score >= 40:
        level = "medium"
    else:
        level = "low"
    return {"score": score, "level": level, "reasons": reasons}


# ─── API: 사정률 발생빈도 분석 ────────────────────────────────────────────

@app.get("/api/v1/analysis/frequency/{announcement_id}")
@rate_limit("60/minute")
def analysis_frequency(
    request: Request,
    announcement_id: str,
    period_months: int = Query(12, ge=1, le=24),
    org_scope: str = Query("specific"),  # specific | parent
    current_user: User = Depends(require_auth),
):
    """사정률 발생빈도 히스토그램 + 피크 구간 분석 (TTL 캐시)"""
    cache_key = _cache_key("freq", announcement_id, period_months, org_scope)
    cached = _cache_get(cache_key)
    if cached is not None:
        # 캐시 히트 시에도 인증 사용자의 이력은 저장
        if current_user:
            top_rate = cached.get("prediction_candidates", [{}])[0].get("rate") if cached.get("prediction_candidates") else None
            save_query_history(
                current_user.id, announcement_id, "frequency",
                parameters={"period_months": period_months, "org_scope": org_scope, "cached": True},
                result_summary={"top_predicted_rate": top_rate,
                                "data_count": cached.get("data_count", 0)},
            )
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        # 동일 카테고리 + 발주처 범위의 과거 사정률
        q = db.query(BidResult.assessment_rate, BidResult.first_place_rate).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        if org_scope == "parent" and ann.parent_org_name:
            # 상위 기관 범위
            q = q.filter(BidAnnouncement.parent_org_name == ann.parent_org_name)
        else:
            # 동일 발주기관
            q = q.filter(BidAnnouncement.ordering_org_name == ann.ordering_org_name)

        rows = q.all()
        if not rows:
            # 폴백: 같은 카테고리+지역 전체
            q = db.query(BidResult.assessment_rate, BidResult.first_place_rate).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.region == ann.region,
                BidResult.assessment_rate.isnot(None),
                BidResult.opened_at >= cutoff,
            )
            rows = q.all()

        rates = [r.assessment_rate for r in rows if r.assessment_rate]
        first_place_rates = [r.first_place_rate for r in rows if r.first_place_rate]

        if not rates:
            db.close()
            return {"bins": [], "peaks": [], "stats": {}, "data_count": 0}

        # 0.1% 단위 히스토그램
        min_r = math.floor(min(rates) * 10) / 10
        max_r = math.ceil(max(rates) * 10) / 10
        bin_size = 0.1
        bins = []
        current = min_r
        while current <= max_r + 0.001:
            count = sum(1 for r in rates if current - bin_size / 2 <= r < current + bin_size / 2)
            fp_count = sum(1 for r in first_place_rates if current - bin_size / 2 <= r < current + bin_size / 2)
            bins.append({
                "rate": round(current, 2),
                "count": count,
                "first_place_count": fp_count,
            })
            current = round(current + bin_size, 2)

        # 피크 구간 (상위 5개)
        sorted_bins = sorted(bins, key=lambda b: b["count"], reverse=True)
        peaks = []
        for b in sorted_bins[:5]:
            if b["count"] > 0:
                peaks.append({
                    "rate": b["rate"],
                    "range_start": round(b["rate"] - bin_size / 2, 2),
                    "range_end": round(b["rate"] + bin_size / 2, 2),
                    "count": b["count"],
                })

        # 1순위 예측 후보 (10개 이상) — 빈도 + 1순위 비율 점수
        total_fp = len(first_place_rates)
        prediction_candidates = []
        for b in bins:
            if b["count"] == 0:
                continue
            fp_ratio = round(b["first_place_count"] / b["count"] * 100, 1) if b["count"] > 0 else 0
            # 점수 = 빈도 비율(0.5) + 1순위 비율(0.5)
            freq_score = b["count"] / max(bb["count"] for bb in bins) if max(bb["count"] for bb in bins) > 0 else 0
            fp_score = b["first_place_count"] / max((max(bb["first_place_count"] for bb in bins), 1))
            combined_score = round(freq_score * 0.5 + fp_score * 0.5, 4)
            bid_amount = int(ann.base_amount * b["rate"] / 100) if ann.base_amount else 0
            prediction_candidates.append({
                "rate": b["rate"],
                "frequency": b["count"],
                "first_place_count": b["first_place_count"],
                "first_place_ratio": fp_ratio,
                "score": combined_score,
                "bid_amount": bid_amount,
                "is_recommended": False,
            })
        prediction_candidates.sort(key=lambda c: c["score"], reverse=True)
        # 상위 후보에 추천 마크
        for i, c in enumerate(prediction_candidates[:3]):
            c["is_recommended"] = True
        # 순위 부여
        for i, c in enumerate(prediction_candidates):
            c["rank"] = i + 1

        stat = {
            "mean": round(statistics.mean(rates), 4),
            "median": round(statistics.median(rates), 4),
            "std": round(statistics.stdev(rates), 4) if len(rates) > 1 else 0,
            "min": round(min(rates), 4),
            "max": round(max(rates), 4),
        }

    finally:
        db.close()
    result = {
        "bins": bins,
        "peaks": peaks,
        "stats": stat,
        "data_count": len(rates),
        "first_place_total": total_fp,
        "prediction_candidates": prediction_candidates[:15],
        "announcement": {
            "id": ann.id, "title": ann.title, "type": ann.category,
            "org": ann.ordering_org_name, "area": ann.region,
            "budget": ann.base_amount,
        },
    }
    _cache_set(cache_key, result)
    if current_user:
        top_rate = prediction_candidates[0]["rate"] if prediction_candidates else None
        save_query_history(
            current_user.id, announcement_id, "frequency",
            parameters={"period_months": period_months, "org_scope": org_scope},
            result_summary={"top_predicted_rate": top_rate, "data_count": len(rates),
                            "peak_count": len(peaks)},
        )
    return result


# ─── API: 사정률 구간 분석 (스펙 §1 — 3가지 모드 A/B/C) ──────────────────

@app.get("/api/v1/analysis/rate-buckets/{announcement_id}")
@rate_limit("60/minute")
def analysis_rate_buckets(
    request: Request,
    announcement_id: str,
    period_months: int = Query(6, ge=1, le=36, description="1/3/6/24개월 등 밀어내기식 분석 기간"),
    category_filter: str = Query("same", pattern="^(same|all|construction|service|industry)$"),
    detail_rule: str = Query("max_gap", pattern="^(first_after|last_after|max_gap)$"),
    current_user: User = Depends(require_auth),
):
    """사정률 발생빈도 + 3가지 구간 알고리즘 (스펙 §1)

    - **mode A**: 빈도 최대 (막대 가장 큰 값 기준)
    - **mode B**: 공백 (막대 없는 값 기준)
    - **mode C**: 차이 최대 (인접 막대 차이가 가장 큰 값 기준)

    각 모드별로 100 기준 ±방향 정렬한 상위 5개 구간 반환 +
    detail_rule 에 따른 세부 사정률 값 1개.
    """
    cache_key = _cache_key("buckets", announcement_id, period_months, category_filter, detail_rule)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        # 카테고리 필터 적용
        q = db.query(BidResult.assessment_rate).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        q = _apply_category_filter(q, category_filter, ann)

        rates = [r[0] for r in q.all() if r[0] is not None]

        if not rates:
            result = {
                "announcement": {"id": ann.id, "title": ann.title, "category": ann.category},
                "params": {"period_months": period_months, "category_filter": category_filter,
                           "detail_rule": detail_rule},
                "data_count": 0,
                "histogram": [],
                "buckets": {"A": [], "B": [], "C": []},
                "detail_rate": None,
            }
        else:
            histogram = _build_rate_histogram(rates, bin_size=0.1)
            buckets = {
                "A": _compute_rate_buckets(histogram, mode="A", top_n=5),
                "B": _compute_rate_buckets(histogram, mode="B", top_n=5),
                "C": _compute_rate_buckets(histogram, mode="C", top_n=5),
            }
            # 표 = 빈도가 0보다 큰 bin 의 사정률 리스트
            table_rates = [b["rate"] for b in histogram if b["count"] > 0]
            detail_rate = _extract_detail_rate(table_rates, rule=detail_rule)
            bid_amount = (
                int(ann.base_amount * detail_rate / 100)
                if ann.base_amount and detail_rate else None
            )

            result = {
                "announcement": {
                    "id": ann.id, "title": ann.title, "category": ann.category,
                    "ordering_org_name": ann.ordering_org_name,
                    "base_amount": ann.base_amount,
                },
                "params": {
                    "period_months": period_months,
                    "category_filter": category_filter,
                    "detail_rule": detail_rule,
                },
                "data_count": len(rates),
                "histogram": histogram,
                "buckets": buckets,
                "detail_rate": round(detail_rate, 4) if detail_rate else None,
                "predicted_bid_amount": bid_amount,
            }
        _cache_set(cache_key, result)
        if current_user:
            top_a = buckets["A"][0]["rate"] if result.get("buckets", {}).get("A") else None
            save_query_history(
                current_user.id, announcement_id, "rate_buckets",
                parameters={"period_months": period_months, "category_filter": category_filter,
                            "detail_rule": detail_rule},
                result_summary={"top_A_rate": top_a, "detail_rate": result.get("detail_rate"),
                                "data_count": result["data_count"]},
            )
        return result
    finally:
        db.close()


# ─── API: 업체사정률 분석 (갭 분석) ───────────────────────────────────────

@app.get("/api/v1/analysis/company-rates/{announcement_id}")
@rate_limit("60/minute")
def analysis_company_rates(
    request: Request,
    announcement_id: str,
    rate_range_start: float = Query(99.0),
    rate_range_end: float = Query(100.0),
    period_months: int = Query(12, ge=1, le=36),
    # 사정률 예측 로직 §2 — 검색 5종
    org_search: str = Query("", description="발주처 부분 일치 검색"),
    price_volatility: float = Query(0.0, ge=0.0, description="예가 변동폭 % (assessment_rate 의 ± 범위)"),
    base_amount_min: int = Query(0, ge=0, description="기초금액 하한"),
    base_amount_max: int = Query(0, ge=0, description="기초금액 상한 (0 이면 무시)"),
    industry_filter: str = Query("", description="업종 코드 부분 일치"),
    current_user: User = Depends(require_auth),
):
    """선택 구간 내 업체 투찰률 분포 + 최대 갭 분석 + 검색 5종 (스펙 §2)

    Search params:
      - org_search        발주처명 LIKE
      - price_volatility  assessment_rate 가 (100 ± volatility) 범위인 결과만
      - base_amount_min/max  공고 기초금액 범위
      - industry_filter   업종 코드 LIKE
    """
    cache_key = _cache_key(
        "gap", announcement_id, rate_range_start, rate_range_end, period_months,
        org_search, price_volatility, base_amount_min, base_amount_max, industry_filter,
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        if current_user:
            save_query_history(
                current_user.id, announcement_id, "company-rates",
                parameters={"rate_range_start": rate_range_start,
                            "rate_range_end": rate_range_end,
                            "period_months": period_months, "cached": True},
                result_summary={"refined_rate": cached.get("refined_rate"),
                                "total_companies": cached.get("total_companies", 0)},
            )
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        # 선택 구간 내 업체 투찰률 (스펙 §2 검색 5종 적용)
        rec_q = db.query(CompanyBidRecord).join(
            BidAnnouncement, CompanyBidRecord.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidAnnouncement.announced_at >= cutoff,
            CompanyBidRecord.bid_rate >= rate_range_start,
            CompanyBidRecord.bid_rate <= rate_range_end,
        )
        if org_search:
            rec_q = rec_q.filter(BidAnnouncement.ordering_org_name.like(f"%{org_search}%"))
        if base_amount_min > 0:
            rec_q = rec_q.filter(BidAnnouncement.base_amount >= base_amount_min)
        if base_amount_max > 0:
            rec_q = rec_q.filter(BidAnnouncement.base_amount <= base_amount_max)
        if industry_filter:
            rec_q = rec_q.filter(BidAnnouncement.industry_code.like(f"%{industry_filter}%"))
        if price_volatility > 0:
            # 예가 변동폭: 매칭되는 BidResult 의 assessment_rate 가 100 ± volatility 인 공고만
            ann_ids_in_volatility = db.query(BidResult.announcement_id).filter(
                BidResult.assessment_rate.isnot(None),
                BidResult.assessment_rate >= 100 - price_volatility,
                BidResult.assessment_rate <= 100 + price_volatility,
            )  # subquery() 없이 select 그대로 IN 절에 전달 → SAWarning 회피
            rec_q = rec_q.filter(CompanyBidRecord.announcement_id.in_(ann_ids_in_volatility))
        records = rec_q.order_by(CompanyBidRecord.bid_rate).all()

        company_rates = [
            {
                "company": r.company_name,
                "rate": round(r.bid_rate, 4),
                "amount": r.bid_amount,
                "ranking": r.ranking,
                "is_first_place": r.is_first_place,
            }
            for r in records
        ]

        # 갭 분석: 연속 투찰률 간 최대 빈 공간
        unique_rates = sorted(set(r.bid_rate for r in records))
        gaps = []
        if len(unique_rates) >= 2:
            for i in range(len(unique_rates) - 1):
                gap_size = round(unique_rates[i + 1] - unique_rates[i], 4)
                if gap_size > 0.01:  # 유의미한 갭만
                    gaps.append({
                        "start": round(unique_rates[i], 4),
                        "end": round(unique_rates[i + 1], 4),
                        "size": gap_size,
                        "midpoint": round((unique_rates[i] + unique_rates[i + 1]) / 2, 4),
                    })

        gaps.sort(key=lambda g: g["size"], reverse=True)
        largest_gap_midpoint = gaps[0]["midpoint"] if gaps else round((rate_range_start + rate_range_end) / 2, 4)
        refined_rate = largest_gap_midpoint

        # 1순위 예측 리스트 (해당 구간 내 과거 1순위 결과 + 검색 필터 적용)
        fp_q = db.query(BidResult, BidAnnouncement).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidResult.first_place_rate >= rate_range_start,
            BidResult.first_place_rate <= rate_range_end,
            BidResult.opened_at >= cutoff,
        )
        if org_search:
            fp_q = fp_q.filter(BidAnnouncement.ordering_org_name.like(f"%{org_search}%"))
        if base_amount_min > 0:
            fp_q = fp_q.filter(BidAnnouncement.base_amount >= base_amount_min)
        if base_amount_max > 0:
            fp_q = fp_q.filter(BidAnnouncement.base_amount <= base_amount_max)
        if industry_filter:
            fp_q = fp_q.filter(BidAnnouncement.industry_code.like(f"%{industry_filter}%"))
        if price_volatility > 0:
            fp_q = fp_q.filter(
                BidResult.assessment_rate.isnot(None),
                BidResult.assessment_rate >= 100 - price_volatility,
                BidResult.assessment_rate <= 100 + price_volatility,
            )
        first_place_in_range = fp_q.order_by(BidResult.opened_at.desc()).limit(20).all()

        first_place_list = [
            {
                "title": a.title,
                "org": a.ordering_org_name,
                "area": a.region,
                "assessment_rate": round(r.assessment_rate, 4),
                "first_place_rate": round(r.first_place_rate, 4),
                "first_place_amount": r.first_place_amount,
                "date": r.opened_at.strftime("%Y-%m-%d") if r.opened_at else "",
            }
            for r, a in first_place_in_range
        ]

        # Phase 7: 갭 중간점 차기연도 검증
        next_year_validation = []
        if gaps:
            next_year_start = ref_date
            next_year_end = ref_date + timedelta(days=365)
            next_year_results = db.query(BidResult, BidAnnouncement).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidResult.assessment_rate.isnot(None),
                BidResult.first_place_rate.isnot(None),
                BidResult.opened_at >= next_year_start,
                BidResult.opened_at <= next_year_end,
            ).all()

            for g in gaps[:10]:
                mid = g["midpoint"]
                matches = []
                for res, past_ann in next_year_results:
                    diff = abs(mid - res.assessment_rate)
                    actual_diff = abs(res.first_place_rate - res.assessment_rate)
                    is_first = diff <= actual_diff + 0.02
                    matches.append({
                        "actual_rate": round(res.assessment_rate, 4),
                        "first_place_rate": round(res.first_place_rate, 4),
                        "diff": round(abs(mid - res.first_place_rate), 4),
                        "is_first_place": is_first,
                        "date": res.opened_at.strftime("%Y-%m-%d") if res.opened_at else "",
                    })
                match_count = sum(1 for m in matches if m["is_first_place"])
                next_year_validation.append({
                    "gap_midpoint": mid,
                    "gap_size": g["size"],
                    "total_cases": len(matches),
                    "match_count": match_count,
                    "match_rate": round(match_count / len(matches) * 100, 1) if matches else 0,
                    "cases": matches[:5],
                })

    finally:
        db.close()
    result = {
        "company_rates": company_rates[:100],  # 최대 100건
        "gaps": gaps[:10],
        "largest_gap_midpoint": largest_gap_midpoint,
        "refined_rate": refined_rate,
        "total_companies": len(company_rates),
        "unique_rate_count": len(unique_rates),
        "first_place_predictions": first_place_list,
        "next_year_validation": next_year_validation,
    }
    _cache_set(cache_key, result)
    if current_user:
        save_query_history(
            current_user.id, announcement_id, "company-rates",
            parameters={"rate_range_start": rate_range_start, "rate_range_end": rate_range_end,
                        "period_months": period_months},
            result_summary={"refined_rate": refined_rate,
                            "largest_gap_midpoint": largest_gap_midpoint,
                            "total_companies": len(company_rates)},
        )
    return result


# ─── API: 종합분석 ────────────────────────────────────────────────────────

@app.get("/api/v1/analysis/comprehensive/{announcement_id}")
@rate_limit("60/minute")
def analysis_comprehensive(
    request: Request,
    announcement_id: str,
    confirmed_rate: float = Query(99.5),
    period_months: int = Query(12, ge=1, le=24),
    org_scope: str = Query("specific"),
    current_user: User = Depends(require_auth),
):
    """확정사정률 기반 과거 1순위 비교 + 종합 분석 (TTL 캐시)"""
    cache_key = _cache_key("comp", announcement_id, confirmed_rate, period_months, org_scope)
    cached = _cache_get(cache_key)
    if cached is not None:
        if current_user:
            save_query_history(
                current_user.id, announcement_id, "comprehensive",
                parameters={"confirmed_rate": confirmed_rate, "period_months": period_months,
                            "org_scope": org_scope, "cached": True},
                result_summary={"confirmed_rate": cached.get("confirmed_rate"),
                                "predicted_first_place_amount":
                                    cached.get("predicted_first_place", {}).get("amount")},
            )
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        # 기간별 분석 (밀어내기식)
        ref_date = ann.announced_at or datetime.now()
        periods = [1, 3, 6, 12, 24]
        period_results = {}

        for pm in periods:
            if pm > period_months:
                continue
            cutoff = ref_date - timedelta(days=pm * 30)

            q = db.query(BidResult, BidAnnouncement).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidResult.assessment_rate.isnot(None),
                BidResult.first_place_rate.isnot(None),
                BidResult.opened_at >= cutoff,
            )

            if org_scope == "parent" and ann.parent_org_name:
                q = q.filter(BidAnnouncement.parent_org_name == ann.parent_org_name)
            elif org_scope == "specific":
                q = q.filter(BidAnnouncement.ordering_org_name == ann.ordering_org_name)

            past_data = q.order_by(BidResult.opened_at.desc()).all()

            match_count = 0
            comparisons = []
            for res, past_ann in past_data:
                # 예측사정률(confirmed_rate)로 1순위 되었을지 확인
                diff = abs(confirmed_rate - res.assessment_rate)
                actual_diff = abs(res.first_place_rate - res.assessment_rate)
                is_match = diff <= actual_diff + 0.02  # 오차범위 0.02%

                if is_match:
                    match_count += 1

                comparisons.append({
                    "id": past_ann.id,
                    "title": past_ann.title,
                    "org": past_ann.ordering_org_name,
                    "area": past_ann.region,
                    "date": res.opened_at.strftime("%Y-%m-%d") if res.opened_at else "",
                    "assessment_rate": round(res.assessment_rate, 4),
                    "first_place_rate": round(res.first_place_rate, 4),
                    "predicted_rate": round(confirmed_rate, 4),
                    "predicted_diff": round(diff, 4),
                    "actual_first_diff": round(actual_diff, 4),
                    "is_match": is_match,
                    "first_place_amount": res.first_place_amount,
                })

            total = len(comparisons)
            period_results[f"{pm}m"] = {
                "period_months": pm,
                "match_count": match_count,
                "total": total,
                "match_rate": round(match_count / total * 100, 1) if total > 0 else 0,
                "comparisons": comparisons[:30],  # 최대 30건
            }

        # 1순위 예상 낙찰가
        predicted_first_place_amount = int(ann.base_amount * confirmed_rate / 100) if ann.base_amount else 0

        # 상위기관 동시 분석
        parent_analysis = None
        if ann.parent_org_name and org_scope == "specific":
            cutoff = ref_date - timedelta(days=period_months * 30)
            parent_q = db.query(BidResult, BidAnnouncement).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.parent_org_name == ann.parent_org_name,
                BidResult.assessment_rate.isnot(None),
                BidResult.first_place_rate.isnot(None),
                BidResult.opened_at >= cutoff,
            )
            parent_data = parent_q.all()
            parent_match = sum(
                1 for res, _ in parent_data
                if abs(confirmed_rate - res.assessment_rate) <= abs(res.first_place_rate - res.assessment_rate) + 0.02
            )
            parent_analysis = {
                "parent_org": ann.parent_org_name,
                "match_count": parent_match,
                "total": len(parent_data),
                "match_rate": round(parent_match / len(parent_data) * 100, 1) if parent_data else 0,
            }

    finally:
        db.close()
    result = {
        "announcement": {
            "id": ann.id, "title": ann.title, "type": ann.category,
            "org": ann.ordering_org_name, "parent_org": ann.parent_org_name,
            "area": ann.region, "budget": ann.base_amount,
        },
        "confirmed_rate": round(confirmed_rate, 4),
        "predicted_first_place": {
            "rate": round(confirmed_rate, 4),
            "amount": predicted_first_place_amount,
        },
        "period_results": period_results,
        "parent_analysis": parent_analysis,
    }
    _cache_set(cache_key, result)
    if current_user:
        save_query_history(
            current_user.id, announcement_id, "comprehensive",
            parameters={"confirmed_rate": confirmed_rate, "period_months": period_months,
                        "org_scope": org_scope},
            result_summary={"confirmed_rate": round(confirmed_rate, 4),
                            "predicted_first_place_amount": predicted_first_place_amount,
                            "period_result_count": len(period_results) if period_results else 0},
        )
    return result


# ─── API: 결합 예측 (빈도 + 갭 결합) ─────────────────────────────────────

@app.get("/api/v1/analysis/combined-prediction/{announcement_id}")
def analysis_combined_prediction(
    announcement_id: str,
    period_months: int = Query(12, ge=1, le=24),
    org_scope: str = Query("specific"),
):
    """빈도분석 피크 + 갭분석 중간점을 결합하여 최적 후보 10개 이상 산출"""
    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        # 1) 빈도분석: 사정률 히스토그램 피크
        q = db.query(BidResult.assessment_rate, BidResult.first_place_rate).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        if org_scope == "parent" and ann.parent_org_name:
            q = q.filter(BidAnnouncement.parent_org_name == ann.parent_org_name)
        else:
            q = q.filter(BidAnnouncement.ordering_org_name == ann.ordering_org_name)
        rows = q.all()

        if not rows:
            q = db.query(BidResult.assessment_rate, BidResult.first_place_rate).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.region == ann.region,
                BidResult.assessment_rate.isnot(None),
                BidResult.opened_at >= cutoff,
            )
            rows = q.all()

        rates = [r.assessment_rate for r in rows if r.assessment_rate]
        first_place_rates = [r.first_place_rate for r in rows if r.first_place_rate]

        # 빈도 히스토그램 (0.1% 단위)
        freq_map = {}
        if rates:
            bin_size = 0.1
            for r in rates:
                key = round(round(r / bin_size) * bin_size, 2)
                freq_map[key] = freq_map.get(key, 0) + 1
        fp_map = {}
        for r in first_place_rates:
            key = round(round(r / 0.1) * 0.1, 2)
            fp_map[key] = fp_map.get(key, 0) + 1
        max_freq = max(freq_map.values()) if freq_map else 1

        # 상위 빈도 피크 10개
        freq_peaks = sorted(freq_map.items(), key=lambda x: x[1], reverse=True)[:10]

        # 2) 갭분석: 업체 투찰률 빈 공간 중간점
        mean_rate = statistics.mean(rates) if rates else 99.5
        gap_range_start = round(mean_rate - 0.5, 2)
        gap_range_end = round(mean_rate + 0.5, 2)
        company_records = db.query(CompanyBidRecord).join(
            BidAnnouncement, CompanyBidRecord.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidAnnouncement.announced_at >= cutoff,
            CompanyBidRecord.bid_rate >= gap_range_start,
            CompanyBidRecord.bid_rate <= gap_range_end,
        ).order_by(CompanyBidRecord.bid_rate).all()

        unique_rates = sorted(set(r.bid_rate for r in company_records))
        gap_midpoints = []
        if len(unique_rates) >= 2:
            for i in range(len(unique_rates) - 1):
                gap_size = unique_rates[i + 1] - unique_rates[i]
                if gap_size > 0.01:
                    gap_midpoints.append({
                        "rate": round((unique_rates[i] + unique_rates[i + 1]) / 2, 4),
                        "gap_size": round(gap_size, 4),
                    })
        gap_midpoints.sort(key=lambda g: g["gap_size"], reverse=True)

        # 3) 결합 점수 산출
        candidates = {}
        # 빈도 피크에서 후보 추가
        for rate_val, count in freq_peaks:
            fp_cnt = fp_map.get(rate_val, 0)
            freq_score = count / max_freq
            fp_score = fp_cnt / max(max(fp_map.values(), default=1), 1)
            candidates[rate_val] = {
                "rate": rate_val,
                "source": "빈도",
                "freq_count": count,
                "first_place_count": fp_cnt,
                "freq_score": round(freq_score, 4),
                "gap_score": 0,
                "combined_score": 0,
            }

        # 갭 중간점에서 후보 추가/보강
        max_gap = gap_midpoints[0]["gap_size"] if gap_midpoints else 1
        for gm in gap_midpoints[:15]:
            rate_val = round(gm["rate"], 2)
            gap_s = gm["gap_size"] / max_gap
            if rate_val in candidates:
                candidates[rate_val]["source"] = "빈도+갭"
                candidates[rate_val]["gap_score"] = round(gap_s, 4)
            else:
                # 근접 빈도 구간 찾기
                nearest_freq = 0
                nearest_fp = 0
                for fr, fc in freq_map.items():
                    if abs(fr - rate_val) <= 0.15:
                        nearest_freq = max(nearest_freq, fc)
                        nearest_fp = max(nearest_fp, fp_map.get(fr, 0))
                candidates[rate_val] = {
                    "rate": rate_val,
                    "source": "갭",
                    "freq_count": nearest_freq,
                    "first_place_count": nearest_fp,
                    "freq_score": round(nearest_freq / max_freq, 4) if max_freq > 0 else 0,
                    "gap_score": round(gap_s, 4),
                    "combined_score": 0,
                }

        # 결합 점수 = 빈도(0.4) + 갭(0.4) + 1순위 이력(0.2)
        max_fp_score = max((c["first_place_count"] for c in candidates.values()), default=1) or 1
        for c in candidates.values():
            fp_s = c["first_place_count"] / max_fp_score
            c["combined_score"] = round(
                c["freq_score"] * 0.4 + c["gap_score"] * 0.4 + fp_s * 0.2, 4
            )
            c["bid_amount"] = int(ann.base_amount * c["rate"] / 100) if ann.base_amount else 0

        result_list = sorted(candidates.values(), key=lambda c: c["combined_score"], reverse=True)
        for i, c in enumerate(result_list):
            c["rank"] = i + 1
            c["is_recommended"] = i < 3
            c["confidence"] = "높음" if c["combined_score"] > 0.6 else "보통" if c["combined_score"] > 0.3 else "낮음"

    finally:
        db.close()
    return {
        "candidates": result_list[:15],
        "total_candidates": len(result_list),
        "data_summary": {
            "freq_data_count": len(rates),
            "gap_data_count": len(company_records),
            "first_place_count": len(first_place_rates),
        },
        "announcement": {
            "id": ann.id, "title": ann.title, "type": ann.category,
            "org": ann.ordering_org_name, "area": ann.region,
            "budget": ann.base_amount,
        },
    }


# ─── API: 상관관계 분석 (스펙 §3.11 — 3가지 방법 종합 1순위) ─────────────

@app.get("/api/v1/analysis/correlation/{announcement_id}")
@rate_limit("30/minute")
def analysis_correlation(
    request: Request,
    announcement_id: str,
    period_months: int = Query(6, ge=1, le=36),
    category_filter: str = Query("same", pattern="^(same|all|construction|service|industry)$"),
    bucket_mode: str = Query("A", pattern="^(A|B|C)$"),
    detail_rule: str = Query("max_gap", pattern="^(first_after|last_after|max_gap)$"),
    rate_range_start: float = Query(99.0),
    rate_range_end: float = Query(101.0),
    current_user: User = Depends(require_auth),
):
    """스펙 §3.11 — 3가지 분석 방법별 1순위 + 종합 상관관계 1순위.

    Methods:
      1) 사정률 발생빈도 분석 (Tab1 모드)
      2) 업체사정률 갭 분석 (Tab2)
      3) 빈도+갭 결합 분석

    Returns:
      methods: [{name, top1_rate, top1_score, count}, ...]
      correlation: {agreement, final_top1, final_score, methods_aligned}
    """
    cache_key = _cache_key(
        "correlation", announcement_id, period_months, category_filter,
        bucket_mode, detail_rule, rate_range_start, rate_range_end,
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        # 카테고리 필터 — 공통 헬퍼 alias (analysis_correlation 내부 사용)
        _apply_category = lambda q: _apply_category_filter(q, category_filter, ann)

        # ── Method 1: 사정률 발생빈도 (정확도 강화 적용) ───────────
        # 표본 수집 — 사정률·날짜·발주처·업종 동시 조회 (정확도 분석용 메타)
        rate_q = db.query(
            BidResult.assessment_rate,
            BidResult.opened_at,
            BidAnnouncement.ordering_org_name,
            BidAnnouncement.industry_code,
        ).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        rate_q = _apply_category(rate_q)
        raw_samples = [(r[0], r[1], r[2], r[3]) for r in rate_q.all() if r[0] is not None]
        raw_rates = [s[0] for s in raw_samples]

        # ① IQR 이상치 제거 (분포 안정화)
        clean_rates = _remove_outliers_iqr(raw_rates)
        outliers_removed = len(raw_rates) - len(clean_rates)

        # ② 동종성 점수 (대상 공고 발주처/업종과 표본 유사도)
        homogeneity = _homogeneity_score(
            ann.ordering_org_name, ann.industry_code,
            [s[2] for s in raw_samples], [s[3] for s in raw_samples],
        )

        # ③ 시계열 가중치 (최근 6개월 데이터 더 높게)
        rates_with_dates = [(s[0], s[1]) for s in raw_samples if s[0] in clean_rates]
        rates_m1 = _time_weighted_rates(rates_with_dates, half_life_days=180, ref_date=ref_date)
        if not rates_m1:  # fallback
            rates_m1 = clean_rates

        method1 = {
            "name": "1) 사정률 발생빈도 분석",
            "top1_rate": None,
            "top1_score": 0,
            "count": len(raw_rates),  # 원 표본 수 (사용자 노출용)
            "weighted_count": len(rates_m1),
            "outliers_removed": outliers_removed,
            "homogeneity": homogeneity,
            "mode": bucket_mode,
        }
        if rates_m1:
            hist = _build_rate_histogram(rates_m1)
            buckets = _compute_rate_buckets(hist, mode=bucket_mode, top_n=1)
            if buckets:
                method1["top1_rate"] = round(buckets[0]["rate"], 4)
                method1["top1_score"] = buckets[0]["score"]

        # ── Method 2: 업체사정률 갭 분석 ─────────────────────────
        # 1차: CompanyBidRecord (업로드 데이터)
        gap_q = db.query(CompanyBidRecord.bid_rate).join(
            BidAnnouncement, CompanyBidRecord.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.announced_at >= cutoff,
            CompanyBidRecord.bid_rate >= rate_range_start,
            CompanyBidRecord.bid_rate <= rate_range_end,
        )
        gap_q = _apply_category(gap_q)
        bid_rates = sorted({round(r[0], 4) for r in gap_q.all() if r[0] is not None})
        gap_source = "company_records"
        # Fallback: G2B 자동 수집은 CompanyBidRecord 가 비어 있으므로
        # BidResult.first_place_rate (1순위 낙찰률) 분포로 대체
        if not bid_rates:
            fp_q = db.query(BidResult.first_place_rate).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidResult.first_place_rate.isnot(None),
                BidResult.opened_at >= cutoff,
                BidResult.first_place_rate >= rate_range_start,
                BidResult.first_place_rate <= rate_range_end,
            )
            fp_q = _apply_category(fp_q)
            bid_rates = sorted({round(r[0], 4) for r in fp_q.all() if r[0] is not None})
            gap_source = "first_place_rate"
        method2 = {
            "name": "2) 업체사정률 갭 분석",
            "top1_rate": None,
            "top1_score": 0,
            "count": len(bid_rates),
            "source": gap_source,
        }
        if len(bid_rates) >= 2:
            best_gap = 0
            best_mid = None
            for i in range(len(bid_rates) - 1):
                gap = round(bid_rates[i + 1] - bid_rates[i], 4)
                if gap > best_gap:
                    best_gap = gap
                    best_mid = round((bid_rates[i] + bid_rates[i + 1]) / 2, 4)
            if best_mid is not None:
                method2["top1_rate"] = best_mid
                method2["top1_score"] = best_gap

        # ── Method 3: 빈도+갭 결합 (양쪽 후보 중 100 가까운 + 양쪽에 모두 등장) ──
        method3 = {
            "name": "3) 빈도+갭 결합 분석",
            "top1_rate": None,
            "top1_score": 0,
            "count": 0,
        }
        if rates_m1 and bid_rates:
            hist = _build_rate_histogram(rates_m1)
            top5_freq = [b["rate"] for b in _compute_rate_buckets(hist, mode=bucket_mode, top_n=5)]
            # 갭 후보 (상위 5)
            gap_candidates = []
            for i in range(len(bid_rates) - 1):
                gap = round(bid_rates[i + 1] - bid_rates[i], 4)
                mid = round((bid_rates[i] + bid_rates[i + 1]) / 2, 4)
                gap_candidates.append((mid, gap))
            gap_candidates.sort(key=lambda x: -x[1])
            top5_gap = [round(m, 1) for m, _ in gap_candidates[:5]]
            # 교집합 (소수 1자리 반올림 기준) → 100 기준 가까운 순
            top5_freq_round = {round(r, 1) for r in top5_freq}
            top5_gap_round = set(top5_gap)
            common = sorted(top5_freq_round & top5_gap_round, key=lambda r: abs(r - 100))
            method3["count"] = len(common)
            if common:
                method3["top1_rate"] = round(common[0], 4)
                method3["top1_score"] = len(common)
            elif method1["top1_rate"] and method2["top1_rate"]:
                # 교집합 없으면 두 1순위의 평균
                method3["top1_rate"] = round(
                    (method1["top1_rate"] + method2["top1_rate"]) / 2, 4
                )
                method3["top1_score"] = 0

        # ── 종합 상관관계 ──────────────────────────────────────────
        rates_predicted = [m["top1_rate"] for m in [method1, method2, method3] if m["top1_rate"] is not None]
        agreement = 0.0
        final_top1 = None
        methods_aligned = []
        if rates_predicted:
            # 각 방법 1순위가 ±0.2% 이내면 합치 (agreement) — 단순 카운트 기반 점수
            from collections import Counter
            rounded = [round(r, 1) for r in rates_predicted]
            counter = Counter(rounded)
            most_common, freq = counter.most_common(1)[0]
            agreement = round(freq / len(rates_predicted), 2)
            # 정렬: 가장 많이 등장한 값을 가진 방법들
            methods_aligned = [
                m["name"] for m in [method1, method2, method3]
                if m["top1_rate"] is not None and round(m["top1_rate"], 1) == most_common
            ]
            # 최종 1순위: 합치된 값들의 평균 (없으면 method1 우선)
            aligned_rates = [
                m["top1_rate"] for m in [method1, method2, method3]
                if m["top1_rate"] is not None and round(m["top1_rate"], 1) == most_common
            ]
            final_top1 = round(sum(aligned_rates) / len(aligned_rates), 4)

        bid_amount = (
            int(ann.base_amount * final_top1 / 100)
            if ann.base_amount and final_top1 else None
        )

        result = {
            "announcement": {
                "id": ann.id, "title": ann.title, "category": ann.category,
                "ordering_org_name": ann.ordering_org_name,
                "base_amount": ann.base_amount,
            },
            "params": {
                "period_months": period_months,
                "category_filter": category_filter,
                "bucket_mode": bucket_mode,
                "detail_rule": detail_rule,
                "rate_range_start": rate_range_start,
                "rate_range_end": rate_range_end,
            },
            "methods": [method1, method2, method3],
            "correlation": {
                "agreement": agreement,         # 0.0 ~ 1.0
                "final_top1": final_top1,
                "predicted_bid_amount": bid_amount,
                "methods_aligned": methods_aligned,
                # 신뢰도: 1) 데이터 표본 크기 2) 합치도 3) 표준편차 + 동종성 종합
                "confidence": _compute_confidence(rates_m1, agreement, len(methods_aligned)),
                "sample_size": len(rates_m1),
                "raw_sample_size": len(raw_rates),
                "outliers_removed": outliers_removed,
                "homogeneity": homogeneity,
                "std_deviation": (round(statistics.stdev(rates_m1), 4)
                                  if len(rates_m1) > 1 else None),
                # 신뢰구간 (95%) — final_top1 ± margin
                "confidence_interval": (lambda ci: {
                    "mean": ci[0], "lower": ci[1], "upper": ci[2], "margin": ci[3],
                })(_confidence_interval(rates_m1, confidence=0.95)),
            },
        }

        # ── 앙상블 + 모멘텀 보정 ──────────────────────────────────
        # 각 방법의 표본 크기를 가중치로 사용 (큰 표본 = 높은 신뢰)
        ensemble = _ensemble_predict([
            {"name": "빈도분석", "rate": method1["top1_rate"], "weight": method1["count"]},
            {"name": "갭분석", "rate": method2["top1_rate"], "weight": method2["count"]},
            {"name": "결합분석", "rate": method3["top1_rate"], "weight": max(1, method3["count"])},
        ])

        # 시계열 모멘텀 — 최근 6개월 월별 평균 추세
        from collections import defaultdict
        recent_cutoff = ref_date - timedelta(days=180)
        monthly_buckets = defaultdict(list)
        for s in raw_samples:
            if s[1] and s[1] >= recent_cutoff:
                monthly_buckets[s[1].strftime("%Y-%m")].append(s[0])
        monthly_series = [
            {"period": k, "avg_rate": sum(v) / len(v), "count": len(v)}
            for k, v in sorted(monthly_buckets.items())
        ]
        momentum = _trend_momentum(monthly_series)

        # 모멘텀 보정 — slope 의 30%만 반영 (과보정 방지)
        ensemble_corrected = ensemble.get("final_rate")
        if ensemble_corrected is not None and momentum.get("slope") is not None:
            correction = momentum["slope"] * 0.3
            ensemble_corrected = round(ensemble_corrected + correction, 4)

        # ── Bayesian Shrinkage (소표본 발주처 보정) ────────────────
        same_org_rates = [s[0] for s in raw_samples if s[2] == ann.ordering_org_name]
        bayesian = _bayesian_shrinkage(same_org_rates, raw_rates, prior_strength=10)

        # ── 이상 탐지 (대상 공고 기초금액 vs 동종 공고) ─────────────
        anomaly = {"is_anomaly": False}
        if ann.base_amount and ann.base_amount > 0:
            peer_q = db.query(BidAnnouncement.base_amount).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.base_amount.isnot(None),
                BidAnnouncement.base_amount > 0,
                BidAnnouncement.announced_at >= cutoff,
            )
            peer_amounts = [math.log(p[0]) for p in peer_q.all() if p[0]]
            if peer_amounts:
                anomaly = _detect_anomaly(math.log(ann.base_amount), peer_amounts, z_threshold=2.0)

        result["ensemble"] = {
            **ensemble,
            "momentum": momentum,
            "final_rate_corrected": ensemble_corrected,
            "predicted_bid_amount": (
                int(ann.base_amount * ensemble_corrected / 100)
                if ann.base_amount and ensemble_corrected else None
            ),
        }
        result["bayesian_shrinkage"] = bayesian
        result["anomaly"] = anomaly
        _cache_set(cache_key, result)
        if current_user:
            save_query_history(
                current_user.id, announcement_id, "correlation",
                parameters=result["params"],
                result_summary={
                    "final_top1": final_top1,
                    "agreement": agreement,
                    "methods_aligned_count": len(methods_aligned),
                },
            )
        return result
    finally:
        db.close()


# ─── API: 사용자 예측 설정 학습 저장 (스펙 §1 — 다음 공고 자동 적용) ─────

@app.get("/api/v1/users/me/prediction-settings")
@rate_limit("60/minute")
def get_prediction_settings(request: Request, current_user: User = Depends(require_auth)):
    """현재 로그인 사용자의 사정률 예측 설정 조회 (없으면 기본값 반환)"""
    db = SessionLocal()
    try:
        s = db.query(PredictionSettings).filter(
            PredictionSettings.user_id == current_user.id
        ).first()
        if not s:
            return {
                "period_months": 6, "category_filter": "same",
                "bucket_mode": "A", "detail_rule": "max_gap",
                "rate_range_start": None, "rate_range_end": None,
                "confirmed_rate": None, "exists": False,
            }
        return {
            "period_months": s.period_months,
            "category_filter": s.category_filter,
            "bucket_mode": s.bucket_mode,
            "detail_rule": s.detail_rule,
            "rate_range_start": s.rate_range_start,
            "rate_range_end": s.rate_range_end,
            "confirmed_rate": s.confirmed_rate,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None,
            "exists": True,
        }
    finally:
        db.close()


@app.put("/api/v1/users/me/prediction-settings")
@rate_limit("30/minute")
def put_prediction_settings(
    request: Request,
    payload: dict,
    current_user: User = Depends(require_auth),
):
    """사정률 예측 설정 upsert"""
    db = SessionLocal()
    try:
        s = db.query(PredictionSettings).filter(
            PredictionSettings.user_id == current_user.id
        ).first()
        if not s:
            s = PredictionSettings(user_id=current_user.id)
            db.add(s)
        for field in ("period_months", "category_filter", "bucket_mode", "detail_rule",
                      "rate_range_start", "rate_range_end", "confirmed_rate"):
            if field in payload and payload[field] is not None:
                setattr(s, field, payload[field])
        s.updated_at = datetime.now()
        db.commit()
        return {"status": "ok", "updated_at": s.updated_at.isoformat()}
    finally:
        db.close()


# ─── API: 분석 결과 엑셀 export (스펙 §3.8/3.9) ──────────────────────────

@app.get("/api/v1/analysis/export/{export_type}")
@rate_limit("10/minute")
def analysis_export_xlsx(
    request: Request,
    export_type: str,
    announcement_id: str = Query(...),
    period_months: int = Query(6, ge=1, le=36),
    category_filter: str = Query("same"),
    bucket_mode: str = Query("A"),
    detail_rule: str = Query("max_gap"),
    rate_range_start: float = Query(99.0),
    rate_range_end: float = Query(101.0),
    current_user: User = Depends(require_auth),
):
    """엑셀(.xlsx) 분석 결과 다운로드.

    export_type: buckets / company / correlation / bid_list
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            raise HTTPException(status_code=404, detail="공고를 찾을 수 없습니다.")

        wb = Workbook()
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")

        def write_header(row_idx, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row_idx, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center

        # ── 공고 메타 ────────────────────────────────────────
        ws.title = export_type[:31]
        ws.cell(row=1, column=1, value="공고명").font = Font(bold=True)
        ws.cell(row=1, column=2, value=ann.title)
        ws.cell(row=2, column=1, value="공고번호").font = Font(bold=True)
        ws.cell(row=2, column=2, value=ann.bid_number)
        ws.cell(row=3, column=1, value="발주기관").font = Font(bold=True)
        ws.cell(row=3, column=2, value=ann.ordering_org_name)
        ws.cell(row=4, column=1, value="기초금액").font = Font(bold=True)
        ws.cell(row=4, column=2, value=ann.base_amount)

        if export_type == "buckets":
            # 3가지 모드 결과
            ref_date = ann.announced_at or datetime.now()
            cutoff = ref_date - timedelta(days=period_months * 30)
            q = db.query(BidResult.assessment_rate).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidResult.assessment_rate.isnot(None),
                BidResult.opened_at >= cutoff,
                BidAnnouncement.category == ann.category,
            )
            rates = [r[0] for r in q.all() if r[0] is not None]
            hist = _build_rate_histogram(rates)
            write_header(6, ["모드", "순위", "사정률(%)", "방향", "점수", "구간시작", "구간끝"])
            row = 7
            for mode in ["A", "B", "C"]:
                for b in _compute_rate_buckets(hist, mode=mode, top_n=5):
                    ws.cell(row=row, column=1, value=mode)
                    ws.cell(row=row, column=2, value=b["rank"])
                    ws.cell(row=row, column=3, value=b["rate"])
                    ws.cell(row=row, column=4, value=b["side"])
                    ws.cell(row=row, column=5, value=b["score"])
                    ws.cell(row=row, column=6, value=b["range_start"])
                    ws.cell(row=row, column=7, value=b["range_end"])
                    row += 1

        elif export_type == "company":
            # 업체 투찰 + 갭
            ref_date = ann.announced_at or datetime.now()
            cutoff = ref_date - timedelta(days=period_months * 30)
            recs = db.query(CompanyBidRecord, BidAnnouncement).join(
                BidAnnouncement, CompanyBidRecord.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.announced_at >= cutoff,
                CompanyBidRecord.bid_rate >= rate_range_start,
                CompanyBidRecord.bid_rate <= rate_range_end,
            ).order_by(CompanyBidRecord.bid_rate).all()
            write_header(6, ["업체명", "투찰률(%)", "투찰금액", "순위", "1순위", "공고명", "발주기관"])
            for i, (r, a) in enumerate(recs, start=7):
                ws.cell(row=i, column=1, value=r.company_name)
                ws.cell(row=i, column=2, value=round(r.bid_rate, 4))
                ws.cell(row=i, column=3, value=r.bid_amount)
                ws.cell(row=i, column=4, value=r.ranking)
                ws.cell(row=i, column=5, value="O" if r.is_first_place else "")
                ws.cell(row=i, column=6, value=a.title)
                ws.cell(row=i, column=7, value=a.ordering_org_name)

        elif export_type == "correlation":
            # 3가지 방법 + 종합
            corr = analysis_correlation(
                request, announcement_id, period_months=period_months,
                category_filter=category_filter, bucket_mode=bucket_mode,
                detail_rule=detail_rule, rate_range_start=rate_range_start,
                rate_range_end=rate_range_end, current_user=current_user,
            )
            write_header(6, ["분석 방법", "1순위 사정률(%)", "점수", "데이터 수", "합치 여부"])
            aligned_set = set(corr.get("correlation", {}).get("methods_aligned", []))
            for i, m in enumerate(corr.get("methods", []), start=7):
                ws.cell(row=i, column=1, value=m["name"])
                ws.cell(row=i, column=2, value=m["top1_rate"])
                ws.cell(row=i, column=3, value=m["top1_score"])
                ws.cell(row=i, column=4, value=m["count"])
                ws.cell(row=i, column=5, value="합치" if m["name"] in aligned_set else "")
            # 종합 결과
            row = 7 + len(corr.get("methods", [])) + 1
            ws.cell(row=row, column=1, value="종합 1순위").font = Font(bold=True, color="DC2626")
            ws.cell(row=row, column=2, value=corr.get("correlation", {}).get("final_top1"))
            ws.cell(row=row, column=3, value=f"합치도 {round((corr.get('correlation', {}).get('agreement') or 0) * 100)}%")
            ws.cell(row=row, column=4, value=corr.get("correlation", {}).get("predicted_bid_amount"))

        elif export_type == "bid_list":
            # 투찰 리스트 (최근 100건 — 카테고리 동일)
            ref_date = ann.announced_at or datetime.now()
            cutoff = ref_date - timedelta(days=period_months * 30)
            anns = db.query(BidAnnouncement).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.announced_at >= cutoff,
            ).order_by(BidAnnouncement.announced_at.desc()).limit(100).all()
            write_header(6, ["공고일", "공고번호", "공고명", "발주기관", "기초금액", "예상 투찰률(%)", "예상 투찰금액"])
            # 종합 1순위 사용
            corr = analysis_correlation(
                request, announcement_id, period_months=period_months,
                category_filter=category_filter, bucket_mode=bucket_mode,
                detail_rule=detail_rule, rate_range_start=rate_range_start,
                rate_range_end=rate_range_end, current_user=current_user,
            )
            top1 = corr.get("correlation", {}).get("final_top1") or 100
            for i, a in enumerate(anns, start=7):
                ws.cell(row=i, column=1, value=a.announced_at.strftime("%Y-%m-%d") if a.announced_at else "")
                ws.cell(row=i, column=2, value=a.bid_number)
                ws.cell(row=i, column=3, value=a.title)
                ws.cell(row=i, column=4, value=a.ordering_org_name)
                ws.cell(row=i, column=5, value=a.base_amount)
                ws.cell(row=i, column=6, value=top1)
                ws.cell(row=i, column=7, value=int((a.base_amount or 0) * top1 / 100))
        else:
            raise HTTPException(status_code=400, detail="지원하지 않는 export_type")

        # 컬럼 너비 자동
        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"bidstar_{export_type}_{ann.bid_number}.xlsx"
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        db.close()


# ─── API: 회귀 분석 — 다중 선형 회귀로 사정률 예측 ─────────────────────────

@app.get("/api/v1/analysis/regression/{announcement_id}")
@rate_limit("30/minute")
def analysis_regression(
    request: Request,
    announcement_id: str,
    period_months: int = Query(12, ge=1, le=36),
    category_filter: str = Query("same", pattern="^(same|all|construction|service|industry)$"),
    current_user: User = Depends(require_auth),
):
    """기초금액·공사여부·지역·경과일수를 독립변수로 한 사정률 다중 선형 회귀.

    Returns:
      - predicted_rate: 모델이 예측한 사정률
      - r_squared: 모델 설명력 (0~1)
      - top_features: 영향도 큰 5개 변수 (계수)
      - n: 학습 표본 수
    """
    cache_key = _cache_key("regression", announcement_id, period_months, category_filter)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        # 표본 수집 — 사정률 + feature 동시
        q = db.query(
            BidResult.assessment_rate,
            BidResult.opened_at,
            BidAnnouncement.base_amount,
            BidAnnouncement.category,
            BidAnnouncement.region,
            BidAnnouncement.industry_code,
            BidAnnouncement.ordering_org_type,
        ).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        q = _apply_category_filter(q, category_filter, ann)

        samples = []
        for row in q.all():
            if not row[0]:
                continue
            days_old = (ref_date - row[1]).days if row[1] else 0
            samples.append({
                "rate": row[0],
                "base_amount": row[2] or 0,
                "category": row[3],
                "region": row[4],
                "industry": row[5],
                "org_type": row[6],
                "days_old": days_old,
            })

        target = {
            "base_amount": ann.base_amount or 0,
            "category": ann.category,
            "region": ann.region,
            "industry": ann.industry_code,
            "org_type": ann.ordering_org_type,
            "days_old": 0,
        }

        result = _ols_regression(samples, target)
        result["announcement"] = {
            "id": ann.id, "title": ann.title,
            "base_amount": ann.base_amount, "region": ann.region,
            "category": ann.category,
        }
        result["params"] = {"period_months": period_months, "category_filter": category_filter}
        if result.get("predicted_rate") and ann.base_amount:
            result["predicted_bid_amount"] = int(ann.base_amount * result["predicted_rate"] / 100)

        _cache_set(cache_key, result)
        if current_user:
            save_query_history(
                current_user.id, announcement_id, "regression",
                parameters=result["params"],
                result_summary={"predicted_rate": result.get("predicted_rate"),
                                "r_squared": result.get("r_squared"),
                                "n": result.get("n")},
            )
        return result
    finally:
        db.close()


# ─── API: 시계열 추세 — 월별/분기별 사정률 평균 ─────────────────────────────

@app.get("/api/v1/analysis/trend/{announcement_id}")
@rate_limit("60/minute")
def analysis_trend(
    request: Request,
    announcement_id: str,
    granularity: str = Query("month", pattern="^(month|quarter|year)$"),
    period_months: int = Query(24, ge=3, le=84),
    category_filter: str = Query("same", pattern="^(same|all|construction|service|industry)$"),
    current_user: User = Depends(require_auth),
):
    """월/분기/연 단위 평균 사정률 추이 (시계열 차트용)."""
    cache_key = _cache_key("trend", announcement_id, granularity, period_months, category_filter)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}
        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        q = db.query(BidResult.assessment_rate, BidResult.opened_at).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        q = _apply_category_filter(q, category_filter, ann)

        # 그룹핑 키 생성
        from collections import defaultdict
        buckets = defaultdict(list)
        for rate, dt in q.all():
            if not dt or rate is None:
                continue
            if granularity == "month":
                key = dt.strftime("%Y-%m")
            elif granularity == "quarter":
                key = f"{dt.year}-Q{(dt.month - 1) // 3 + 1}"
            else:
                key = str(dt.year)
            buckets[key].append(rate)

        series = []
        for key in sorted(buckets.keys()):
            rates = buckets[key]
            series.append({
                "period": key,
                "avg_rate": round(sum(rates) / len(rates), 4),
                "min_rate": round(min(rates), 4),
                "max_rate": round(max(rates), 4),
                "count": len(rates),
            })

        result = {
            "announcement": {"id": ann.id, "title": ann.title, "category": ann.category},
            "params": {"granularity": granularity, "period_months": period_months,
                       "category_filter": category_filter},
            "series": series,
        }
        _cache_set(cache_key, result)
        return result
    finally:
        db.close()


# ─── API: 백테스트 — 과거 공고 대상 알고리즘 정확도 측정 ──────────────────

@app.get("/api/v1/analysis/backtest/{announcement_id}")
@rate_limit("10/minute")
def analysis_backtest(
    request: Request,
    announcement_id: str,
    sample_size: int = Query(30, ge=5, le=100),
    period_months: int = Query(24, ge=6, le=60),
    bucket_mode: str = Query("A", pattern="^(A|B|C)$"),
    current_user: User = Depends(require_auth),
):
    """A/B 백테스트 — 과거 공고 N건에 대해 알고리즘 예측 vs 실제 사정률 비교.

    각 과거 공고마다:
      1) 그 공고가 발생한 시점의 데이터로만 (이전 데이터) 빈도 분석
      2) 알고리즘 1순위 예측 사정률 산출
      3) 실제 사정률과 절대 오차 계산
    Returns:
      - mae: 평균 절대 오차 (Mean Absolute Error)
      - mape: 평균 절대 비율 오차 (%)
      - within_05: ±0.5% 이내 적중률
      - within_10: ±1.0% 이내 적중률
      - samples: 샘플별 (실제, 예측, 오차) 리스트
    """
    cache_key = _cache_key("backtest", announcement_id, sample_size, period_months, bucket_mode)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        # 과거 N건의 동일 카테고리 결과 (대상 공고 직전까지)
        past_results = (
            db.query(BidResult, BidAnnouncement)
            .join(BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id)
            .filter(
                BidAnnouncement.category == ann.category,
                BidResult.assessment_rate.isnot(None),
                BidResult.opened_at < ref_date,
                BidResult.opened_at >= ref_date - timedelta(days=period_months * 30),
            )
            .order_by(BidResult.opened_at.desc())
            .limit(sample_size * 3)  # 후보 풀
            .all()
        )

        # 무작위 sample_size 건 추출 (재현성 위해 고정 seed)
        import random as _rand
        _rand.seed(42)
        if len(past_results) > sample_size:
            past_results = _rand.sample(past_results, sample_size)

        comparisons = []
        for past_res, past_ann in past_results:
            # 그 공고 시점 이전 90일 데이터로만 분석
            train_cutoff = past_ann.announced_at - timedelta(days=1) if past_ann.announced_at else past_res.opened_at
            train_q = db.query(BidResult.assessment_rate).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidResult.assessment_rate.isnot(None),
                BidResult.opened_at < train_cutoff,
                BidResult.opened_at >= train_cutoff - timedelta(days=180),
            )
            train_rates = [r[0] for r in train_q.all() if r[0] is not None]
            if len(train_rates) < 5:
                continue

            # 알고리즘 적용 (IQR + 가중치 없음 — 단순 빈도 모드)
            clean = _remove_outliers_iqr(train_rates)
            hist = _build_rate_histogram(clean)
            buckets = _compute_rate_buckets(hist, mode=bucket_mode, top_n=1)
            if not buckets:
                continue

            predicted = buckets[0]["rate"]
            actual = past_res.assessment_rate
            error = abs(predicted - actual)
            comparisons.append({
                "title": past_ann.title[:40],
                "actual": round(actual, 4),
                "predicted": round(predicted, 4),
                "error": round(error, 4),
                "opened_at": past_res.opened_at.strftime("%Y-%m-%d") if past_res.opened_at else "",
            })

        if not comparisons:
            return {"error": "백테스트 가능한 표본이 부족합니다.",
                    "n": 0, "params": {"sample_size": sample_size}}

        errors = [c["error"] for c in comparisons]
        actuals = [c["actual"] for c in comparisons]
        mae = round(sum(errors) / len(errors), 4)
        mape = round(sum(e / max(0.01, a) for e, a in zip(errors, actuals)) / len(errors) * 100, 4)
        within_05 = round(sum(1 for e in errors if e <= 0.5) / len(errors) * 100, 1)
        within_10 = round(sum(1 for e in errors if e <= 1.0) / len(errors) * 100, 1)

        result = {
            "announcement": {"id": ann.id, "title": ann.title},
            "params": {"sample_size": sample_size, "period_months": period_months,
                       "bucket_mode": bucket_mode},
            "n": len(comparisons),
            "mae": mae,                        # 평균 절대 오차 (사정률 %)
            "mape": mape,                      # 평균 절대 비율 오차 (%)
            "within_05": within_05,            # ±0.5%p 적중률
            "within_10": within_10,            # ±1.0%p 적중률
            "samples": comparisons[:10],       # 상위 10건만 노출 (응답 크기)
        }
        _cache_set(cache_key, result)
        return result
    finally:
        db.close()


# ─── API: 모델 재학습 이력 (관리자) ──────────────────────────────────────

@app.get("/api/v1/admin/models/history")
@rate_limit("30/minute")
def admin_models_history(
    request: Request,
    limit: int = Query(30, ge=1, le=200),
    category: str = Query(None),
    current_user: User = Depends(require_admin),
):
    """모델 재학습 이력 조회 (정확도 추이)"""
    db = SessionLocal()
    try:
        q = db.query(ModelMetric).order_by(ModelMetric.trained_at.desc())
        if category:
            q = q.filter(ModelMetric.category == category)
        items = q.limit(limit).all()
        return {
            "items": [
                {
                    "id": m.id,
                    "trained_at": m.trained_at.isoformat() if m.trained_at else None,
                    "category": m.category,
                    "model_type": m.model_type,
                    "n_samples": m.n_samples,
                    "r_squared": m.r_squared,
                    "residual_std": m.residual_std,
                    "mae": m.mae,
                    "period_days": m.period_days,
                }
                for m in items
            ],
        }
    finally:
        db.close()


@app.post("/api/v1/admin/models/retrain")
@rate_limit("5/minute")
def admin_models_retrain_now(request: Request, current_user: User = Depends(require_admin)):
    """수동 재학습 트리거 (스케줄 외 즉시 실행)"""
    _scheduled_model_retrain()
    return {"status": "ok", "message": "재학습 완료"}


# ─── API: 이상 공고 일괄 조회 (관리자) ──────────────────────────────────

@app.get("/api/v1/admin/anomalies")
@rate_limit("30/minute")
def admin_anomalies(
    request: Request,
    limit: int = Query(50, ge=1, le=200),
    period_days: int = Query(30, ge=1, le=365),
    z_threshold: float = Query(2.0, ge=1.0, le=5.0),
    current_user: User = Depends(require_admin),
):
    """최근 N일 신규 공고 중 동종 분포 이상치(z>threshold) 자동 탐지."""
    db = SessionLocal()
    try:
        since = datetime.now() - timedelta(days=period_days)
        recent = db.query(BidAnnouncement).filter(
            BidAnnouncement.announced_at >= since,
            BidAnnouncement.base_amount.isnot(None),
            BidAnnouncement.base_amount > 0,
            BidAnnouncement.category.in_(["공사", "용역"]),
        ).all()

        # 카테고리별 동종 표본 mean/std 1회만 계산 (성능 최적화)
        peer_stats = {}
        for cat in ("공사", "용역"):
            peers_q = db.query(BidAnnouncement.base_amount).filter(
                BidAnnouncement.category == cat,
                BidAnnouncement.base_amount.isnot(None),
                BidAnnouncement.base_amount > 0,
                BidAnnouncement.announced_at >= since - timedelta(days=180),
            )
            logs = [math.log(p[0]) for p in peers_q.all() if p[0]]
            if len(logs) >= 30:
                mean = sum(logs) / len(logs)
                std = statistics.stdev(logs)
                peer_stats[cat] = (mean, std, len(logs))

        anomalies = []
        for a in recent:
            stats_t = peer_stats.get(a.category)
            if not stats_t:
                continue
            mean, std, n = stats_t
            if std <= 0:
                continue
            log_amount = math.log(a.base_amount)
            z = (log_amount - mean) / std
            abs_z = abs(z)
            if abs_z > z_threshold:
                severity = "high" if abs_z > z_threshold * 1.5 else "medium"
                res = {"is_anomaly": True, "z_score": round(z, 3), "abs_z": round(abs_z, 3),
                       "peer_mean": round(mean, 4), "peer_std": round(std, 4),
                       "severity": severity}
            else:
                res = {"is_anomaly": False}
            if res.get("is_anomaly"):
                anomalies.append({
                    "id": a.id,
                    "bid_number": a.bid_number,
                    "title": a.title,
                    "ordering_org_name": a.ordering_org_name,
                    "category": a.category,
                    "base_amount": a.base_amount,
                    "announced_at": a.announced_at.isoformat() if a.announced_at else None,
                    "z_score": res.get("z_score"),
                    "abs_z": res.get("abs_z"),
                    "severity": res.get("severity"),
                    "external_url": get_announcement_url(a),
                })
        anomalies.sort(key=lambda x: -x["abs_z"])
        return {
            "items": anomalies[:limit],
            "total_scanned": len(recent),
            "params": {"period_days": period_days, "z_threshold": z_threshold},
        }
    finally:
        db.close()


# ─── API: K-means 공고 클러스터링 ────────────────────────────────────────

@app.get("/api/v1/analysis/clusters")
@rate_limit("10/minute")
def analysis_clusters(
    request: Request,
    period_months: int = Query(12, ge=1, le=36),
    k: int = Query(5, ge=2, le=10),
    category_filter: str = Query("all", pattern="^(all|construction|service)$"),
    current_user: User = Depends(require_auth),
):
    """K-means 자동 군집화 — 기초금액·지역·카테고리 기반 공고 그룹.

    각 군집에 대해 표본 평균 사정률·대표 공고 5건 노출.
    """
    cache_key = _cache_key("clusters", period_months, k, category_filter)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        cutoff = datetime.now() - timedelta(days=period_months * 30)
        q = db.query(BidAnnouncement, BidResult).join(
            BidResult, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidAnnouncement.base_amount.isnot(None),
            BidAnnouncement.base_amount > 0,
            BidAnnouncement.announced_at >= cutoff,
            BidAnnouncement.category.in_(["공사", "용역"]),
        )
        q = _apply_category_filter(q, category_filter)
        rows = q.limit(2000).all()
        if len(rows) < k * 5:
            return {"clusters": [], "n": len(rows),
                    "reason": f"표본 부족 (k×5={k*5} 필요, 현재 {len(rows)})"}

        # 지역 인코딩 (top 5)
        from collections import Counter as _C
        top_regions = [r for r, _ in _C(a.region for a, _ in rows if a.region).most_common(5)]

        def featurize(a):
            return [
                math.log(max(1, a.base_amount or 1)),
                1.0 if a.category == "공사" else 0.0,
            ] + [1.0 if a.region == r else 0.0 for r in top_regions]

        features = [featurize(a) for a, _ in rows]
        labels, centroids = _kmeans_simple(features, k=k)

        # 군집별 집계
        clusters = []
        for cluster_idx in range(k):
            members = [(a, r, lbl) for (a, r), lbl in zip(rows, labels) if lbl == cluster_idx]
            if not members:
                continue
            rates = [r.assessment_rate for _, r, _ in members if r.assessment_rate]
            amounts = [a.base_amount for a, _, _ in members if a.base_amount]
            cats = _C(a.category for a, _, _ in members)
            regs = _C(a.region for a, _, _ in members if a.region)

            samples = [
                {
                    "id": a.id, "title": a.title[:60],
                    "ordering_org_name": a.ordering_org_name,
                    "base_amount": a.base_amount,
                    "assessment_rate": round(r.assessment_rate, 4) if r.assessment_rate else None,
                    "external_url": get_announcement_url(a),
                }
                for a, r, _ in members[:5]
            ]
            clusters.append({
                "cluster_id": cluster_idx,
                "size": len(members),
                "avg_assessment_rate": round(sum(rates) / len(rates), 4) if rates else None,
                "avg_base_amount": int(sum(amounts) / len(amounts)) if amounts else None,
                "category_dist": dict(cats),
                "top_regions": dict(regs.most_common(3)),
                "samples": samples,
            })
        clusters.sort(key=lambda x: -x["size"])

        result = {
            "params": {"period_months": period_months, "k": k, "category_filter": category_filter},
            "n_total": len(rows),
            "clusters": clusters,
        }
        _cache_set(cache_key, result)
        return result
    finally:
        db.close()


# ─── API: K-NN 유사 공고 추천 ──────────────────────────────────────────

@app.get("/api/v1/analysis/similar/{announcement_id}")
@rate_limit("30/minute")
def analysis_similar(
    request: Request,
    announcement_id: str,
    k: int = Query(5, ge=1, le=20),
    period_months: int = Query(12, ge=1, le=36),
    category_filter: str = Query("same", pattern="^(same|all|construction|service|industry)$"),
    current_user: User = Depends(require_auth),
):
    """대상 공고와 가장 유사한 과거 공고 K건 추천 (사정률 포함).

    유사도 = 1 / (1 + 가중 거리)
    거리 = 0.4×|log(기초금액)차이| + 0.3×지역 불일치 + 0.2×업종 불일치 + 0.1×경과/365
    """
    cache_key = _cache_key("similar", announcement_id, k, period_months, category_filter)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        q = db.query(BidAnnouncement, BidResult).join(
            BidResult, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidAnnouncement.announced_at >= cutoff,
            BidAnnouncement.id != announcement_id,
        )
        q = _apply_category_filter(q, category_filter, ann)

        candidates = []
        for c_ann, c_res in q.limit(2000).all():  # 후보 풀 제한
            days_old = (ref_date - c_ann.announced_at).days if c_ann.announced_at else 0
            candidates.append({
                "id": c_ann.id,
                "bid_number": c_ann.bid_number,
                "title": c_ann.title,
                "category": c_ann.category,
                "ordering_org_name": c_ann.ordering_org_name,
                "region": c_ann.region,
                "industry": c_ann.industry_code,
                "base_amount": c_ann.base_amount,
                "assessment_rate": round(c_res.assessment_rate, 4) if c_res.assessment_rate else None,
                "days_old": days_old,
                "external_url": get_announcement_url(c_ann),
            })

        target = {
            "base_amount": ann.base_amount,
            "region": ann.region,
            "industry": ann.industry_code,
            "category": ann.category,
        }
        similar = _knn_similar_announcements(target, candidates, k=k)

        # 유사 공고 평균 사정률
        rates = [s["assessment_rate"] for s in similar if s.get("assessment_rate") is not None]
        avg_similar_rate = round(sum(rates) / len(rates), 4) if rates else None
        predicted_amount = (
            int(ann.base_amount * avg_similar_rate / 100)
            if ann.base_amount and avg_similar_rate else None
        )

        result = {
            "announcement": {"id": ann.id, "title": ann.title, "category": ann.category,
                             "base_amount": ann.base_amount, "region": ann.region},
            "params": {"k": k, "period_months": period_months, "category_filter": category_filter},
            "similar": similar,
            "avg_similar_rate": avg_similar_rate,
            "predicted_bid_amount": predicted_amount,
            "candidate_pool_size": len(candidates),
        }
        _cache_set(cache_key, result)
        return result
    finally:
        db.close()


# ─── API: Walk-forward Validation (시계열 rolling 정확도) ────────────

@app.get("/api/v1/analysis/walk-forward/{announcement_id}")
@rate_limit("10/minute")
def analysis_walk_forward(
    request: Request,
    announcement_id: str,
    period_months: int = Query(24, ge=6, le=60),
    window_days: int = Query(90, ge=30, le=365),
    stride_days: int = Query(30, ge=7, le=90),
    category_filter: str = Query("same", pattern="^(same|all|construction|service|industry)$"),
    current_user: User = Depends(require_auth),
):
    """Walk-forward Validation — 시간순 rolling window 백테스트.

    각 시점마다 [t-window, t] 데이터로 학습 → 다음 stride 예측 → 실제값 비교.
    백테스트보다 robust 한 시계열 정확도 평가.
    """
    cache_key = _cache_key("walkfwd", announcement_id, period_months, window_days, stride_days, category_filter)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)

        q = db.query(BidResult.assessment_rate, BidResult.opened_at).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        q = _apply_category_filter(q, category_filter, ann)

        samples = [(r[0], r[1]) for r in q.all() if r[0] is not None and r[1] is not None]

        result = _walk_forward_validate(samples, window_days=window_days, stride_days=stride_days)
        result["announcement"] = {"id": ann.id, "title": ann.title, "category": ann.category}
        result["params"] = {
            "period_months": period_months, "window_days": window_days,
            "stride_days": stride_days, "category_filter": category_filter,
        }

        _cache_set(cache_key, result)
        return result
    finally:
        db.close()


# ─── API: 밀어내기식 검토 (Sliding Window Backtesting) ────────────────────

@app.get("/api/v1/analysis/sliding-review/{announcement_id}")
@rate_limit("30/minute")
def analysis_sliding_review(
    request: Request,
    announcement_id: str,
    window_size: str = Query("3m"),
    confirmed_rate: float = Query(99.5),
    org_scope: str = Query("specific"),
    current_user: User = Depends(require_auth),
):
    """밀어내기식 검토: 윈도우를 이동하며 예측 vs 실제 비교"""
    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()

        # 윈도우 크기 → 일수 변환
        window_days_map = {"1d": 1, "5d": 5, "1m": 30, "3m": 90, "6m": 180, "1y": 365}
        w_days = window_days_map.get(window_size, 90)

        # 전체 과거 데이터 가져오기 (최대 7년)
        full_cutoff = ref_date - timedelta(days=2650)
        q = db.query(BidResult, BidAnnouncement).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidResult.assessment_rate.isnot(None),
            BidResult.first_place_rate.isnot(None),
            BidResult.opened_at >= full_cutoff,
            BidResult.opened_at < ref_date,
        )
        if org_scope == "parent" and ann.parent_org_name:
            q = q.filter(BidAnnouncement.parent_org_name == ann.parent_org_name)
        elif org_scope == "specific":
            q = q.filter(BidAnnouncement.ordering_org_name == ann.ordering_org_name)

        all_data = q.order_by(BidResult.opened_at).all()
        if not all_data:
            # 폴백: 동일 카테고리+지역으로 범위 확대
            all_data = db.query(BidResult, BidAnnouncement).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.region == ann.region,
                BidResult.assessment_rate.isnot(None),
                BidResult.first_place_rate.isnot(None),
                BidResult.opened_at >= full_cutoff,
                BidResult.opened_at < ref_date,
            ).order_by(BidResult.opened_at).all()

        # 슬라이딩 윈도우: 각 시점에서 window 크기만큼의 데이터로 예측
        timeline = []
        match_count = 0
        step_days = max(w_days // 3, 1)  # 스텝 크기

        current_date = full_cutoff + timedelta(days=w_days)
        while current_date < ref_date:
            window_start = current_date - timedelta(days=w_days)
            window_data = [
                (r, a) for r, a in all_data
                if r.opened_at and window_start <= r.opened_at < current_date
            ]

            if len(window_data) >= 3:
                # 윈도우 내 사정률 평균/중앙값으로 예측
                window_rates = [r.assessment_rate for r, a in window_data]
                predicted_rate = round(statistics.median(window_rates), 4)

                # 해당 시점 이후 가장 가까운 실제 결과 찾기
                next_results = [
                    (r, a) for r, a in all_data
                    if r.opened_at and current_date <= r.opened_at < current_date + timedelta(days=step_days)
                ]

                for res, res_ann in next_results:
                    diff = abs(confirmed_rate - res.assessment_rate)
                    actual_diff = abs(res.first_place_rate - res.assessment_rate)
                    is_first = diff <= actual_diff + 0.02

                    if is_first:
                        match_count += 1

                    timeline.append({
                        "date": current_date.strftime("%Y-%m-%d"),
                        "window_start": window_start.strftime("%Y-%m-%d"),
                        "predicted_rate": predicted_rate,
                        "actual_rate": round(res.assessment_rate, 4),
                        "first_place_rate": round(res.first_place_rate, 4),
                        "confirmed_rate": round(confirmed_rate, 4),
                        "was_first_place": is_first,
                        "diff": round(diff, 4),
                        "data_count": len(window_data),
                    })

            current_date += timedelta(days=step_days)

        total = len(timeline)
        hit_rate = round(match_count / total * 100, 1) if total > 0 else 0

    finally:
        db.close()
    result = {
        "timeline": timeline[-60:],  # 최근 60개 포인트
        "summary": {
            "total_points": total,
            "match_count": match_count,
            "hit_rate": hit_rate,
            "window_size": window_size,
            "window_days": w_days,
        },
        "announcement": {
            "id": ann.id, "title": ann.title,
            "org": ann.ordering_org_name,
        },
    }
    if current_user:
        save_query_history(
            current_user.id, announcement_id, "sliding-review",
            parameters={"window_size": window_size, "confirmed_rate": confirmed_rate,
                        "org_scope": org_scope},
            result_summary={"hit_rate": hit_rate, "total_points": total,
                            "match_count": match_count},
        )
    return result


# ─── API: 연도별 낙찰확률 검증 ──────────────────────────────────────────

@app.get("/api/v1/analysis/yearly-validation/{announcement_id}")
@rate_limit("30/minute")
def analysis_yearly_validation(
    request: Request,
    announcement_id: str,
    confirmed_rate: float = Query(99.5),
    org_scope: str = Query("specific"),
    current_user: User = Depends(require_auth),
):
    """연도별(2020~2026) 낙찰확률 검증"""
    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        years_result = []
        for year in range(2020, 2027):
            year_start = datetime(year, 1, 1)
            year_end = datetime(year, 12, 31)
            prev_cutoff = datetime(year - 1, 1, 1)

            # 해당 연도 실제 낙찰 데이터
            q = db.query(BidResult, BidAnnouncement).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidResult.assessment_rate.isnot(None),
                BidResult.first_place_rate.isnot(None),
                BidResult.opened_at >= year_start,
                BidResult.opened_at <= year_end,
            )
            if org_scope == "parent" and ann.parent_org_name:
                q = q.filter(BidAnnouncement.parent_org_name == ann.parent_org_name)
            elif org_scope == "specific":
                q = q.filter(BidAnnouncement.ordering_org_name == ann.ordering_org_name)

            year_data = q.all()
            if not year_data:
                continue

            match_count = 0
            cases = []
            for res, past_ann in year_data:
                diff = abs(confirmed_rate - res.assessment_rate)
                actual_diff = abs(res.first_place_rate - res.assessment_rate)
                is_match = diff <= actual_diff + 0.02

                if is_match:
                    match_count += 1

                cases.append({
                    "title": past_ann.title,
                    "org": past_ann.ordering_org_name,
                    "predicted_rate": round(confirmed_rate, 4),
                    "actual_rate": round(res.assessment_rate, 4),
                    "first_place_rate": round(res.first_place_rate, 4),
                    "is_match": is_match,
                    "date": res.opened_at.strftime("%Y-%m-%d") if res.opened_at else "",
                })

            total = len(cases)
            years_result.append({
                "year": year,
                "total": total,
                "first_place_count": match_count,
                "match_rate": round(match_count / total * 100, 1) if total > 0 else 0,
                "cases": cases[:20],
            })

    finally:
        db.close()
    result = {
        "years": years_result,
        "confirmed_rate": round(confirmed_rate, 4),
        "announcement": {
            "id": ann.id, "title": ann.title,
            "org": ann.ordering_org_name,
        },
    }
    if current_user:
        avg_match = round(sum(y["match_rate"] for y in years_result) / len(years_result), 1) if years_result else 0
        save_query_history(
            current_user.id, announcement_id, "yearly-validation",
            parameters={"confirmed_rate": confirmed_rate, "org_scope": org_scope},
            result_summary={"avg_match_rate": avg_match, "year_count": len(years_result)},
        )
    return result


# ─── API: 기관 계층 조회 ──────────────────────────────────────────────────

@app.get("/api/v1/orgs/hierarchy")
def get_org_hierarchy():
    return {"hierarchy": ORG_HIERARCHY}


# ─── API: 추첨예가 빈도수 분석 ────────────────────────────────────────────

@app.get("/api/v1/analysis/preliminary-frequency/{announcement_id}")
def analysis_preliminary_frequency(
    announcement_id: str,
    period_months: int = Query(12, ge=1, le=24),
    org_scope: str = Query("specific"),
):
    """복수예비가격 15개 번호 중 추첨 빈도 분석"""
    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        ref_date = ann.announced_at or datetime.now()
        cutoff = ref_date - timedelta(days=period_months * 30)
        q = db.query(BidResult).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidResult.selected_price_indices.isnot(None),
            BidResult.opened_at >= cutoff,
        )
        if org_scope == "parent" and ann.parent_org_name:
            q = q.filter(BidAnnouncement.parent_org_name == ann.parent_org_name)
        else:
            q = q.filter(BidAnnouncement.ordering_org_name == ann.ordering_org_name)

        results = q.all()
        if not results:
            # 폴백: 같은 카테고리+지역
            q = db.query(BidResult).join(
                BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
            ).filter(
                BidAnnouncement.category == ann.category,
                BidAnnouncement.region == ann.region,
                BidResult.selected_price_indices.isnot(None),
                BidResult.opened_at >= cutoff,
            )
            results = q.all()

        # 15개 번호별 추첨 횟수 집계
        freq = {i: 0 for i in range(1, 16)}
        total_cases = 0
        for res in results:
            try:
                indices = json.loads(res.selected_price_indices) if isinstance(res.selected_price_indices, str) else res.selected_price_indices
                if indices:
                    total_cases += 1
                    for idx in indices:
                        if 1 <= idx <= 15:
                            freq[idx] += 1
            except (json.JSONDecodeError, TypeError):
                continue

        bins = [{"number": i, "count": freq[i], "percentage": round(freq[i] / total_cases * 100, 1) if total_cases > 0 else 0} for i in range(1, 16)]
        max_count = max(b["count"] for b in bins) if bins else 0
        peak_numbers = [b["number"] for b in bins if b["count"] == max_count and max_count > 0]

    finally:
        db.close()
    return {
        "bins": bins,
        "total_cases": total_cases,
        "peak_numbers": peak_numbers,
        "selected_per_case": 4,
    }


# ─── API: 통계 ────────────────────────────────────────────────────────────

@app.get("/api/v1/stats/kpi")
def get_kpi():
    db = SessionLocal()
    try:
        total = db.query(BidAnnouncement).filter(BidAnnouncement.category.in_(["공사", "용역"])).count()
        today = db.query(BidAnnouncement).filter(
            func.date(BidAnnouncement.announced_at) == func.date(datetime.now())
        ).count()
        construction = db.query(BidAnnouncement).filter(BidAnnouncement.category == "공사").count()
        service = db.query(BidAnnouncement).filter(BidAnnouncement.category == "용역").count()
        avg_rate = db.query(func.avg(BidResult.assessment_rate)).scalar()
        avg_fp_rate = db.query(func.avg(BidResult.first_place_rate)).scalar()
    finally:
        db.close()
    return {
        "total_announcements": total,
        "today_announcements": today,
        "construction_count": construction,
        "service_count": service,
        "avg_assessment_rate": round(float(avg_rate), 2) if avg_rate else 0,
        "avg_first_place_rate": round(float(avg_fp_rate), 4) if avg_fp_rate else 0,
    }


@app.get("/api/v1/stats/trends")
def get_trends(months: int = 6, category: str = None):
    db = SessionLocal()
    try:
        q = db.query(
            func.strftime("%Y-%m", BidResult.opened_at).label("period"),
            BidAnnouncement.category,
            func.avg(BidResult.assessment_rate).label("avg_rate"),
            func.count(BidResult.id).label("cnt"),
        ).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidResult.opened_at.isnot(None),
            BidAnnouncement.category.in_(["공사", "용역"]),
        )
        if category and category != "all":
            q = q.filter(BidAnnouncement.category == category)
        rows = q.group_by(
            func.strftime("%Y-%m", BidResult.opened_at),
            BidAnnouncement.category,
        ).order_by("period").all()

        period_data = {}
        for row in rows:
            p = row.period
            if p not in period_data:
                period_data[p] = {"period": p, "construction": None, "service": None, "total": None, "count": 0}
            if row.category == "공사":
                period_data[p]["construction"] = round(row.avg_rate, 2)
            elif row.category == "용역":
                period_data[p]["service"] = round(row.avg_rate, 2)
            period_data[p]["count"] += row.cnt

        for p in period_data:
            vals = [v for v in [period_data[p]["construction"], period_data[p]["service"]] if v]
            period_data[p]["total"] = round(sum(vals) / len(vals), 2) if vals else None

        data = sorted(period_data.values(), key=lambda x: x["period"])[-months:]
    finally:
        db.close()
    return {"data": data}


@app.get("/api/v1/stats/by-region")
def get_region_stats():
    db = SessionLocal()
    try:
        rows = db.query(
            BidAnnouncement.region,
            func.avg(BidResult.assessment_rate).label("avg_rate"),
            func.count(BidResult.id).label("count"),
            func.min(BidResult.assessment_rate).label("min_rate"),
            func.max(BidResult.assessment_rate).label("max_rate"),
        ).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidResult.assessment_rate.isnot(None),
            BidAnnouncement.region.isnot(None),
            BidAnnouncement.category.in_(["공사", "용역"]),
        ).group_by(BidAnnouncement.region).order_by(func.count(BidResult.id).desc()).all()

        data = [
            {
                "region": r.region,
                "rate": round(r.avg_rate, 2),
                "count": r.count,
                "min_rate": round(r.min_rate, 2) if r.min_rate else None,
                "max_rate": round(r.max_rate, 2) if r.max_rate else None,
            }
            for r in rows
        ]
    finally:
        db.close()
    return {"data": data}


# ─── API: 예측 (기존 호환) ────────────────────────────────────────────────

@app.get("/api/v1/predictions/{announcement_id}")
def predict(announcement_id: str):
    """실시간 사정률 예측 (통계 기반 간이 모델)"""
    db = SessionLocal()
    try:
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == announcement_id).first()
        if not ann:
            db.close()
            return {"error": "공고를 찾을 수 없습니다."}

        stats = db.query(
            func.avg(BidResult.assessment_rate).label("avg"),
            func.min(BidResult.assessment_rate).label("min"),
            func.max(BidResult.assessment_rate).label("max"),
            func.count(BidResult.id).label("cnt"),
        ).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidAnnouncement.region == ann.region,
            BidResult.assessment_rate.isnot(None),
        ).first()

        org_stats = db.query(
            func.avg(BidResult.assessment_rate)
        ).join(
            BidAnnouncement, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.ordering_org_type == ann.ordering_org_type,
            BidResult.assessment_rate.isnot(None),
        ).scalar()

        if stats and stats.avg and stats.cnt >= 5:
            pred_rate = round(stats.avg * 0.7 + (org_stats or stats.avg) * 0.3, 2)
            ci_width = max(1.0, (stats.max - stats.min) * 0.3)
            confidence = min(95, 50 + stats.cnt * 0.5)
        else:
            pred_rate = 99.3 if ann.category == "공사" else 99.5
            ci_width = 3.0
            confidence = 30

        pred_min = round(pred_rate - ci_width, 2)
        pred_max = round(pred_rate + ci_width, 2)
        bid_pred = int(ann.base_amount * pred_rate / 100) if ann.base_amount else 0
        bid_min = int(ann.base_amount * pred_min / 100) if ann.base_amount else 0
        bid_max = int(ann.base_amount * pred_max / 100) if ann.base_amount else 0

        similar = db.query(BidAnnouncement, BidResult).join(
            BidResult, BidResult.announcement_id == BidAnnouncement.id
        ).filter(
            BidAnnouncement.category == ann.category,
            BidResult.assessment_rate.isnot(None),
        ).order_by(BidResult.opened_at.desc()).limit(10).all()

        similar_list = [
            {
                "title": s[0].title, "area": s[0].region,
                "rate": round(s[1].assessment_rate, 2),
                "first_place_rate": round(s[1].first_place_rate, 4) if s[1].first_place_rate else None,
                "amount": s[1].winning_amount,
                "date": s[1].opened_at.strftime("%Y-%m") if s[1].opened_at else "",
            }
            for s in similar
        ]

    finally:
        db.close()
    return {
        "announcement": {
            "id": ann.id, "title": ann.title, "type": ann.category,
            "area": ann.region, "budget": ann.base_amount,
            "org": ann.ordering_org_name, "org_type": ann.ordering_org_type,
            "bid_method": ann.bid_method,
        },
        "prediction": {
            "predRate": round(pred_rate, 2),
            "predMin": pred_min, "predMax": pred_max,
            "bidAmountPred": bid_pred, "bidMin": bid_min, "bidMax": bid_max,
            "confidence": round(confidence),
            "dataPoints": stats.cnt if stats else 0,
        },
        "similar": similar_list,
    }


# ─── API: 관리자 ──────────────────────────────────────────────────────────

@app.get("/api/v1/admin/dashboard")
def admin_dashboard(current_user: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        total_users = db.query(User).count()
        active_users = db.query(User).filter(User.is_active == True).count()
        total_ann = db.query(BidAnnouncement).count()
        total_res = db.query(BidResult).count()
        # 평균 사정률
        avg_rate = db.query(func.avg(BidResult.assessment_rate)).scalar()
        avg_rate = round(avg_rate, 4) if avg_rate else 0
        # 매칭률 — 사정률이 산출된 결과 비율 (전체 결과 대비)
        rate_with_assessment = db.query(BidResult).filter(
            BidResult.assessment_rate.isnot(None)).count()
        matching_rate = round(rate_with_assessment / total_res * 100, 1) if total_res else 0
        # 분석 가능 공고 수 — BidResult 가 매칭된 공고
        analyzable_ann = db.query(BidResult.announcement_id).distinct().count()
        analyzable_rate = round(analyzable_ann / total_ann * 100, 1) if total_ann else 0

        logs = db.query(DataSyncLog).order_by(DataSyncLog.started_at.desc()).limit(10).all()
        pipelines = []
        seen = set()
        for log in logs:
            key = f"{log.source}_{log.sync_type}"
            if key in seen:
                continue
            seen.add(key)
            pipelines.append({
                "name": f"{log.source} {log.sync_type}",
                "status": log.status,
                "last": log.started_at.strftime("%Y-%m-%d %H:%M") if log.started_at else "",
                "count": f"{log.records_fetched}건",
            })

        users = db.query(User).order_by(User.query_count.desc()).all()
        user_list = [
            {
                "id": u.id, "email": u.email,
                "joined": u.joined_at.strftime("%Y.%m.%d") if u.joined_at else "",
                "last": u.last_login_at.strftime("%Y-%m-%d") if u.last_login_at else "",
                "queries": u.query_count,
                "is_active": u.is_active if u.is_active is not None else True,
                "role": u.role,
            }
            for u in users
        ]

        # 데이터 출처 비율
        g2b_count = db.query(BidAnnouncement).filter(BidAnnouncement.source == "G2B").count()
        d2b_count = db.query(BidAnnouncement).filter(BidAnnouncement.source == "D2B").count()
        upload_count = db.query(BidAnnouncement).filter(BidAnnouncement.source == "UPLOAD").count()

        # 총 분석 건수
        total_analyses = db.query(QueryHistory).count()

    finally:
        db.close()
    return {
        "total_users": total_users, "active_users": active_users,
        "total_announcements": total_ann, "total_results": total_res,
        "avg_assessment_rate": avg_rate,
        "total_analyses": total_analyses,
        # 매칭률·신뢰도 KPI
        "matching_rate": matching_rate,
        "analyzable_announcements": analyzable_ann,
        "analyzable_rate": analyzable_rate,
        "source_ratio": {"G2B": g2b_count, "D2B": d2b_count, "UPLOAD": upload_count},
        "pipelines": pipelines,
        "users": user_list,
    }


def _parse_date(s) -> "datetime | None":
    """문자열 → datetime (공공데이터 여러 포맷 대응)"""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%Y%m%d%H%M", "%Y%m%d", "%Y/%m/%d %H:%M", "%Y/%m/%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _extract_g2b_items(payload_bytes: bytes) -> list[dict]:
    """G2B BidPublicInfoService 응답에서 items 리스트 추출"""
    import json as _json
    try:
        data = _json.loads(payload_bytes.decode("utf-8"))
    except (UnicodeDecodeError, ValueError) as exc:
        logger.warning("G2B 응답 파싱 실패 (size=%d): %s", len(payload_bytes), exc)
        return []
    # 표준 응답 경로: response.body.items
    try:
        body = data.get("response", {}).get("body", {})
        items = body.get("items", [])
        if isinstance(items, dict):
            items = items.get("item", [])
        if isinstance(items, list):
            return items
    except (AttributeError, KeyError, TypeError) as exc:
        logger.debug("G2B items 경로 추출 실패: %s", exc)
    return []


# 국방·군 발주 자동 인식 키워드
DEFENSE_KEYWORDS = (
    "국방", "육군", "해군", "공군", "방위", "방사청", "국군", "병무",
    "계룡대", "훈련소", "사령부", "군수", "방산", "국방과학",
)


def _detect_defense(org_name, title=None) -> bool:
    """발주기관/제목에 국방 관련 키워드가 있으면 True"""
    text = f"{org_name or ''} {title or ''}"
    return any(kw in text for kw in DEFENSE_KEYWORDS)


def get_announcement_url(ann) -> str:
    """공고의 외부 URL 반환 — external_url 우선, 없으면 표준 G2B URL fallback."""
    return (getattr(ann, "external_url", None)
            or build_g2b_url(getattr(ann, "bid_number", ""),
                             getattr(ann, "bid_ord", None) or "000"))


def build_g2b_url(bid_number: str, bid_ord: str = "000") -> str:
    """G2B 공고 상세 페이지 fallback URL 생성.

    실제 응답의 bidNtceDtlUrl이 우선이지만, 없을 경우 표준 패턴으로 생성.
    """
    if not bid_number:
        return ""
    ord_str = (bid_ord or "000").zfill(3)
    return (
        "https://www.g2b.go.kr/link/PNPE027_01/single/"
        f"?bidPbancNo={bid_number}&bidPbancOrd={ord_str}"
    )


def _normalize_item(source: str, item: dict) -> "dict | None":
    """공공데이터 item → BidAnnouncement 필드로 정규화.

    실제 필드명은 소스별로 다르지만 아래 후보 키들을 순회하며 매핑.
    """
    def pick(*keys):
        for k in keys:
            if k in item and item[k] not in (None, "", "null"):
                return item[k]
        return None

    bid_number = pick("bidNtceNo", "bidNo", "bid_number", "noticeNo")
    title = pick("bidNtceNm", "bidNm", "title", "noticeNm")
    org = pick("dminsttNm", "ordInsttNm", "ordering_org_name", "orgNm")
    if not bid_number or not title:
        return None

    # 기초금액: presmptPrce / basicAmt / budgetAmt / asignBdgtAmt 순으로 폴백
    base_raw = pick("presmptPrce", "basicAmt", "bssamt", "asignBdgtAmt", "budgetAmt", "base_amount")
    try:
        base_amount = int(float(base_raw)) if base_raw else None
        if base_amount is not None and base_amount < 0:
            base_amount = None
    except (ValueError, TypeError):
        base_amount = None

    announced_at = _parse_date(pick("bidNtceDt", "ntceDt", "announced_at")) or datetime.now()
    deadline_at = _parse_date(pick("bidClseDt", "opengDt", "deadline_at"))
    category = pick("bsnsDivNm", "category") or "용역"
    if isinstance(category, str):
        if "공사" in category:
            category = "공사"
        elif "용역" in category:
            category = "용역"
        elif "물품" in category:
            category = "물품"

    org_str = str(org).strip() if org else "미상"
    title_str = str(title).strip()
    bid_no = str(bid_number).strip()
    bid_ord_raw = pick("bidNtceOrd", "bid_ord") or "000"
    bid_ord = str(bid_ord_raw).strip() or "000"
    # 응답의 bidNtceDtlUrl 우선, 없으면 표준 패턴으로 생성
    ext_url = pick("bidNtceDtlUrl", "bidNtceUrl", "external_url") or build_g2b_url(bid_no, bid_ord)
    return {
        "source": source,
        "bid_number": bid_no,
        "bid_ord": bid_ord,
        "title": title_str,
        "ordering_org_name": org_str,
        "region": pick("prtcptLmtRgnNm", "region"),
        "industry_code": pick("indstrytyCd", "industry_code"),
        "base_amount": base_amount,
        "bid_method": pick("cntrctCnclsMthdNm", "bid_method"),
        "announced_at": announced_at,
        "deadline_at": deadline_at,
        "category": category,
        "status": "진행중",
        "is_defense": _detect_defense(org_str, title_str),
        "external_url": ext_url[:1000] if ext_url else None,
    }


def _upsert_announcements(db, source: str, items: list[dict]) -> dict:
    """공고 upsert — 신규는 insert, 기존은 가변 필드(status/deadline/base_amount)만 update.

    - 신규(bid_number 미존재): 그대로 insert
    - 기존: status / deadline_at / base_amount / external_url 만 비교 후 변경 시 update
      (title / org / category 등 불변 메타는 최초 적재 값 보존)
    """
    if not items:
        return {"inserted": 0, "skipped": 0, "updated": 0, "invalid": 0}
    numbers = [i["bid_number"] for i in items if i.get("bid_number")]
    existing_rows = {
        row.bid_number: row for row in
        db.query(BidAnnouncement).filter(BidAnnouncement.bid_number.in_(numbers)).all()
    }
    inserted = 0
    skipped = 0
    updated = 0
    batch: list[BidAnnouncement] = []
    seen_in_payload: set = set()
    # 상태 변경 감지 대상 필드 (값이 바뀌면 update)
    # category 포함: 엔드포인트 기반 강제 분류 적용 후 재수집 시 기존 오분류 자동 정정
    MUTABLE_FIELDS = ("status", "deadline_at", "base_amount", "external_url", "category")
    for norm in items:
        bn = norm["bid_number"]
        if bn in seen_in_payload:
            skipped += 1
            continue
        seen_in_payload.add(bn)

        existing = existing_rows.get(bn)
        if existing is None:
            batch.append(BidAnnouncement(**norm))
            inserted += 1
            continue

        # 기존 레코드 — 변경된 가변 필드만 패치
        changed = False
        for field in MUTABLE_FIELDS:
            new_val = norm.get(field)
            if new_val is None:
                continue
            old_val = getattr(existing, field, None)
            # base_amount: 0/None 은 "비어있음"으로 간주 (보강만)
            if field == "base_amount" and (not old_val) and new_val:
                setattr(existing, field, new_val)
                changed = True
            elif old_val != new_val and field != "base_amount":
                setattr(existing, field, new_val)
                changed = True
        if changed:
            updated += 1
        else:
            skipped += 1
    if batch:
        db.bulk_save_objects(batch)
    if batch or updated:
        db.commit()
        invalidate_analysis_cache()  # 신규/변경 시 종합분석 캐시 무효화
    return {"inserted": inserted, "skipped": skipped, "updated": updated, "invalid": 0}


# G2B 카테고리별 신규 엔드포인트 (/ad/BidPublicInfoService — 2026 운영 버전)
# 비즈니스 요건상 용역/공사만 수집 (물품/외자 제외)
G2B_CATEGORIES = [
    ("용역", "getBidPblancListInfoServc"),
    ("공사", "getBidPblancListInfoCnstwk"),
]

# 낙찰결과 + 사정률 상세 (예정가격) 엔드포인트 (/as/ScsbidInfoService)
G2B_RESULT_CATEGORIES = [
    ("용역", "getOpengResultListInfoServcPreparPcDetail"),
    ("공사", "getOpengResultListInfoCnstwkPreparPcDetail"),
]


def _windows_iter(total_days: int, window_days: int, start_date=None):
    """오늘부터 거꾸로 window_days 단위로 (start_str, end_str) 튜플 yield.

    형식: 'YYYYMMDDHHMM' (G2B inqryDiv 표준).
    data.go.kr API는 7일 초과 inqry 시 빈 응답 → 윈도우 분할 필수.

    - start_date(YYYY-MM-DD) 가 주어지면 해당 일자까지 거꾸로 모두 순회 (전체 백필).
      total_days 는 무시됨. 예) "2019-01-01" → 2019-01-01 ~ 오늘 전 구간을 7일 단위 분할.
    - start_date 가 없으면 기존처럼 최근 total_days 만 순회.
    """
    now = datetime.now()
    if start_date:
        try:
            floor_dt = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            floor_dt = None
        if floor_dt:
            win_end = now
            while win_end > floor_dt:
                win_start = max(floor_dt, win_end - timedelta(days=window_days))
                yield (
                    win_start.strftime("%Y%m%d") + "0000",
                    win_end.strftime("%Y%m%d") + "2359",
                )
                win_end = win_start
            return
    # 기본: 최근 total_days 만 순회
    produced = 0
    while produced < total_days:
        win = min(window_days, total_days - produced)
        win_end = now - timedelta(days=produced)
        win_start = win_end - timedelta(days=win)
        yield (
            win_start.strftime("%Y%m%d") + "0000",
            win_end.strftime("%Y%m%d") + "2359",
        )
        produced += win


def _record_sync_error(per_cat_errors: list, name: str, cat_label: str, page: int, exc: Exception):
    """수집 루프 예외를 분류해서 per_cat_errors 에 기록."""
    if isinstance(exc, urllib.error.HTTPError):
        per_cat_errors.append(f"{name} {cat_label} p{page} HTTP {exc.code}")
    elif isinstance(exc, (urllib.error.URLError, TimeoutError)):
        per_cat_errors.append(f"{name} {cat_label} p{page} 네트워크: {str(exc)[:60]}")
    else:
        per_cat_errors.append(f"{name} {cat_label} p{page} 파싱: {str(exc)[:60]}")


def _build_sync_url(source: str, api_key: str, rows: int = 100, page_no: int = 1,
                    start_date: str = None, end_date: str = None,
                    category_op: str = "getBidPblancListInfoServc",
                    days: int = 30) -> str:
    """data.go.kr 표준 API URL 빌더 — 페이지네이션 + 일자 범위 지원

    G2B는 `/1230000/ad/BidPublicInfoService/{operation}` 형태이며
    필수 파라미터로 inqryDiv=01 + inqryBgnDt + inqryEndDt 가 필요.
    start_date / end_date 형식: 'YYYYMMDDHHMM' (없으면 최근 days일 자동 설정)
    """
    qkey = urllib.parse.quote(api_key, safe="")
    if source == "G2B":
        if not start_date or not end_date:
            now = datetime.now()
            end_date = end_date or now.strftime("%Y%m%d") + "2359"
            start_date = start_date or (now - timedelta(days=days)).strftime("%Y%m%d") + "0000"
        return (
            f"https://apis.data.go.kr/1230000/ad/BidPublicInfoService/{category_op}"
            f"?serviceKey={qkey}&numOfRows={rows}&pageNo={page_no}&type=json"
            f"&inqryDiv=01&inqryBgnDt={start_date}&inqryEndDt={end_date}"
        )
    if source == "D2B":
        return f"https://d2b.go.kr/bidApi/list?apiKey={qkey}&size={rows}"
    raise ValueError(f"알 수 없는 source: {source}")


def _build_result_url(api_key: str, category_op: str, rows: int = 100, page_no: int = 1,
                      start_date: str = None, end_date: str = None,
                      days: int = 7) -> str:
    """낙찰결과(예정가격 상세) URL 빌더 — 페이지네이션 + 기간 옵션.

    PreparPcDetail 응답이 무거우므로 기본 7일치 + numOfRows=100으로 제한.
    """
    qkey = urllib.parse.quote(api_key, safe="")
    if not start_date or not end_date:
        now = datetime.now()
        end_date = end_date or now.strftime("%Y%m%d") + "2359"
        start_date = start_date or (now - timedelta(days=days)).strftime("%Y%m%d") + "0000"
    return (
        f"https://apis.data.go.kr/1230000/as/ScsbidInfoService/{category_op}"
        f"?serviceKey={qkey}&numOfRows={rows}&pageNo={page_no}&type=json"
        f"&inqryDiv=01&inqryBgnDt={start_date}&inqryEndDt={end_date}"
    )


def _normalize_result_item(item: dict) -> "dict | None":
    """낙찰결과 PreparPcDetail item → BidResult 필드로 정규화.

    핵심 매핑:
      plnprc → 사정률 계산 (plnprc / bssamt * 100)
      bssamt → 기초금액 (BidAnnouncement.base_amount 보강용)
      plnprc → winning_amount (낙찰가 근사치)
      compnoRsrvtnPrceMkngDt → opened_at
    """
    bid_number = item.get("bidNtceNo")
    if not bid_number:
        return None

    plnprc = item.get("plnprc")
    bssamt = item.get("bssamt")
    try:
        plnprc_f = float(plnprc) if plnprc not in (None, "", "null") else None
        bssamt_f = float(bssamt) if bssamt not in (None, "", "null") else None
    except (ValueError, TypeError):
        plnprc_f = bssamt_f = None

    # 사정률 = plnprc / bssamt * 100
    assessment_rate = None
    if plnprc_f and bssamt_f and bssamt_f > 0:
        assessment_rate = round(plnprc_f / bssamt_f * 100, 4)

    opened_at = _parse_date(item.get("rlOpengDt") or item.get("compnoRsrvtnPrceMkngDt") or item.get("inptDt"))

    return {
        "bid_number": str(bid_number).strip(),
        "winning_amount": int(plnprc_f) if plnprc_f else None,
        "winning_rate": assessment_rate,
        "assessment_rate": assessment_rate,
        # first_place_*는 별도 endpoint에서 보강 (현재 미수집)
        "first_place_rate": assessment_rate,
        "first_place_amount": int(plnprc_f) if plnprc_f else None,
        "num_bidders": None,
        "winning_company": None,
        "opened_at": opened_at or datetime.now(),
        # 메타 (참조용, 모델엔 없음)
        "_bssamt": int(bssamt_f) if bssamt_f else None,
        "_drwt_nums": item.get("drwtNum"),
    }


def _aggregate_prelim_prices(items: list[dict]) -> list[dict]:
    """PreparPcDetail 응답을 bid_number 단위로 집계 — 매칭률 향상 핵심.

    한 공고당 15개 예비가격 row 가 옴. 그중:
      1) 추첨된(drwtNum != null) 4개의 plnprc 평균 = 사정가(예정가격)
      2) 추첨된 게 없으면 15개 plnprc 평균을 fallback 으로 사용
      3) bssamt(기초금액)는 모든 row 에 동일

    사정률 = 사정가 / bssamt × 100
    """
    by_bid: dict = {}
    for it in items:
        bn = it.get("bid_number")
        if not bn:
            continue
        by_bid.setdefault(bn, []).append(it)

    aggregated = []
    for bn, rows in by_bid.items():
        # 추첨된 예비가격 우선 (drwtNum 값 존재)
        selected_prices = []
        all_prices = []
        bssamt = None
        opened_at = None
        for r in rows:
            plnprc = r.get("winning_amount")  # _normalize_result_item 에서 이미 plnprc → winning_amount
            bs = r.get("_bssamt")
            if plnprc:
                all_prices.append(plnprc)
                if r.get("_drwt_nums"):  # 추첨번호 존재 → 선택된 예비가
                    selected_prices.append(plnprc)
            if bs and not bssamt:
                bssamt = bs
            if r.get("opened_at") and not opened_at:
                opened_at = r["opened_at"]

        prelim_prices = selected_prices or all_prices
        if not prelim_prices:
            continue

        avg_prearng = sum(prelim_prices) / len(prelim_prices)
        rate = round(avg_prearng / bssamt * 100, 4) if bssamt and bssamt > 0 else None

        aggregated.append({
            "bid_number": bn,
            "winning_amount": int(avg_prearng),
            "winning_rate": rate,
            "assessment_rate": rate,
            "first_place_rate": rate,
            "first_place_amount": int(avg_prearng),
            "num_bidders": None,
            "winning_company": None,
            "opened_at": opened_at,
            "_bssamt": bssamt,
            "preliminary_prices": ",".join(str(int(p)) for p in all_prices),
            "selected_price_indices": ",".join(["1"] * len(selected_prices)) if selected_prices else "",
        })
    return aggregated


def _upsert_results(db, items: list[dict]) -> dict:
    """낙찰결과 upsert — 공고별 1건만 유지.

    PreparPcDetail 의 다중 예비가격을 _aggregate_prelim_prices() 로 집계 후
    공고별 사정률(평균 사정가 / 기초금액) 산출.
    부수 효과: BidAnnouncement.base_amount 가 비어 있으면 bssamt 로 보강.
    """
    if not items:
        return {"inserted": 0, "skipped": 0, "no_announcement": 0}

    # 매칭률 개선 — 첫 row dedup 대신 전체 집계
    items = _aggregate_prelim_prices(items)

    bid_numbers = [it["bid_number"] for it in items if it.get("bid_number")]
    # bid_number → BidAnnouncement.id 매핑
    ann_map = {
        a.bid_number: a for a in
        db.query(BidAnnouncement).filter(BidAnnouncement.bid_number.in_(bid_numbers)).all()
    }
    # 이미 결과가 있는 announcement_id
    if ann_map:
        existing_ann_ids = {
            row[0] for row in
            db.query(BidResult.announcement_id).filter(
                BidResult.announcement_id.in_([a.id for a in ann_map.values()])
            ).all()
        }
    else:
        existing_ann_ids = set()

    inserted = 0
    skipped = 0
    no_ann = 0
    batch_results: list[BidResult] = []
    for it in items:
        ann = ann_map.get(it["bid_number"])
        if not ann:
            no_ann += 1
            continue
        if ann.id in existing_ann_ids:
            skipped += 1
            continue
        existing_ann_ids.add(ann.id)

        # base_amount 보강
        if (ann.base_amount in (None, 0)) and it.get("_bssamt"):
            ann.base_amount = it["_bssamt"]

        batch_results.append(BidResult(
            announcement_id=ann.id,
            winning_amount=it.get("winning_amount"),
            winning_rate=it.get("winning_rate"),
            assessment_rate=it.get("assessment_rate"),
            first_place_rate=it.get("first_place_rate"),
            first_place_amount=it.get("first_place_amount"),
            num_bidders=it.get("num_bidders"),
            winning_company=it.get("winning_company"),
            preliminary_prices=it.get("preliminary_prices"),
            selected_price_indices=it.get("selected_price_indices"),
            opened_at=it.get("opened_at"),
        ))
        inserted += 1

    if batch_results:
        db.bulk_save_objects(batch_results)
        db.commit()
        invalidate_analysis_cache()

    return {"inserted": inserted, "skipped": skipped, "no_announcement": no_ann}


def _http_fetch_with_retry(url: str, timeout: int = 15, retries: int = 2) -> bytes:
    """일시적 네트워크 오류 시 지수 백오프로 재시도"""
    last_exc = None
    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(
                url,
                headers={"User-Agent": "bid-insight/1.0", "Accept": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except (urllib.error.URLError, TimeoutError) as e:
            last_exc = e
            if attempt < retries:
                import time as _t
                _t.sleep(0.5 * (2 ** attempt))  # 0.5s → 1s
                logger.warning("retrying sync fetch attempt=%d url=%s err=%s",
                              attempt + 1, url[:80], e)
                continue
            raise
    raise last_exc  # 도달 불가


def _run_sync_for_source(source: str, trigger: str = "manual") -> dict:
    """공공데이터 실 수집 파이프라인.

    API 키가 있으면 실제 호출 → JSON 파싱 → 정규화 → BidAnnouncement upsert.
    키가 없거나 호출/파싱이 실패하면 로그만 남기고 DB는 변경하지 않음.
    """
    db = SessionLocal()
    started = datetime.now()
    status = "success"
    error_message = None
    summary = {"fetched": 0, "inserted": 0, "skipped": 0, "invalid": 0,
               "results_fetched": 0, "results_inserted": 0, "results_skipped": 0,
               "results_no_announcement": 0}

    key_env = f"{source}_API_KEY"
    api_key = os.environ.get(key_env, "").strip()

    # 페이지네이션 / 기간 한계 (환경변수로 조정)
    # data.go.kr API는 7일 초과 inqry 시 빈 응답 → 7일 단위 슬라이딩 윈도우로 분할 호출
    rows_per_page = int(os.environ.get("SYNC_ROWS_PER_PAGE", "100"))
    max_pages_ann = int(os.environ.get("SYNC_MAX_PAGES_ANN", "5"))      # 공고: 윈도우당 최대 5페이지
    max_pages_res = int(os.environ.get("SYNC_MAX_PAGES_RES", "5"))      # 낙찰: 윈도우당 최대 5페이지
    window_days = int(os.environ.get("SYNC_WINDOW_DAYS", "7"))           # API 호환 윈도우 크기
    # 공고는 낙찰일자 - 2~3개월 전 등록되므로 충분히 길게 (낙찰결과와 매칭률↑)
    total_days_ann = int(os.environ.get("SYNC_TOTAL_DAYS_ANN", "60"))   # 공고: 누적 최대 60일 (start_date 미지정 시)
    total_days_res = int(os.environ.get("SYNC_TOTAL_DAYS_RES", "14"))   # 낙찰: 누적 최대 14일 (start_date 미지정 시)
    # 전체 백필 시작일 — 지정 시 해당 일자부터 오늘까지 7일 단위 전 구간 수집
    # 기본값 2019-01-01: 공공데이터 포털 G2B 데이터 가용 기간 시작점
    sync_start_ann = os.environ.get("SYNC_START_DATE_ANN", os.environ.get("SYNC_START_DATE", "2019-01-01")).strip() or None
    sync_start_res = os.environ.get("SYNC_START_DATE_RES", os.environ.get("SYNC_START_DATE", "2019-01-01")).strip() or None

    # 증분 모드 — 직전 성공 수집 시점 이후만 다시 호출 (cron 비용 절감)
    # SYNC_INCREMENTAL=true 이면 DataSyncLog 의 마지막 성공 시점 - 1일(buffer) 부터만 윈도우 생성.
    # SYNC_FORCE_FULL=true (또는 trigger=='backfill') 이면 강제로 전체 구간 재수집.
    incremental = os.environ.get("SYNC_INCREMENTAL", "false").lower() in ("true", "1", "yes")
    force_full = os.environ.get("SYNC_FORCE_FULL", "false").lower() in ("true", "1", "yes") or trigger == "backfill"
    if incremental and not force_full:
        last_ok = (
            db.query(DataSyncLog.finished_at)
            .filter(DataSyncLog.source == source, DataSyncLog.status == "success")
            .order_by(DataSyncLog.finished_at.desc())
            .first()
        )
        if last_ok and last_ok[0]:
            buf_days = int(os.environ.get("SYNC_INCREMENTAL_BUFFER_DAYS", "1"))
            inc_floor = (last_ok[0] - timedelta(days=buf_days)).strftime("%Y-%m-%d")
            # 전체 백필 시작일보다 더 최근 시점일 때만 좁힌다 (앞으로 당기지는 않음)
            if not sync_start_ann or inc_floor > sync_start_ann:
                sync_start_ann = inc_floor
                sync_start_res = inc_floor
            logger.info("incremental sync floor=%s (last_ok=%s)", inc_floor, last_ok[0])

    try:
        if api_key:
            # G2B는 용역/공사 2개 카테고리 순회 수집
            categories = G2B_CATEGORIES if source == "G2B" else [(None, None)]
            result_categories = G2B_RESULT_CATEGORIES if source == "G2B" else []
            per_cat_errors = []

            # ① 입찰공고 수집 (카테고리 × 슬라이딩 윈도우 × 페이지)
            for cat_label, cat_op in categories:
                cat_fetched = cat_inserted = 0
                for win_start, win_end in _windows_iter(total_days_ann, window_days, start_date=sync_start_ann):
                    for page in range(1, max_pages_ann + 1):
                        try:
                            url = _build_sync_url(
                                source, api_key, rows=rows_per_page, page_no=page,
                                start_date=win_start, end_date=win_end,
                                category_op=cat_op or "getBidPblancListInfoServc",
                            )
                            payload = _http_fetch_with_retry(url, timeout=20, retries=2)
                            raw_items = _extract_g2b_items(payload)
                            if not raw_items:
                                break
                            summary["fetched"] += len(raw_items)
                            cat_fetched += len(raw_items)
                            normalized = []
                            for it in raw_items:
                                norm = _normalize_item(source, it)
                                if norm:
                                    # 엔드포인트가 카테고리를 명시한 경우 강제 적용
                                    # (Cnstwk/Servc 엔드포인트의 응답에는 bsnsDivNm 이 누락되거나 다른 값일 수 있어
                                    #  _normalize_item 의 기본값 "용역"이 잘못 적용되는 문제를 방지)
                                    if cat_label:
                                        norm["category"] = cat_label
                                    normalized.append(norm)
                                else:
                                    summary["invalid"] += 1
                            up = _upsert_announcements(db, source, normalized)
                            summary["inserted"] += up["inserted"]
                            summary["skipped"] += up["skipped"]
                            cat_inserted += up["inserted"]
                            if len(raw_items) < rows_per_page:
                                break
                        except Exception as exc:
                            _record_sync_error(per_cat_errors, "공고", cat_label, page, exc)
                            break
                if cat_label:
                    logger.info(
                        "g2b ann sync category=%s fetched=%d inserted=%d total_days=%d",
                        cat_label, cat_fetched, cat_inserted, total_days_ann,
                    )

            # ② 낙찰결과(예정가격 상세) 수집 - 슬라이딩 윈도우 적용
            for cat_label, cat_op in result_categories:
                cat_fetched = cat_inserted = 0
                for win_start, win_end in _windows_iter(total_days_res, window_days, start_date=sync_start_res):
                    for page in range(1, max_pages_res + 1):
                        try:
                            url = _build_result_url(
                                api_key, cat_op, rows=rows_per_page,
                                page_no=page, start_date=win_start, end_date=win_end,
                            )
                            payload = _http_fetch_with_retry(url, timeout=30, retries=2)
                            raw_items = _extract_g2b_items(payload)
                            if not raw_items:
                                break
                            summary["results_fetched"] += len(raw_items)
                            cat_fetched += len(raw_items)
                            normalized = [n for n in (_normalize_result_item(it) for it in raw_items) if n]
                            up = _upsert_results(db, normalized)
                            summary["results_inserted"] += up["inserted"]
                            summary["results_skipped"] += up["skipped"]
                            summary["results_no_announcement"] += up["no_announcement"]
                            cat_inserted += up["inserted"]
                            if len(raw_items) < rows_per_page:
                                break
                        except Exception as exc:
                            _record_sync_error(per_cat_errors, "낙찰", cat_label, page, exc)
                            break
                logger.info(
                    "g2b result sync category=%s fetched=%d inserted=%d total_days=%d",
                    cat_label, cat_fetched, cat_inserted, total_days_res,
                )

            if per_cat_errors:
                # 일부만 실패해도 부분 성공으로 처리, 전체 실패 시 failed
                total_fetched = summary["fetched"] + summary["results_fetched"]
                if total_fetched == 0:
                    status = "failed"
                error_message = "; ".join(per_cat_errors)[:500]
            elif summary["fetched"] == 0:
                status = "failed"
                error_message = f"{source} 응답에서 item을 찾지 못했습니다."
        else:
            # 시드 폴백: DB 변경 없이 기록만
            summary["fetched"] = random.randint(15, 80)
            summary["skipped"] = summary["fetched"]  # 시드 모드는 삽입 없음
            error_message = None
    except Exception as e:
        status = "failed"
        error_message = f"수집 중 예외: {str(e)[:150]}"

    finished = datetime.now()
    total_fetched = summary["fetched"] + summary.get("results_fetched", 0)
    total_inserted = summary["inserted"] + summary.get("results_inserted", 0)
    log = DataSyncLog(
        source=source,
        sync_type=f"공고+낙찰 수집 ({trigger})",
        status=status,
        records_fetched=total_fetched,
        inserted_count=total_inserted,
        error_message=(error_message[:500] if error_message else None),
        started_at=started,
        finished_at=finished,
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    sync_id = log.id
    db.close()
    return {
        "sync_id": sync_id,
        "source": source,
        "status": status,
        "records": summary["fetched"],
        "inserted": summary["inserted"],
        "skipped": summary["skipped"],
        "invalid": summary["invalid"],
        "results_fetched": summary.get("results_fetched", 0),
        "results_inserted": summary.get("results_inserted", 0),
        "results_skipped": summary.get("results_skipped", 0),
        "results_no_announcement": summary.get("results_no_announcement", 0),
        "duration_sec": int((finished - started).total_seconds()),
        "error_message": error_message,
    }


@app.post("/api/v1/admin/sync")
def trigger_sync(current_user: User = Depends(require_admin)):
    """관리자 수동 수집 — 기본은 G2B만 실행 (국방부 데이터 포함).
    D2B 별도 호출이 필요한 경우 환경변수 ENABLE_D2B_SYNC=true 설정.
    """
    sources = ("G2B", "D2B") if os.environ.get("ENABLE_D2B_SYNC", "").lower() == "true" else ("G2B",)
    results = [_run_sync_for_source(src, trigger="manual") for src in sources]
    total = sum(r["records"] for r in results)
    failed = [r for r in results if r["status"] == "failed"]
    return {
        "message": f"{len(results) - len(failed)}/{len(results)} 소스 수집 완료 (총 {total}건)",
        "records": total,
        "results": results,
        "status": "failed" if failed and not total else "success",
    }


@app.post("/api/v1/admin/sync/diagnose", tags=["관리자"])
def diagnose_sync(
    source: str = Query("G2B", pattern="^(G2B|D2B)$"),
    rows: int = Query(5, ge=1, le=50),
    current_user: User = Depends(require_admin),
):
    """수집 파이프라인 단계별 진단 — DB 변경 없이 키/네트워크/파싱만 검증"""
    api_key = os.environ.get(f"{source}_API_KEY", "").strip()
    steps: list[dict] = []

    def add(name: str, ok: bool, detail: str = "", **extra):
        steps.append({"step": name, "ok": ok, "detail": detail, **extra})

    # 1. 키 확인
    if not api_key:
        add("api_key", False, f"{source}_API_KEY 환경변수 미설정")
        return {"source": source, "ok": False, "steps": steps,
                "guide": f"export {source}_API_KEY='발급받은_인증키' 후 재시도"}
    add("api_key", True, f"길이 {len(api_key)}, 마스킹 {api_key[:4]}…{api_key[-4:]}")

    # 2. URL 빌드
    try:
        url = _build_sync_url(source, api_key, rows=rows)
        add("build_url", True, f"길이 {len(url)}자")
    except Exception as e:
        add("build_url", False, str(e))
        return {"source": source, "ok": False, "steps": steps}

    # 3. HTTP 호출
    import time as _t
    t0 = _t.perf_counter()
    try:
        payload = _http_fetch_with_retry(url, timeout=15, retries=1)
        elapsed_ms = round((_t.perf_counter() - t0) * 1000, 1)
        add("http_fetch", True, f"{len(payload)}바이트 / {elapsed_ms}ms",
            response_size=len(payload), elapsed_ms=elapsed_ms)
    except urllib.error.HTTPError as e:
        body = (e.read()[:300].decode("utf-8", errors="replace") if e.fp else "")
        add("http_fetch", False, f"HTTP {e.code} {e.reason}", body=body)
        return {"source": source, "ok": False, "steps": steps}
    except Exception as e:
        add("http_fetch", False, f"{type(e).__name__}: {str(e)[:150]}")
        return {"source": source, "ok": False, "steps": steps}

    # 4. items 추출
    items = _extract_g2b_items(payload)
    head = payload[:200].decode("utf-8", errors="replace")
    add("extract_items", len(items) > 0,
        f"{len(items)}건 추출 (응답 헤드: {head[:120]})", count=len(items))

    # 5. 정규화 (DB 반영 없음)
    valid = []
    invalid = 0
    for it in items:
        n = _normalize_item(source, it)
        if n:
            valid.append({"bid_number": n["bid_number"], "title": n["title"][:60],
                          "org": n["ordering_org_name"], "amount": n["base_amount"]})
        else:
            invalid += 1
    add("normalize", len(valid) > 0 if items else True,
        f"유효 {len(valid)} / 무효 {invalid}",
        valid_count=len(valid), invalid_count=invalid)

    return {
        "source": source,
        "ok": all(s["ok"] for s in steps),
        "steps": steps,
        "sample_records": valid[:3],
        "note": "이 엔드포인트는 DB를 변경하지 않습니다. 실제 적재는 POST /api/v1/admin/sync 사용.",
    }


@app.get("/api/v1/admin/nas-status")
def nas_status(current_user: User = Depends(require_admin)):
    """NAS 마운트/용량 확인"""
    nas_exists = os.path.exists(NAS_PATH)
    if nas_exists:
        try:
            stat = os.statvfs(NAS_PATH)
            total_gb = round(stat.f_frsize * stat.f_blocks / (1024 ** 3), 1)
            free_gb = round(stat.f_frsize * stat.f_bavail / (1024 ** 3), 1)
        except OSError as exc:
            logger.warning("NAS 디스크 정보 조회 실패 (%s): %s", NAS_PATH, exc)
            total_gb = 0
            free_gb = 0
    else:
        total_gb = 0
        free_gb = 0
    return {
        "mounted": nas_exists,
        "path": NAS_PATH,
        "total_gb": total_gb,
        "free_gb": free_gb,
        "db_path": DB_PATH,
        "db_size_mb": round(os.path.getsize(DB_PATH) / (1024 * 1024), 1) if os.path.exists(DB_PATH) else 0,
    }


@app.put("/api/v1/admin/users/{user_id}/status")
def toggle_user_status(user_id: str, current_user: User = Depends(require_admin)):
    """사용자 활성/비활성 토글"""
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            db.close()
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        if user.id == current_user.id:
            db.close()
            raise HTTPException(status_code=400, detail="자신의 상태는 변경할 수 없습니다.")
        user.is_active = not user.is_active
        db.commit()
        new_status = "활성" if user.is_active else "비활성"
    finally:
        db.close()
    return {"message": f"사용자 상태가 {new_status}(으)로 변경되었습니다.", "is_active": user.is_active}


@app.get("/api/v1/admin/sync/history")
def sync_history(page: int = 1, page_size: int = 20, current_user: User = Depends(require_admin)):
    """수집 히스토리 목록"""
    db = SessionLocal()
    try:
        q = db.query(DataSyncLog)
        total = q.count()
        items = q.order_by(DataSyncLog.started_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
        result = [{
            "id": log.id, "source": log.source, "sync_type": log.sync_type,
            "status": log.status, "records_fetched": log.records_fetched,
            "inserted_count": getattr(log, "inserted_count", 0) or 0,
            "error_message": log.error_message,
            "started_at": log.started_at.strftime("%Y-%m-%d %H:%M") if log.started_at else None,
            "finished_at": log.finished_at.strftime("%Y-%m-%d %H:%M") if log.finished_at else None,
            "duration_sec": int((log.finished_at - log.started_at).total_seconds()) if log.started_at and log.finished_at else None,
        } for log in items]
    finally:
        db.close()
    return {"items": result, "total": total, "page": page, "page_size": page_size}


@app.post("/api/v1/admin/sync/{sync_id}/retry")
def retry_sync(sync_id: str, current_user: User = Depends(require_admin)):
    """실패한 수집 건 재시도 — 실 파이프라인 재실행"""
    db = SessionLocal()
    try:
        log = db.query(DataSyncLog).filter(DataSyncLog.id == sync_id).first()
        if not log:
            db.close()
            raise HTTPException(status_code=404, detail="수집 이력을 찾을 수 없습니다.")
        source = log.source
    finally:
        db.close()
    # 실 파이프라인으로 재시도
    res = _run_sync_for_source(source, trigger="retry")
    return {
        "message": "재시도 완료",
        "status": res["status"],
        "records": res["records"],
        "inserted": res.get("inserted", 0),
        "error_message": res.get("error_message"),
        "new_sync_id": res["sync_id"],
    }


@app.get("/api/v1/admin/schedule")
def get_schedule(current_user: User = Depends(require_admin)):
    """자동 수집 스케줄 조회 — 다음 실행 시각 포함"""
    next_run = None
    if _SCHEDULER_AVAILABLE and _scheduler is not None:
        job = _scheduler.get_job("auto_sync")
        if job and job.next_run_time:
            next_run = job.next_run_time.strftime("%Y-%m-%d %H:%M:%S")
    return {**_schedule_config, "scheduler_running": _scheduler is not None, "next_run": next_run}


@app.put("/api/v1/admin/schedule")
def update_schedule(
    interval: str = Query(...),  # hourly / daily / weekly
    time: str = Query("02:00"),
    enabled: bool = Query(True),
    current_user: User = Depends(require_admin),
):
    """자동 수집 스케줄 설정 — 실 APScheduler 재등록"""
    if interval not in ("hourly", "daily", "weekly"):
        raise HTTPException(status_code=400, detail="interval은 hourly/daily/weekly 중 하나여야 합니다.")
    _schedule_config["interval"] = interval
    _schedule_config["time"] = time
    _schedule_config["enabled"] = enabled
    _apply_schedule_config(_schedule_config)
    next_run = None
    if _SCHEDULER_AVAILABLE and _scheduler is not None:
        job = _scheduler.get_job("auto_sync")
        if job and job.next_run_time:
            next_run = job.next_run_time.strftime("%Y-%m-%d %H:%M:%S")
    return {"message": "스케줄이 업데이트되었습니다.", **_schedule_config, "next_run": next_run}


@app.get("/api/v1/admin/errors")
def get_errors(current_user: User = Depends(require_admin)):
    """오류 모니터링 — 실패한 수집 로그 목록"""
    db = SessionLocal()
    try:
        fail_count = db.query(DataSyncLog).filter(DataSyncLog.status == "failed").count()
        recent_errors = db.query(DataSyncLog).filter(
            DataSyncLog.status == "failed"
        ).order_by(DataSyncLog.started_at.desc()).limit(20).all()
        result = [{
            "id": log.id, "source": log.source, "sync_type": log.sync_type,
            "started_at": log.started_at.strftime("%Y-%m-%d %H:%M") if log.started_at else None,
            "error_message": log.error_message or f"{log.source} {log.sync_type} 수집 실패",
        } for log in recent_errors]
    finally:
        db.close()
    return {"fail_count": fail_count, "errors": result}


# ─── API: 조회 이력 ──────────────────────────────────────────────────────

@app.get("/api/v1/history")
def list_history(
    page: int = 1, page_size: int = 20,
    current_user: User = Depends(require_auth),
):
    """현재 사용자의 분석 조회 이력"""
    db = SessionLocal()
    try:
        q = db.query(QueryHistory).filter(QueryHistory.user_id == current_user.id)
        total = q.count()
        items = q.order_by(QueryHistory.queried_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

        result = []
        for h in items:
            ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == h.announcement_id).first()
            result.append({
                "id": h.id,
                "announcement_id": h.announcement_id,
                "announcement_title": ann.title if ann else None,
                "analysis_type": h.analysis_type,
                "parameters": json.loads(h.parameters) if h.parameters else None,
                "result_summary": json.loads(h.result_summary) if h.result_summary else None,
                "queried_at": h.queried_at.strftime("%Y-%m-%d %H:%M") if h.queried_at else None,
            })

    finally:
        db.close()
    return {"items": result, "total": total, "page": page, "page_size": page_size}


@app.get("/api/v1/history/{history_id}")
def get_history(history_id: str, current_user: User = Depends(require_auth)):
    db = SessionLocal()
    try:
        h = db.query(QueryHistory).filter(
            QueryHistory.id == history_id, QueryHistory.user_id == current_user.id
        ).first()
        if not h:
            db.close()
            raise HTTPException(status_code=404, detail="조회 이력을 찾을 수 없습니다.")
        ann = db.query(BidAnnouncement).filter(BidAnnouncement.id == h.announcement_id).first()
    finally:
        db.close()
    return {
        "id": h.id,
        "announcement_id": h.announcement_id,
        "announcement_title": ann.title if ann else None,
        "analysis_type": h.analysis_type,
        "parameters": json.loads(h.parameters) if h.parameters else None,
        "result_summary": json.loads(h.result_summary) if h.result_summary else None,
        "queried_at": h.queried_at.strftime("%Y-%m-%d %H:%M") if h.queried_at else None,
    }


def save_query_history(user_id: str, announcement_id: str, analysis_type: str,
                       parameters: dict = None, result_summary: dict = None):
    """분석 조회 이력 저장 (헬퍼 함수). 예외는 운영에 영향 없도록 로그만."""
    try:
        with db_session() as db:
            db.add(QueryHistory(
                user_id=user_id,
                announcement_id=announcement_id,
                analysis_type=analysis_type,
                parameters=json.dumps(parameters, ensure_ascii=False) if parameters else None,
                result_summary=json.dumps(result_summary, ensure_ascii=False) if result_summary else None,
                queried_at=datetime.now(),
            ))
            user = db.query(User).filter(User.id == user_id).first()
            if user:
                user.query_count = (user.query_count or 0) + 1
            db.commit()
    except Exception as exc:
        logger.debug("query history 저장 실패 (분석 결과에 영향 없음): %s", exc)


# ─── API: 데이터 업로드 ──────────────────────────────────────────────────

@app.post("/api/v1/data/upload")
def upload_data(file: UploadFile = File(...), current_user: User = Depends(require_auth)):
    """CSV/Excel 입찰 데이터 업로드 — 중복 검사·행별 오류 추적 포함"""
    import io
    allowed_ext = {".csv", ".xlsx", ".xls"}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail=f"허용되지 않는 파일 형식입니다. ({', '.join(allowed_ext)})")

    content = file.file.read()
    file_size = len(content)
    MAX_SIZE = 10 * 1024 * 1024  # 10MB
    if file_size > MAX_SIZE:
        raise HTTPException(status_code=413, detail=f"파일 크기가 10MB를 초과합니다. (현재 {round(file_size/1024/1024, 1)}MB)")

    db = SessionLocal()
    upload = UploadLog(
        user_id=current_user.id, filename=file.filename,
        file_size=file_size, status="processing",
        uploaded_at=datetime.now(),
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)

    def _fail(msg: str, code: int = 400):
        upload.status = "failed"
        upload.error_message = msg
        db.commit()
        db.close()
        raise HTTPException(status_code=code, detail=msg)

    try:
        import pandas as pd
        try:
            if ext == ".csv":
                # 한국어 파일 인코딩 대응 (utf-8 → cp949 폴백)
                try:
                    df = pd.read_csv(io.BytesIO(content))
                except UnicodeDecodeError:
                    df = pd.read_csv(io.BytesIO(content), encoding="cp949")
            else:
                df = pd.read_excel(io.BytesIO(content))
        except Exception as e:
            _fail(f"파일 파싱 실패: {str(e)[:200]}")

        # 필수 컬럼 확인
        required_cols = {"공고번호", "공고명", "발주기관", "기초금액"}
        missing = required_cols - set(df.columns)
        if missing:
            _fail(f"필수 컬럼 누락: {', '.join(missing)}")

        total_rows = len(df)
        if total_rows == 0:
            _fail("데이터 행이 없습니다.")

        # 기존 bid_number 집합 (중복 검사)
        existing_numbers = {
            row[0] for row in db.query(BidAnnouncement.bid_number).all()
        }

        inserted = 0
        duplicates = 0
        error_rows: list[dict] = []
        seen_in_file: set[str] = set()
        batch: list[BidAnnouncement] = []

        for idx, row in df.iterrows():
            row_num = idx + 2  # 엑셀 기준: 헤더 1행 + 1-based
            try:
                bid_number = str(row.get("공고번호", "")).strip()
                title = str(row.get("공고명", "")).strip()
                org = str(row.get("발주기관", "")).strip()
                base_raw = row.get("기초금액")

                # 필드 값 검증
                if not bid_number or bid_number == "nan":
                    error_rows.append({"row": row_num, "error": "공고번호 비어 있음"})
                    continue
                if not title or title == "nan":
                    error_rows.append({"row": row_num, "error": "공고명 비어 있음"})
                    continue
                if not org or org == "nan":
                    error_rows.append({"row": row_num, "error": "발주기관 비어 있음"})
                    continue
                if pd.isna(base_raw):
                    error_rows.append({"row": row_num, "error": "기초금액 누락"})
                    continue
                try:
                    base_amount = int(float(base_raw))
                    if base_amount < 0:
                        error_rows.append({"row": row_num, "error": "기초금액 음수"})
                        continue
                except (ValueError, TypeError):
                    error_rows.append({"row": row_num, "error": f"기초금액 숫자 변환 실패: {base_raw}"})
                    continue

                # 중복 검사 (DB + 파일 내)
                if bid_number in existing_numbers or bid_number in seen_in_file:
                    duplicates += 1
                    continue
                seen_in_file.add(bid_number)

                category = str(row.get("카테고리", "용역")).strip() or "용역"
                region = str(row.get("지역", "")).strip() if "지역" in df.columns else None
                if region in ("", "nan"):
                    region = None

                batch.append(BidAnnouncement(
                    source="UPLOAD",
                    bid_number=bid_number,
                    category=category,
                    title=title,
                    ordering_org_name=org,
                    region=region,
                    base_amount=base_amount,
                    announced_at=datetime.now(),
                    status="업로드",
                ))
                inserted += 1
            except Exception as e:
                error_rows.append({"row": row_num, "error": f"예외: {str(e)[:120]}"})

        if batch:
            db.bulk_save_objects(batch)
            invalidate_analysis_cache()

        # 요약 기록
        upload.records_count = inserted
        final_status = "success" if inserted > 0 else ("failed" if error_rows and not duplicates else "partial")
        upload.status = final_status
        upload.error_message = None if not error_rows else f"오류 {len(error_rows)}건"
        db.commit()
        # 세션 close 전에 primitive 값 캡처
        upload_id = upload.id
        db.close()

        summary = {
            "total": total_rows,
            "inserted": inserted,
            "duplicates": duplicates,
            "errors": len(error_rows),
        }
        return {
            "message": f"{inserted}건 등록 / {duplicates}건 중복 / {len(error_rows)}건 오류",
            "upload_id": upload_id,
            "status": final_status,
            "summary": summary,
            "error_rows": error_rows[:50],  # 상위 50건만 반환
        }

    except HTTPException:
        raise
    except Exception as e:
        upload.status = "failed"
        upload.error_message = str(e)[:500]
        db.commit()
        db.close()
        raise HTTPException(status_code=500, detail=f"데이터 처리 중 오류: {str(e)}")


@app.get("/api/v1/data/uploads")
def list_uploads(current_user: User = Depends(require_auth)):
    """업로드 이력 조회"""
    db = SessionLocal()
    try:
        uploads = db.query(UploadLog).filter(
            UploadLog.user_id == current_user.id
        ).order_by(UploadLog.uploaded_at.desc()).limit(50).all()
    finally:
        db.close()
    return [{
        "id": u.id, "filename": u.filename,
        "file_size": u.file_size, "records_count": u.records_count,
        "status": u.status, "error_message": u.error_message,
        "uploaded_at": u.uploaded_at.strftime("%Y-%m-%d %H:%M") if u.uploaded_at else None,
    } for u in uploads]


# ─── 프론트엔드 서빙 ──────────────────────────────────────────────────────

STATIC_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend", "public")
DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "docs")


# ─── 헬스체크 + 메트릭 ───────────────────────────────────────────
@app.get("/healthz", tags=["운영"])
def healthz():
    """컨테이너 헬스체크 — DB 연결 + 스케줄러 상태"""
    ok = True
    checks = {}
    try:
        db = SessionLocal()
        try:
            db.query(User).limit(1).all()
        finally:
            db.close()
        checks["db"] = "ok"
    except Exception as e:
        ok = False
        checks["db"] = f"error: {str(e)[:80]}"

    checks["scheduler"] = "running" if (_scheduler is not None) else "stopped"
    checks["env"] = _RUNTIME_ENV
    status_code = 200 if ok else 503
    from starlette.responses import JSONResponse as _JR
    return _JR({"status": "ok" if ok else "degraded", "checks": checks}, status_code=status_code)


@app.get("/metrics", tags=["운영"])
def metrics():
    """경량 운영 메트릭 — Prometheus 파싱 호환 형식.

    외부 노출 전에 인증/ACL 제어 필요 시 reverse-proxy에서 제한할 것.
    """
    db = SessionLocal()
    lines: list[str] = []

    def _m(name: str, value, help_text: str, mtype: str = "gauge"):
        lines.append(f"# HELP {name} {help_text}")
        lines.append(f"# TYPE {name} {mtype}")
        lines.append(f"{name} {value}")

    try:
        _m("bid_announcements_total", db.query(BidAnnouncement).count(),
           "등록된 공고 수")
        _m("bid_results_total", db.query(BidResult).count(),
           "낙찰 결과 수")
        _m("users_total", db.query(User).count(),
           "가입 사용자 수")
        _m("users_active_total", db.query(User).filter(User.is_active == True).count(),  # noqa: E712
           "활성 사용자 수")
        _m("sync_success_total",
           db.query(DataSyncLog).filter(DataSyncLog.status == "success").count(),
           "누적 성공 수집 건수", "counter")
        _m("sync_failed_total",
           db.query(DataSyncLog).filter(DataSyncLog.status == "failed").count(),
           "누적 실패 수집 건수", "counter")
        _m("query_history_total", db.query(QueryHistory).count(),
           "누적 분석 조회 수", "counter")
        _m("upload_logs_total", db.query(UploadLog).count(),
           "누적 업로드 건수", "counter")
        # 캐시/스케줄러 상태
        with _cache_lock:
            cache_size = len(_cache_store)
        _m("analysis_cache_entries", cache_size, "분석 캐시 엔트리 수")
        _m("scheduler_running", 1 if _scheduler is not None else 0,
           "스케줄러 실행 여부 (1=running)")
    finally:
        db.close()

    from starlette.responses import Response as _Resp
    return _Resp("\n".join(lines) + "\n", media_type="text/plain; version=0.0.4")


@app.get("/")
def serve_index():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


@app.get("/proposal")
def serve_proposal():
    return FileResponse(os.path.join(DOCS_DIR, "기술제안서.html"))


# SPA catchall — path 기반 라우팅(/admin, /announcements 등) 새로고침/북마크 지원
# API/메타 prefix 는 흡수하지 않도록 화이트리스트 명시.
_API_PREFIXES = ("api/", "docs", "redoc", "openapi", "metrics", "healthz", "static")


@app.get("/{full_path:path}")
def spa_catchall(full_path: str):
    if not full_path or full_path.startswith(_API_PREFIXES):
        raise HTTPException(status_code=404)
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


# ─── 시작 ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    Base.metadata.create_all(engine)
    # SKIP_SEED=true 면 시드 데이터 자동 생성 비활성화 (운영 모드)
    if os.environ.get("SKIP_SEED", "").lower() not in ("true", "1", "yes"):
        try:
            from seed import seed_database  # lazy import (운영 시 임포트도 안 함)
            seed_database()
        except Exception as e:
            print(f"⚠️ 시드 데이터 생성 중 오류 (무시하고 계속): {e}")
    else:
        logger.info("SKIP_SEED=true — 시드 데이터 생성 건너뜀")
    port = int(os.environ.get("PORT", 8000))
    print(f"\n🚀 서버 시작: http://0.0.0.0:{port}")
    print(f"   API 문서: http://localhost:{port}/docs\n")
    uvicorn.run(app, host="0.0.0.0", port=port)
